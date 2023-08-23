import fractions
from functools import partial
from itertools import chain
import re
from xmlrpc.client import ResponseError
from django.db import transaction
import json
import csv
from tokenize import String
from django.core import mail
from django.http import HttpResponse, response
from django.shortcuts import render
from branch_user.views import Create_Update_Weekly_Sales
from master.thread import SaleData_create, SaleData_update
from rest_framework import generics
from base.serializers import *
from base.models import *
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import (AllowAny,IsAuthenticated)
from django.conf import settings
import os
import datetime
from rest_framework.generics import RetrieveDestroyAPIView, ListCreateAPIView
from django.contrib.auth.hashers import make_password ,check_password
from django.db.models import Q
from base.tests import generate_access_token, generate_refresh_token
import jwt
import logging
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from django.template.loader import get_template
from django.template import Context
from master.utils import count_invalid_sale_date_wise, count_success_sale_date_wise, create_update_weekly, weekly_generate_on_weekend
from master.thread import SaleData_create, SaleData_create_and_update_1, SaleData_create_and_update_2, SaleData_create_and_update_3, SaleData_create_and_update_4, SaleData_update,first_sale_create_update

from wd import utils
from wd.helper import week_sele_before_date
logger = logging.getLogger(__name__)
from secondary_sales.settings import *
import openpyxl
from master.serializers import *
from master.models import *
from django.contrib.auth.hashers import make_password ,check_password
import requests
from dateutil.parser import parse
from rest_framework import viewsets
from rest_framework.response import Response
import csv
import pandas as pd
from .email_sub import surya_success_sub,surya_failed_sub,sfa_success_sub,sfa_failed_sub,success_email_list,failed_list,cc_list,test_sch
import numpy as np
from secondary_sales.settings import my_eng, web_base_url


def unit_val_calc(sku, wd_town_id, total):

    val = WdSkuCatagory.objects.filter(Q(active_flag = 'Y')|Q(active_flag = 'y'),sku_code = sku, wd_town_id = wd_town_id).last()
    if val:
        print(sku,"======if======",wd_town_id)
        calc_val = float(val.last_price) * total
        return calc_val, val.unit_price, val.cnf_id, val.company
    else:
        print(sku,"======ielf======",wd_town_id)
        val = WdSkuCatagory.objects.filter(sku_id = sku, wd_town_id = wd_town_id).last()
        calc_val = val.last_price * total
        return calc_val, val.unit_price, val.cnf_id, val.company


def emailsend_zoro_or_running(context):
    subject = sfa_failed_sub
    recipient_list = failed_list
    cc_email = cc_list
    # recipient_list = ["VIPULSHARMA-gpi@modi-ent.com","ankit.rakwal@esobene.com"]
    html_message = str(context)
    reset_email = EmailMessage(
                        subject = subject,
                        body = html_message,
                        from_email = settings.EMAIL_HOST_USER,
                        to = recipient_list,
                        cc = cc_email,
                        reply_to = cc_email,
                        )
    reset_email.content_subtype = "html"
    reset_email.send(fail_silently=True)



class Candy_Gms(APIView):
    def post(self, request):
        file_obj = request.FILES['file']
        wb_obj = openpyxl.load_workbook(file_obj)
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row
        dicts = []
        id_list= []
        try:
            
            for i in range(2, m_row + 1):
                empdata={}
                empdata['sku_id'] = str(sheet_obj.cell(row = i, column = 1).value)
                empdata['sku_code'] = str(sheet_obj.cell(row = i, column = 2).value)
                empdata['sku_name'] = str(sheet_obj.cell(row = i, column = 3).value)
                empdata['sku_short_name'] = str(sheet_obj.cell(row = i, column = 4).value)
                empdata['jar_size_gms'] = str(sheet_obj.cell(row = i, column = 5).value)
                serializer = Candy_JAR_GMSSerializer(data=empdata)
                if serializer.is_valid():
                    serializer.save()
                else:
                    return Response(serializer.errors, status=status.HTTP_200_OK)

            message={"ok":"ok"}
            return Response(message, status=status.HTTP_200_OK)
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong','error':error}
            return Response(context, status=status.HTTP_200_OK)

            
        


class WDmaster_upload(APIView):
    # permission_classes = (IsAuthenticated,)
    # serializer_class = WDmasterSerializers


    def post(self, request):
        file_obj = request.FILES['file']
        wb_obj = openpyxl.load_workbook(file_obj)
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row
        dicts = []
        id_list= []
        try:
            
            for i in range(2, m_row + 1):
                empdata={}
                # print(sheet_obj.cell(row = i, column = 16).value)

                empdata['wd_ids']=str(sheet_obj.cell(row = i, column = 1).value)
                empdata['wd_name']=str(sheet_obj.cell(row = i, column = 2).value)
                # if empdata['wd_name'] == "SURYA":
                #     id_list.append(empdata['wd_ids'])
            # uupdate = WDmaster.objects.filter(wd_ids__in= id_list)
            
            # update = WDmaster.objects.filter(wd_ids__in= id_list).update(wd_type = "SURYA")
            # print(update,'success',uupdate)
                # id_list = 

                empdata['wd_address1']=sheet_obj.cell(row = i, column = 3).value
                empdata['wd_address2']=sheet_obj.cell(row = i, column = 4).value
                empdata['wd_address3']=sheet_obj.cell(row = i, column = 3).value
                empdata['wd_address4']=sheet_obj.cell(row = i, column = 4).value
                empdata['wd_city']=sheet_obj.cell(row = i, column = 6).value
                empdata['wd_postal_code']=sheet_obj.cell(row = i, column = 5).value
                empdata['wd_state']=sheet_obj.cell(row = i, column = 7).value
                empdata['wd_country']=sheet_obj.cell(row = i, column = 8).value
                print("====4===",empdata)
                serializer = WDmasterSerializers(data=empdata)
                if serializer.is_valid():
                    serializer.save()
                else:
                    return Response(serializer.errors, status=status.HTTP_200_OK)

            message={"ok":"ok"}
            return Response(message)
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong','error':error}
            return Response(context, status=status.HTTP_200_OK)

class Sales_Hierarchy_Master_upload(APIView):
    # permission_classes = (IsAuthenticated,)
    serializer_class = Saless_Hierarchy_MasterSerializers


    def post(self, request):
        file_obj = request.FILES['file']
        wb_obj = openpyxl.load_workbook(file_obj)
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row
        dicts = []
        try:
            for i in range(2, m_row + 1):
                empdata={}
                empdata['wd_id']=(str(sheet_obj.cell(row = i, column = 1).value)).split('.')[0]
                empdata['wd_town_id']=sheet_obj.cell(row = i, column = 2).value
                empdata['region_code']=sheet_obj.cell(row = i, column = 3).value
                empdata['region']=sheet_obj.cell(row = i, column = 4).value
                empdata['town']=sheet_obj.cell(row = i, column = 5).value
                empdata['town_code']=sheet_obj.cell(row = i, column = 6).value
                empdata['sku_category_code']=sheet_obj.cell(row = i, column = 7).value
                empdata['wd_town_code']=sheet_obj.cell(row = i, column = 8).value
                print("====4===",empdata)
                serializer = Saless_Hierarchy_MasterSerializers(data=empdata)
                if serializer.is_valid():
                    serializer.save()
                else:
                    return Response(serializer.errors, status=status.HTTP_200_OK)

            message={"ok":"ok"}
            return Response(message, status=status.HTTP_200_OK)
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong','error':error}
            return Response(context, status=status.HTTP_200_OK)


class SKU_Master_Productupload(APIView):
    # permission_classes = (IsAuthenticated,)
    serializer_class = SKU_Master_ProductSerializers


    def post(self, request):
        file_obj = request.FILES['file']
        wb_obj = openpyxl.load_workbook(file_obj)
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row
        dicts = []
        try:
            for i in range(2, m_row + 1):
                empdata={}
                empdata['sku_id']=sheet_obj.cell(row = i, column = 1).value
                empdata['sku_code']=sheet_obj.cell(row = i, column = 2).value
                empdata['sku_short_name']=sheet_obj.cell(row = i, column = 3).value
                empdata['sku_name']=sheet_obj.cell(row = i, column = 4).value
                empdata['active_flag']=sheet_obj.cell(row = i, column = 5).value
                empdata['primary_uom_code']=sheet_obj.cell(row = i, column = 6).value
                empdata['effective_from']=str(sheet_obj.cell(row = i, column = 7).value)
                empdata['category_code']=sheet_obj.cell(row = i, column = 8).value
                empdata['category_name']=sheet_obj.cell(row = i, column = 9).value
                empdata['created_by']=sheet_obj.cell(row = i, column = 10).value
                empdata['company'] = sheet_obj.cell(row = i, column = 10).value
                print("====4===",empdata)
                serializer = SKU_Master_ProductSerializers(data=empdata)
                if serializer.is_valid():
                    serializer.save()
                else:
                    return Response(serializer.errors, status=status.HTTP_200_OK)

            message={"ok":"ok"}
            return Response(message, status=status.HTTP_200_OK)
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong','error':error}
            return Response(context, status=status.HTTP_200_OK)



class User_upload(APIView):
    # permission_classes = (IsAuthenticated,)
    serializer_class = User_uoloadSerializers


    def post(self, request):
        user_type=request.data.get('user_type')
        file_obj = request.FILES['file']
        wb_obj = openpyxl.load_workbook(file_obj)
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row
        dicts = []
        try:
            for i in range(2, m_row + 1):
                if (sheet_obj.cell(row = i, column = 4).value).upper() == 'W':
                    empdata={}
                    empdata['user_type']='WD'
                    empdata['username']=i
                    emails=str(i)+'ajayk@triazinesoft.com'
                    empdata['email']=str(emails)
                    empdata['password']=make_password('wpassword@123')
                    empdata['user_id']=str(sheet_obj.cell(row = i, column = 1).value).split('.')[0]
                    empdata['first_name']=sheet_obj.cell(row = i, column = 2).value
                    empdata['locationcode']=sheet_obj.cell(row = i, column = 3).value
                    print("===",empdata)
                    serializer = User_uoloadSerializers(data=empdata)
                    if serializer.is_valid():
                        print("=====b====")
                        serializer.save()
                    else:
                        return Response(serializer.errors, status=status.HTTP_200_OK)
                if (sheet_obj.cell(row = i, column = 4).value).upper() == 'B':
                    empdata={}
                    empdata['user_type']='BRANCH USER'
                    empdata['username']=i
                    emails=str(i)+'ajayk@triazinesoft.com'
                    empdata['email']=str(emails)
                    empdata['password']=make_password('bpassword@123')
                    empdata['user_id']=str(sheet_obj.cell(row = i, column = 1).value).split('.')[0]
                    empdata['first_name']=sheet_obj.cell(row = i, column = 2).value
                    empdata['locationcode']=sheet_obj.cell(row = i, column = 3).value
                    print("===",empdata)
                    serializer = User_uoloadSerializers(data=empdata)
                    if serializer.is_valid():
                        print("=====m======")
                        serializer.save()
                    else:
                        return Response(serializer.errors, status=status.HTTP_200_OK)

            message={"ok":"ok"}
            return Response(message, status=status.HTTP_200_OK)
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong','error':error}
            return Response(context, status=status.HTTP_200_OK)


class WdSkuCatagoryupload(APIView):
    # permission_classes = (IsAuthenticated,)
    serializer_class = WdSkuCatagorySerializers


    def post(self, request):
        file_obj = request.FILES['file']
        print("=======",type(file_obj))
        wb_obj = openpyxl.load_workbook(file_obj)
        print("====3=====",wb_obj)
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row
        dicts = []
        try:
            for i in range(2, m_row + 1):
                empdata={}
                
                empdata['active_flag']=sheet_obj.cell(row = i, column = 5).value
                empdata['sku_code']=sheet_obj.cell(row = i, column = 1).value
                empdata['wd_town_id']=str(sheet_obj.cell(row = i, column = 2).value)
                empdata['sku_id']=sheet_obj.cell(row = i, column = 3).value
                empdata['cnf_id']=sheet_obj.cell(row = i, column = 4).value
                
                empdata['last_price']=sheet_obj.cell(row = i, column = 6).value
                
                serializer = WdSkuCatagorySerializers(data=empdata)
                if serializer.is_valid():
                    serializer.save()
                else:
                    return Response(serializer.errors, status=status.HTTP_200_OK)

            message={"ok":"ok"}
            return Response(message, status=status.HTTP_200_OK)
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong','error':error}
            return Response(context, status=status.HTTP_200_OK)



class BranchMaster_upload(APIView):
    permission_classes = (IsAuthenticated,)
    serializer_class = BranchMasterSerializers


    def post(self, request):
        file_obj = request.FILES['file']
        wb_obj = openpyxl.load_workbook(file_obj)
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row
        dicts = []
        try:
            for i in range(2, m_row + 1):
                empdata={}
                empdata['branch_ids']=sheet_obj.cell(row = i, column = 1).value
                empdata['branch_name']=str(sheet_obj.cell(row = i, column = 2).value)
                empdata['branch_address1']=sheet_obj.cell(row = i, column = 3).value
                empdata['branch_address2']=sheet_obj.cell(row = i, column = 4).value
                empdata['branch_address3']=sheet_obj.cell(row = i, column = 5).value
                empdata['branch_country']=sheet_obj.cell(row = i, column = 6).value
                empdata['branch_postal_code']=sheet_obj.cell(row = i, column = 7).value
                empdata['branch_code']=sheet_obj.cell(row = i, column = 8).value
                empdata['region']=sheet_obj.cell(row = i, column = 9).value
                print("====4===",empdata)
                serializer = BranchMasterSerializers(data=empdata)
                if serializer.is_valid():
                    serializer.save()
                else:
                    return Response(serializer.errors, status=status.HTTP_200_OK)

            message={"ok":"ok"}
            return Response(message, status=status.HTTP_200_OK)
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong','error':error}
            return Response(context, status=status.HTTP_200_OK)



class WD_sku_type_upload(APIView):
    

    def get(self, request):
        b = SKU_type.objects.all()
        p=''
        list1=[]
        presanr=[]
        not_presanr=[]
        "18-05-2022"
        wd_list = SalesData.objects.filter(sales_date_time__gte = "2022-05-18",created_by = "SURYA",sales_date_time__lte = "2022-05-19").values_list('wd_id', flat = True)
        print(wd_list,"======")
        WDmaster.objects.filter(wd_ids__in = wd_list).update(wd_type = "SURYA")
        
        ids = WDmaster.objects.filter(wd_type = "SFA").values_list('wd_ids', flat=True)
        User.objects.filter(user_id__in = ids).update(lock_unlock=False)
        # breakpoint()
        try:
            for i in b:
                a=i.wd_id
                if '-' in a:
                    x = a.split("-")
                    p=x[0]
                    wd_type=i.wd_type
                    k=WDmaster.objects.filter(wd_ids=p).exclude(wd_ids__isnull=True).update(wd_type = wd_type)

                    # print("------",i.wd_type)
                    # print("------",k)
                    # presanr.append(p)
                else:
                    p=a
                    wd_type=i.wd_type
                    k=WDmaster.objects.filter(wd_ids=p).exclude(wd_ids__isnull=True).update(wd_type = wd_type)
                    # print("---==---",k)
                    # presanr.append(p)
            # print("======",presanr)
                list1.append(p)
            # print("=====list1==",list1)
            k=WDmaster.objects.all().exclude(wd_ids__isnull=True).values_list('wd_ids',flat=True)
            for j in k:
                if j in list1:
                    presanr.append(j)
                else:
                    not_presanr.append(j)
            print("not_presanr",not_presanr)
            print("presanr",presanr)
            message={"ok":"ok"}
            return Response(message, status=status.HTTP_200_OK)
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong','error':error}
            return Response(context, status=status.HTTP_200_OK)





# Manualy freeze multiple sku===========================>
class Sku_Freeze(APIView):
    permission_classes = (IsAuthenticated,)
    serializer_class = SalesDataSerializer

    def post(self, request):
        try:
            for i in request.data:
                print(i['wd_id'],"======",i['sku_id'],"====",i['date'])
                Salesobj = SalesData.objects.filter(sku_id=i['sku_id'],wd_id=i['wd_id'],town_id=i['town_id'],brand_category=i['category'],created_date__date=i['date']).update(freez_status=True)
                print(Salesobj,"====>>")
            message={'status': True,'massage': 'Request successfull'}
            return Response(message, status=status.HTTP_200_OK)
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong','error':error}
            return Response(context, status=status.HTTP_200_OK)



class Schedule_Surya(generics.ListCreateAPIView):
    def get_queryset(self):
        # n= datetime.datetime.now().date()
        # start_date = n-datetime.timedelta(days=1)
        # print(str(to_date),"=====from_date=====")
        # url = "http://gpisfatest.winitsoftware.com/api/api/GPIInterface/GET?StartDate="+str(start_date)+"&EndDate="+str(start_date)+"9&Passcode=GP!W!N"
        # # url = "http://gpisfatest.winitsoftware.com/api/api/GPIInterface/GET?StartDate=2022-03-23&EndDate=2022-03-23&Passcode=GP!W!N"
        # url="http://gpisfatest.winitsoftware.com/api/api/GPIInterface/GET?StartDate="+str(from_date)+"&EndDate="+str(to_date)+"&Passcode=GP!W!N"
        # n_o_w = datetime.datetime.now()
        # time_rng = datetime.datetime(2009, 10, 5, 9, 10, 20)
        # time_rng_2 = datetime.datetime(2009, 10, 5, 9, 10, 30)
        try:
            # url = "https://ss.godfreyphillips.co/api_ss/api/GPIInterface?Passcode=GP!W!N"
            url = "https://ss.godfreyphillips.co/GPIInterFaceAPI_new/api/GPIInterface?Passcode=GP!W!N"
            
            surya_data = requests.get(url).json()
            print("surya_data",url,"======surya_data ======")
            errors_list = []
            error_obj_list = []
            valid_in_ss = []
            transaction_history = {}
            error_detail_dict = {}
            flag=0
            wd_list_4_week = []
            date_tuple_weekly = [8,15,22,1]
            
            wd_count = 0
            sku_count = 0
            town_count = 0
            n = datetime.datetime.now().date()-datetime.timedelta(days=1)
            # today = utils.now_date
            today = datetime.datetime.now()
            now_day=str(today.date())
            splited_today = now_day.split('-')[2]
            week_details = week_sele_before_date('BRANCH USER')
        
        
        # breakpoint()
        
            surya_api_sale_date=parse(surya_data.get('StartDate'))
            surya_datas=surya_data.get('Data')
            json_object = json.dumps(surya_datas, indent = 4)
            with open(settings.MEDIA_ROOT + 'surya_api_sales/'+"surya_sales_"+str(datetime.datetime.now().date()) + '.json','w')as outfile:
                outfile.write(json_object)
                
            today_date=datetime.datetime.now()
            status_data = Apistatus.objects.filter(running_date=today_date.date(),api="SURYA_API").last()
            running_state =  status_data.status if status_data and status_data.status == True else False
            print(status_data,'aaaasssss',running_state)
            print ((running_state == False),"--------55")
            # if   :
            #     print("aaasssdwwwwwwwwwwww")
                    
                
                
            if surya_datas :
                print("kamalkannnnnnt")
                response = HttpResponse(content_type='text/csv')
                # response['Content-Disposition'] = 'attachment; filename="sec_sales_data.csv"'
                df = pd.DataFrame.from_dict(list(surya_datas))
                df.to_csv(settings.MEDIA_ROOT + 'surya_api_sales_csv/'+"surya_sales_"+str(datetime.datetime.now().date()) + '.csv', index = False)

               
                    
            
            
            if surya_datas and (running_state == False):
               
                aaa=Apistatus.objects.create(running_date=today_date.date(),status=True,api="SURYA_API") 
                      
                
                for json_data in surya_datas:
                    wd = []
                    sku = []
                    town = []
                    
                 
                     
                    if json_data['sku_code'] and json_data['wd_code'] and json_data['town_id']:
                        sku_obj=SKU_Master_Product.objects.filter(sku_code=json_data['sku_code']).last()
                        json_data['wd_id'] = json_data['wd_code']
                        
                        user = User.objects.filter(user_id = json_data['wd_code'])
                        town_cde = str("0")+str(json_data['town_id'])# 0 concatenate for town issue================
                        town_cde2 = str(json_data['town_id']).strip('-')
                        wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(town_code__iexact=json_data['town_id'])| Q(town_code__iexact=str(json_data['town_id'])[1:])| Q(town_code__iexact='0'+str(json_data['town_id']))),wd_id=json_data.get('wd_id')).values_list('wd_town_id',flat=True)
                        wdobj = WdSkuCatagory.objects.filter(sku_code=json_data['sku_code'],wd_town_id__in=wd_town_list).last()
                        
                       
                        if not sku_obj:
                            sku.append(json_data['sku_code'])
                        if not user:
                            wd.append(json_data['wd_code'])
                        if not wdobj:
                            town.append(json_data['town_id'])
                        total_sale = (
                                float(json_data.get('local_sales_reatil') if json_data.get('local_sales_reatil', 0) else 0) 
                                + float(json_data.get('local_sales_dealer') if json_data.get('local_sales_dealer', 0) else 0)
                                + float(json_data.get('local_sales_modern_trade') if json_data.get('local_sales_modern_trade', 0) else 0)
                                + float(json_data.get('local_sales_hawker') if json_data.get('local_sales_hawker', 0) else 0)
                                
                                + float(json_data.get('outstation_sales_reatil') if json_data.get('outstation_sales_reatil', 0) else 0) 
                                + float(json_data.get('outstation_sales_dealer') if json_data.get('outstation_sales_dealer', 0) else 0)
                                + float(json_data.get('outstation_sales_modern_trade') if json_data.get('outstation_sales_modern_trade', 0) else 0)
                                + float(json_data.get('outstation_sales_hawker') if json_data.get('outstation_sales_hawker', 0) else 0)
                                
                                + float(json_data.get('other_sales_reatil') if json_data.get('other_sales_reatil', 0) else 0) 
                                + float(json_data.get('other_sales_dealer') if json_data.get('other_sales_dealer', 0) else 0)
                                + float(json_data.get('other_sales_hawker') if json_data.get('other_sales_hawker', 0) else 0)
                                + float(json_data.get('other_sales_modern_trade') if json_data.get('other_sales_modern_trade', 0) else 0)
                            )

                        if sku_obj and wdobj and user and (total_sale > 0):
                            valid_in_ss.append(json_data)
                        else:
                            json_data['dist_id']=json_data['wd_code']
                            json_data['prodcode']=json_data['sku_code']
                            json_data['town_code']=json_data['town_id']
                            json_data['sale_date']=json_data['sale_date_time']
                            json_data['region']=json_data['region']
                            json_data['local_retail']=json_data['local_sales_reatil']
                            json_data['local_dealer']=json_data['local_sales_dealer']
                            json_data['local_MT']=json_data['local_sales_modern_trade']
                            json_data['local_HA']=json_data['local_sales_hawker']
                            json_data['out_retail']=json_data['outstation_sales_reatil']
                            json_data['out_dealer']=json_data['outstation_sales_dealer']
                            json_data['out_MT']=json_data['outstation_sales_modern_trade']
                            json_data['out_HA']=json_data['outstation_sales_hawker']
                            json_data['other_retail']=json_data['other_sales_reatil']
                            json_data['other_dealer']=json_data['other_sales_dealer']
                            json_data['other_MT']=json_data['other_sales_modern_trade']
                            json_data['other_issued_damage']=json_data['other_issues_damage']
                            json_data['other_issues_return']=json_data['other_issues_return']
                            json_data['transaction_source'] = "SURYA_API"
                            json_data['plan_type'] = "DAILY"
                            json_data['distrcode'] = str(json_data['wd_code'])+"-"+str(json_data['town_code'])
                            log_data_serializer = Invalid_log_dataSerializers(data = json_data)
                            if log_data_serializer.is_valid():
                                log_data_serializer.save()
                            if total_sale == 0:
                                error_detail_dict['reason'] = str(json_data['wd_code']) +" "+" Received '0' sales for this SKU"+ str(json_data['sku_code'])
                            elif len(wd)>0:
                                wd_count = wd_count+1
                                error_detail_dict['repeat_count'] = wd_count
                                error_detail_dict['reason'] = str(json_data['wd_code']) +" "+" wd_id is not present in user_master database."
                            elif len(sku)>0:
                                sku_count = sku_count+1
                                error_detail_dict['repeat_count'] = sku_count
                                error_detail_dict['reason'] = str(json_data['sku_code']) +" "+ " sku_code is not present in sku_master database: sku_master"
                            elif len(town)>0:
                                town_count = town_count+1
                                error_detail_dict['repeat_count'] = town_count
                                error_detail_dict['reason'] = str(json_data['sku_code']) +" sku_code is not mapped with "+ str(json_data['wd_code']) +" wd_id in SS_database: Sales_Hierarchy_Master"
                            
                            const_time = datetime.datetime(2009, 10, 5, 00, 00, 00)
                            error_detail_dict['sku_id'] = str(json_data['sku_code'])
                            error_detail_dict['wd_id'] = str(json_data['wd_code'])
                            error_detail_dict['town_id'] = str(json_data['town_id'])
                    
                            error_detail_dict['tranisition_source'] = "SURYA_API"
                            comb_d_t = datetime.datetime.combine(parse(json_data['sale_date_time']), const_time.time())
                            error_detail_dict['sales_date_time'] = comb_d_t
                            log_detail = Integration_log_details.objects.filter(tranisition_source = "SURYA_API",sales_date_time = comb_d_t, sku_id = json_data['sku_code'],town_id = json_data['town_id']).last()
                            
                            if log_detail:
                                log_serializer = Integration_log_detailsSerializers(log_detail, data = error_detail_dict , partial = True)
                                
                                if log_serializer.is_valid():
                                    
                                    log_serializer.save()
                            else:
                                log_serializer = Integration_log_detailsSerializers(data = error_detail_dict)
                                
                                if log_serializer.is_valid():
                                    
                                    log_serializer.save()
                                else:
                                    return Response(log_serializer.errors)
                a = 0
                if valid_in_ss:
                    # print(a,"====")
                    for ss_store in valid_in_ss:
                        a=a+1
                        wd_list_4_week.append(ss_store.get('wd_id'))
                        sku_obj=SKU_Master_Product.objects.filter(sku_code=ss_store['sku_code']).last()
                        
                        local_sales_retailcalulated = ss_store.get('local_sales_reatil') if ss_store.get('local_sales_reatil', 0) else 0
                        local_sales_dealercalulated = ss_store.get('local_sales_dealer') if ss_store.get('local_sales_dealer', 0) else 0
                        # calculate suryaAPI sale value======================
                        if sku_obj and sku_obj.category_name == "CIGARETTE":
                            umo_code = sku_obj.weight_uom_code if sku_obj.weight_uom_code else 10
                            calculated_val = sku_obj.calculated_by if sku_obj.calculated_by else 100
                            local_sales_retailcalulated = float(ss_store.get('local_sales_reatil') if ss_store.get('local_sales_reatil', 0) else 0)/float(calculated_val)
                            local_sales_dealercalulated = float(ss_store.get('local_sales_dealer') if ss_store.get('local_sales_dealer', 0) else 0)/float(calculated_val)
                            
                            
                        ss_store['wd_id'] = ss_store['wd_code']
                        wd_town_list =Sales_Hierarchy_Master.objects.filter(Q(wd_id=ss_store.get('wd_id'))|Q(town_code=ss_store['town_id'])|Q(town_code__iexact = str(ss_store['town_id'])[1:])).values_list('wd_town_id',flat=True)
                        wdobj = WdSkuCatagory.objects.filter(sku_code=ss_store['sku_code'],wd_town_id__in=wd_town_list).last()
                        wdid = WDmaster.objects.filter(wd_ids = ss_store.get('wd_id')).last()
                        town_code1 = "0"+str(ss_store['town_id'])# 0 concatenate for town issue================
                        town_cde2 = str(ss_store['town_id']).strip('-')
                        wd_town_name =Sales_Hierarchy_Master.objects.filter(wd_id=ss_store.get('wd_id')).last()
                        wd_town_code =Sales_Hierarchy_Master.objects.filter((Q(town_code=ss_store['town_id'])|Q(town_code=town_code1)|Q(town_code=town_cde2)|Q(town_code = str(ss_store['town_id'])[1:])),wd_id=ss_store.get('wd_id')).last()
                        town_name = WDmaster.objects.filter(wd_ids = ss_store.get('wd_id'),wd_postal_code__icontains = wd_town_code.town_code).values('wd_postal_code').last()
                        ss_store['town_code'] = wd_town_code.town_code
                        ss_store['sku_short_name'] = sku_obj.sku_short_name
                        ss_store['sku_code'] = ss_store['sku_code']
                        ss_store['region'] = ss_store['region']
                        ss_store['wd_name'] = wdid.wd_name
                        ss_store['wd_type'] = wdid.wd_type
                        # print(town_name['wd_postal_code'])
                        if town_name:
                            ss_store['town_name'] = town_name['wd_postal_code']
                        else:
                            ss_store['town_name'] = None

                        ss_store['sku_id'] = sku_obj.sku_id
                        ss_store['transaction_source'] = "SURYA_API"
                        ss_store['statename'] = wdid.wd_state if wdid else None
                        ss_store['brand_category'] = sku_obj.category_code
                        ss_store['transaction_type'] = 'DAILY'
                        ss_store['created_by'] = "SURYA_API"
                        ss_store['created_date'] = datetime.datetime.now()
                        ss_store['sales_date_time'] = ss_store.get('sale_date_time')
                        ss_store['local_sales_retail'] = round(float(local_sales_retailcalulated),3)
                        ss_store['local_sales_dealer'] = round(float(local_sales_dealercalulated),3)
                        ss_store['local_sales_modern_trade'] = 0
                        ss_store['local_sales_hawker'] = 0
                        ss_store['outstation_sales_reatil'] = 0
                        ss_store['outstation_sales_dealer'] = 0
                        ss_store['outstation_sales_modern_trade'] = 0
                        ss_store['outstation_sales_hawker'] = 0
                        ss_store['other_sales_reatil'] = 0
                        ss_store['other_sales_dealer'] = 0
                        ss_store['other_sales_modern_trade'] = 0
                        ss_store['other_issues_other'] = 0
                        ss_store['other_issues_damage'] = 0
                        ss_store['other_issues_return'] = 0

                        ss_store['total_local_sales'] = round((float(ss_store['local_sales_retail']) + float(ss_store['local_sales_dealer']) + float(ss_store['local_sales_modern_trade']) + float(ss_store['local_sales_hawker'])),3)
                        ss_store['total_outstation_sales'] = round((float(ss_store['outstation_sales_reatil']) + float(ss_store['outstation_sales_dealer']) + float(ss_store['outstation_sales_hawker']) + float(ss_store['outstation_sales_modern_trade'])),3)
                        ss_store['total_other_sales'] = round((float(ss_store['other_sales_reatil']) + float(ss_store['other_sales_dealer']) + float(ss_store['other_sales_modern_trade'])),3)
                        ss_store['grand_total'] = round((ss_store['total_local_sales'] + ss_store['total_outstation_sales'] + ss_store['total_other_sales']),3)

                        ss_store['total_issue'] = round((float(ss_store['other_issues_other']) + float(ss_store['other_issues_damage']) + float(ss_store['other_issues_return'])),3)
                        
                        ss_store['company'] = sku_obj.company
                        branch = User.objects.filter(user_id = ss_store['wd_code']).last()
                        region = BranchMaster.objects.filter(branch_code = branch.locationcode).last()
                        ss_store['region'] = region.region
                        sale_data=""
                        if wdobj:
                            ss_store['unit_price'] = wdobj.last_price
                            ss_store['cnf_id'] = wdobj.cnf_id
                            ss_store['town_id'] =  wdobj.wd_town_id
                            ss_store['value'] = float(wdobj.last_price) * float(ss_store['grand_total'])
                            
                            sale_data = SalesData.objects.filter(wd_id = ss_store['wd_id'], sku_id = sku_obj.sku_id, sales_date_time=ss_store['sales_date_time'],town_id = wdobj.wd_town_id,town_code = wd_town_code.town_code,transaction_type = 'DAILY').last()
                        if sale_data :
                            ss_store.pop('created_by')
                            ss_store.pop('sales_date_time')
                            ss_store.pop('created_date')
                            ss_store['last_updated'] = "Surya_API"
                            ss_store['last_updated_date'] = datetime.datetime.now()
                            serializer = SalesDataSerializer(sale_data, data = ss_store,partial=True)
                            if serializer.is_valid():
                                serializer.save()
                                flag=flag+1
                            else:
                                errors_list.append(serializer.errors,ss_store['wd_code'],ss_store['sku_code'])
                                
                        else:
                            serializer = SalesDataSerializer(data = ss_store)
                            if serializer.is_valid():
                                serializer.save()
                                flag=flag+1
                            else:
                                errors_list.append(serializer.errors,)

                transaction_history['tranisition_source'] = "SURYA_API"
                transaction_history['sale_date'] = surya_api_sale_date.date()
                transaction_history['total_distributer_sale'] = len(surya_datas)
                transaction_history['total_insart_sale'] = len(valid_in_ss) if len(valid_in_ss) else 0
                transaction_history['created_by'] = "SURYA_API"
                transaction_history['created_date'] = datetime.datetime.now()
                transact_his = Integration_log_summary.objects.filter(sale_date = transaction_history['sale_date'],tranisition_source = "SURYA_API").last()
                print(transact_his,"=====",transaction_history['sale_date'])
                # print(transaction_history,"======transaction_history==== ",transaction_history)
                if transact_his is not None:
                    print("======update")
                    transaction_history.pop('created_date')
                    transaction_history['last_updated_date'] = datetime.datetime.now()
                    serializer_class = Integration_log_summarySerializers(transact_his,data = transaction_history,partial=True)
                    if serializer_class.is_valid():
                        serializer_class.save()
                    else:
                        # return Response(serializer_class.errors)
                        pass
                else:
                    print("======create")
                    
                    serializer_class = Integration_log_summarySerializers(data = transaction_history)
                    if serializer_class.is_valid():
                        # print("hiii__create")
                        serializer_class.save()
                    else:
                        # return Response(serializer_class.errors)
                        pass

                trans_obj=Integration_log_summary.objects.filter(sale_date=transaction_history['sale_date'],email_status = False,tranisition_source = "SURYA_API").values("total_distributer_sale","total_insart_sale","sale_date")
                log_detail = Integration_log_details.objects.filter(sales_date_time__date = surya_datas[0]['sale_date_time'],tranisition_source = "SURYA_API").values("sku_id","wd_id","town_id","sales_date_time","reason","repeat_count")
                row_no = 0
                if trans_obj or log_detail:
                    # subject = '"UAT | SURYA_API Integration Status Report."'
                    subject = surya_success_sub
                    
                    html_message = '<table style="width:100% ; border: 1px solid black;">'
                            
                    html_message +=            '<tr style="border: 1px solid black;">'
                    html_message +=                '<th "border: 1px solid black;">SNo.</th>'
                    html_message +=                '<th "border: 1px solid black;">Sale Date</th>'
                    html_message +=                '<th style="border: 1px solid black;">Total distributer sale</th>'
                    html_message +=                '<th style="border: 1px solid black;">Total insert sale</th>'
                    # html_message +=                '<th>town_id</th>'
                    # html_message +=                '<th>reason</th>'
                    html_message +=            '</tr>'
                    for ils in trans_obj:
                        # print(ils,"=================")
                        html_message +=            '<tr style="border: 1px solid red;">'
                        html_message +=                '<td style="border: 1px solid black;">'+str(row_no+1)+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['sale_date'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_distributer_sale'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_insart_sale'])+'</td>'
                        
                        html_message +=            '</tr>'
                    html_message += '</table> </br></br>'



                    # log details==============
                    html_message += '<table style="width:100% ; border: 1px solid black;">'
                            
                    html_message +=            '<tr style="border: 1px solid black;">'
                    html_message +=                '<th style="border: 1px solid black;">SNo.</th>'
                    html_message +=                '<th style="border: 1px solid black;">Sale Date</th>'
                    html_message +=                '<th style="border: 1px solid black;">Sku id</th>'
                    html_message +=                '<th style="border: 1px solid black;">Wd id</th>'
                    html_message +=                '<th style="border: 1px solid black;">Town id</th>'
                    html_message +=                '<th style="border: 1px solid black;">Reason</th>'
                    # html_message +=                '<th style="border: 1px solid black;">Repeat Count</th>'
                    html_message +=            '</tr>'
                    for ld in log_detail:
                        
                        row_no+=1
                        # print(ld['sales_date_time'].date(),"==========date=======")
                        html_message +=            '<tr style="border: 1px solid red;">'
                        html_message +=                '<td style="border: 1px solid black;">'+str(row_no)+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['sales_date_time'].date())+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['sku_id'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['wd_id'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['town_id'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['reason'])+'</td>'
                        # html_message +=                '<td style="border: 1px solid black;">'+str(ld['repeat_count'])+'</td>'

                        html_message +=            '</tr>'
                    html_message += '</table> </br></br>'

                    
                    email_from = settings.EMAIL_HOST_USER
                    cc_email = cc_list
                    # # ,'saurabha@triazinesoft.com','mayankg@triazinesoft.com'
                    recipient_list = success_email_list
                    # recipient_list = ["VIPULSHARMA-gpi@modi-ent.com","ankit.rakwal@esobene.com"]

                    reset_email = EmailMessage(
                                subject = subject,
                                body = html_message,
                                from_email = email_from,
                                to = recipient_list,
                                cc = cc_email,
                                reply_to = cc_email,
                                )
                    # trans_obj=Integration_log_summary.objects.filter(sale_date = surya_datas[0]['sale_date_time'],email_status = False,tranisition_source = "SURYA_API").update(email_status = True)
                    reset_email.content_subtype = "html"
                    reset_email.send(fail_silently=True)      
                    
                    
                
                # today = utils.now_date
                today = datetime.datetime.now()
                now_day=str(today.date())
                splited_today = now_day.split('-')[2]
                week_details = week_sele_before_date('BRANCH USER')
                print(wd_list_4_week,"==============wd_list_4_week//////==============",splited_today in date_tuple_weekly and week_details[2])
                Apistatus.objects.filter(running_date=today_date.date(),api="SFA_API").update(status=False)
                if int(splited_today) in date_tuple_weekly and week_details[2]:
                    print(wd_list_4_week,"==============wd_list_4_week==============")
                    create_update_weekly(wd_list_4_week)
                    
                return Response({"errors_list":errors_list,"error":error_obj_list})
            
            
            
            else:
                # subject = 'UAT | SURYA_API_No Sale'
                subject = surya_success_sub
                html_message = "<p>SURYA API Having No sales.</p>" if not surya_datas else "<p>SURYA API is still running.</p>"
                email_from = settings.EMAIL_HOST_USER
                cc_email = cc_list
                # # ,'saurabha@triazinesoft.com','mayankg@triazinesoft.com'
                recipient_list = failed_list
                # recipient_list = ["VIPULSHARMA-gpi@modi-ent.com","ankit.rakwal@esobene.com"]

                no_sale_mail = EmailMessage(
                            subject = subject,
                            body = html_message,
                            from_email = email_from,
                            to = recipient_list,
                            cc = cc_email,
                            reply_to = cc_email,
                            )
                # trans_obj=Integration_log_summary.objects.filter(sale_date = surya_datas[0]['sale_date_time'],email_status = False,tranisition_source = "SURYA_API").update(email_status = True)
                no_sale_mail.content_subtype = "html"
                no_sale_mail.send(fail_silently=True)
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong','error':error}
            Apistatus.objects.filter(running_date=today_date.date(),api="SFA_API").update(status=False)
        
            #subject = "UAT | SURYA_API Issues"
            subject = surya_failed_sub
            # recipient_list = ['rasmis@triazinesoft.com','saurabha@triazinesoft.com']
            recipient_list = failed_list
            # recipient_list = ["VIPULSHARMA-gpi@modi-ent.com","ankit.rakwal@esobene.com"]
            cc_email = cc_list
            html_message = str(context)
            reset_email = EmailMessage(
                                subject = subject,
                                body = html_message,
                                from_email = settings.EMAIL_HOST_USER,
                                to = recipient_list,
                                cc = cc_email,
                                reply_to = cc_email,
                                )
            reset_email.content_subtype = "html"
            reset_email.send(fail_silently=True)
            # self.get_queryset()
            return Response(context, status=status.HTTP_200_OK)




# Scheduler for SFA /liteapi data=======================
class Schedule(viewsets.ModelViewSet):
    def get_queryset(self):
        try:
            print(a=10)
            print('success_count',"========success_count=========")
            # breakpoint()
            # url = 'https://uat.lakshya.rsr.cloware.com/secondarysync/secondary_sales/0'
            # hashers = {'Authorization':'Basic  U0VDT05EQVJZX0lOVEVHUkFUSU9OX1RFQU06U0VDT05EQVJZU3luYzEyMyM=','Auth-Key':'secondaryauth','Content-Type':'application/json'}
            # production=======
            url = "https://drop52.reports.cloware.com/secondarysync/secondary_sales/0"
            hashers = {'Authorization':'Basic  U0VDT05EQVJZX0lOVEVHUkFUSU9OX1RFQU06U0VDT05EQVJZU3luYzEyMyM=','Auth-Key':'secondaryauth','Content-Type':'application/json'}

            data=requests.get(url, headers=hashers)
            success_count=data.json()['data']['success_count']
            s=0
            all_data=[]
            errors_list=[]
            flag=0
            repet_flag = 0
            histry_dict={}
            error_detail_dict = {}
            len_val = 0
            town_cd =[]
            wd_list_4_week = []
            date_tuple_weekly = [8,15,22,1]
            log_list = 0
            success_list = 0

            for i in range(0,int(success_count),500):
                if s==0:
                    # url = 'https://uat.lakshya.rsr.cloware.com/secondarysync/secondary_sales/0'
                    url = "https://drop52.reports.cloware.com/secondarysync/secondary_sales/0"
                    data=requests.get(url, headers=hashers)
                    total_data=data.json()['data']['distributer_sale']
                    all_data.append(total_data)
                else:
                    # url = 'https://uat.lakshya.rsr.cloware.com/secondarysync/secondary_sales/'+str(i+1)
                    url = "https://drop52.reports.cloware.com/secondarysync/secondary_sales/"+str(i+1)
                    data=requests.get(url, headers=hashers)
                    total_data=data.json()['data']['distributer_sale']
                    all_data.append(total_data)
                s=s+1
            final_data=list(chain(*all_data))
            json_object = json.dumps(all_data, indent = 4)
            # n = datetime.datetime.now().date()-datetime.timedelta(days=1)
            
            with open(settings.MEDIA_ROOT + 'sfa_api_sales/'+"sfa_sales_"+str(datetime.datetime.now().date()) + '.json','w')as outfile:
                outfile.write(json_object)
            if json_object[0]:
                response = HttpResponse(content_type='text/csv')
                df = pd.DataFrame.from_dict(list(final_data))
                df.to_csv(settings.MEDIA_ROOT + 'sfa_api_sales_csv/'+"sfa_sales_"+str(datetime.datetime.now().date()) + '.csv', index = False)
            # =========testcase===========================================
            # from master import sfa
            # success_count = sfa.aa['data']['success_count']
            # final_data = sfa.aa['data']['distributer_sale']
            # print(len(sfa.aa['data']['distributer_sale']),"==========",success_count)
            # # breakpoint()
            a = 0
            # ===========================end test case====================
            today_date = datetime.datetime.now()
            if int(success_count)>0:
                trans_obj=Integration_log_summary.objects.filter(sale_date=final_data[0].get('sale_date')).last()
                Integration_log_summary.objects.filter(created_date__date=today_date.date(),tranisition_source = "SFA/SFA_LITE_API").delete()
                Integration_log_details.objects.filter(created_date__date = today_date.date(), tranisition_source = "SFA/SFA_LITE_API").delete()
                for json_data in final_data:
                    a = a+1
                    print(a,"=====================================================================================//==")
                    sku_obj=SKU_Master_Product.objects.filter(sku_code__iexact=json_data['prodcode']).last()
                    wd_id = json_data['distrcode'].split("-")[0]
                    
                    town_trim_wd = str(json_data['distrcode']).replace(json_data['town_code'],'')
                    user_id = User.objects.filter(Q(user_id = wd_id)|Q(user_id = town_trim_wd)).last()
                    town_cd.append(json_data['town_code'])
                    # wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(wd_id=wd_id)|Q(wd_id = town_trim_wd)),(Q(town_code=json_data['town_code'])|Q(town_code__iexact=json_data['town_code'])|Q(town_code__iexact=str(json_data['town_code'])[1:])|Q(town_code__iexact='0'+str(json_data['town_code'])))).values_list('wd_town_id',flat=True)
                    wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(wd_id=wd_id)|Q(wd_id = town_trim_wd)),town_code__in=[json_data['town_code'],re.sub('\W+','',json_data['town_code']),str(json_data['town_code']),str(json_data['town_code'])[1:],'0'+str(json_data['town_code'])]).values_list('wd_town_id',flat=True)
                    wdobj = WdSkuCatagory.objects.filter(sku_code__iexact=json_data['prodcode'],wd_town_id__in=wd_town_list).last()
                    sku = []
                    wd = []
                    town = []
                    unmap_sku = []
                    if not user_id:
                        wd.append(wd_id)
                    elif not wd_town_list:
                        town.append(json_data['town_code'])
                    elif not wdobj:
                        unmap_sku.append(json_data['prodcode'])
                    elif not sku_obj:
                        sku.append(json_data['prodcode'])
                    total_sale = (float(json_data.get('local_retail', 0))+float(json_data.get('local_dealer', 0))+float(json_data.get('local_MT', 0))+float(json_data.get('local_HA', 0))
                                +float(json_data.get('out_retail', 0))+float(json_data.get('out_dealer', 0))+float(json_data.get('out_MT', 0))+float(json_data.get('out_HA', 0))
                                +float(json_data.get('other_retail', 0))+float(json_data.get('other_dealer', 0))+float(json_data.get('other_MT', 0)))
                    if sku_obj and wdobj and user_id and (total_sale > 0):
                        print(total_sale,"===================>>createing===================sfa======")
                        success_list = success_list + 1
                        wd_list_4_week.append(user_id.user_id)
                        
                        wds = WDmaster.objects.filter(wd_ids = user_id.user_id).last()
                        # town_nm = Sales_Hierarchy_Master.objects.filter((Q(wd_id = user_id.user_id)|Q(wd_id = town_trim_wd)),(Q(town_code__iexact=json_data['town_code'])|Q(town_code__iexact = str(json_data['town_code'])[1:])|Q(town_code__iexact='0'+str(json_data['town_code'])))).last()
                        town_nm = wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(wd_id=wd_id)|Q(wd_id = town_trim_wd)),town_code__in=[json_data['town_code'],re.sub('\W+','',json_data['town_code']),str(json_data['town_code']),str(json_data['town_code'])[1:],'0'+str(json_data['town_code'])]).last()
                        if town_nm:
                            town_name = town_nm.town
                        else:
                            town_name = " "

                        if wds:
                            wd_id_name = wds.wd_name
                        else:
                            wd_id_name = "Name not Avail in wd_master"
                            
                        json_data['distrcode'] = str(user_id.user_id)+"-"+json_data['town_code']
                        json_data['town_code'] = json_data['town_code']
                        json_data['sku_short_name'] = sku_obj.sku_short_name
                        json_data['sku_code'] = json_data['prodcode']
                        json_data['town_name'] = town_name

                        json_data['wd_name'] = wd_id_name
                        json_data['wd_type'] = json_data['dist_type']
                        
                        json_data['sku_id'] = sku_obj.sku_id
                        json_data['town_id'] = wdobj.wd_town_id
                        json_data['wd_id'] = user_id.user_id
                        json_data['tranisition_source'] = json_data['dist_type']
                        json_data['transaction_type'] = 'DAILY'
                        json_data['created_by'] = "SFA/SFA_LITE_API"
                        json_data['brand_category'] = json_data.get('catcode')
                        json_data['created_date'] = datetime.datetime.now()
                        json_data['sales_date_time'] = json_data.get('sale_date')
                        json_data['local_sales_retail'] = json_data.get('local_retail', 0)
                        json_data['local_sales_dealer'] = json_data.get('local_dealer', 0)
                        json_data['local_sales_modern_trade'] = json_data.get('local_MT', 0)
                        json_data['local_sales_hawker'] = json_data.get('local_HA', 0)
                        json_data['outstation_sales_reatil'] = json_data.get('out_retail', 0)
                        json_data['outstation_sales_dealer'] = json_data.get('out_dealer', 0)
                        json_data['outstation_sales_modern_trade'] = json_data.get('out_MT', 0)
                        json_data['outstation_sales_hawker'] = json_data.get('out_HA', 0)
                        json_data['other_sales_retail'] = json_data.get('other_retail', 0)
                        json_data['other_sales_dealer'] = json_data.get('other_dealer', 0)
                        json_data['other_sales_modern_trade'] = json_data.get('other_MT', 0)
                        json_data['other_issues_other'] = json_data.get('other_issues_other', 0)
                        json_data['other_issues_damage'] = json_data.get('other_issued_damage', 0)
                        json_data['other_issues_return'] = json_data.get('other_issued_returns', 0)
                        # json_data['last_updated_date'] = datetime.datetime.now()
                        json_data['total_local_sales'] = float(json_data['local_sales_retail']) + float(json_data['local_sales_dealer']) +float(json_data['local_sales_modern_trade'])+float(json_data['local_sales_hawker'])
                        json_data['total_outstation_sales'] = float(json_data['outstation_sales_reatil']) +float(json_data['outstation_sales_dealer']) +float(json_data['outstation_sales_modern_trade']) +float(json_data['outstation_sales_hawker'])
                        json_data['total_other_sales'] = float(json_data['other_sales_retail']) +float(json_data['other_sales_dealer']) +float(json_data['other_sales_modern_trade'])
                        json_data['total_issue'] = float(json_data['other_issues_other']) +float(json_data['other_issues_damage']) +float(json_data['other_issues_return'])
                        json_data['grand_total'] = json_data['total_local_sales'] +json_data['total_outstation_sales'] +json_data['total_other_sales']
                        if sku_obj:
                            json_data['company'] = sku_obj.company
                        else:
                            json_data['company'] =None
                        if wdobj:
                            json_data['unit_price'] = wdobj.last_price
                            json_data['cnf_id'] = wdobj.cnf_id
                        else:
                            json_data['unit_price'] = 0
                            json_data['cnf_id'] = None
                        branch = User.objects.filter(user_id = json_data.get('wd_id'), user_type = "WD").last()
                        if branch:
                            region = BranchMaster.objects.filter(branch_code = branch.locationcode).last()
                            json_data['region'] = region.region
                        else:
                            json_data['region'] = None

                        json_data['value'] = float(json_data['unit_price']) * float(json_data['grand_total'])
                        sale_data = SalesData.objects.filter(wd_id = json_data['wd_id'],sku_id = json_data['sku_id'],sales_date_time=json_data['sales_date_time'],town_id = json_data['town_id'],town_code = json_data['town_code'],status = True).last()
                        sale_data_ss = SalesData.objects.filter(wd_id = json_data['wd_id'],sku_id = json_data['sku_id'],sales_date_time=json_data['sales_date_time'],town_id = json_data['town_id'],town_code = json_data['town_code'],status = False).last()
                        # print(sale_data,"======")
                        if sale_data_ss :

                            # json_data['town_code'] = str(sale_data.town_code)
                            # json_data['sku_short_name'] = sku_obj.sku_short_name
                            # json_data['sku_code'] = json_data['prodcode']

                            # json_data['wd_name'] = wd_id_name
                            # json_data['wd_type'] = json_data['dist_type']

                            json_data['local_sales_retail'] = float(json_data['local_sales_retail']) + sale_data_ss.local_sales_retail
                            json_data['local_sales_dealer'] = float(json_data['local_sales_dealer']) + sale_data_ss.local_sales_dealer 
                            json_data['local_sales_modern_trade'] = float(json_data['local_sales_modern_trade']) + sale_data_ss.local_sales_modern_trade
                            json_data['local_sales_hawker'] = float(json_data['local_sales_hawker']) + sale_data_ss.local_sales_hawker
                            json_data['total_local_sales'] = float(json_data['local_sales_retail']) + json_data['local_sales_dealer'] +json_data['local_sales_modern_trade']+json_data['local_sales_hawker']

                            json_data['outstation_sales_reatil'] = float(json_data['outstation_sales_reatil']) + sale_data_ss.outstation_sales_reatil
                            json_data['outstation_sales_dealer'] = float(json_data['outstation_sales_dealer']) + sale_data_ss.outstation_sales_dealer
                            json_data['outstation_sales_modern_trade'] = float(json_data['outstation_sales_modern_trade']) + sale_data_ss.outstation_sales_modern_trade
                            json_data['outstation_sales_hawker'] = float(json_data['outstation_sales_hawker']) + sale_data_ss.outstation_sales_hawker
                            json_data['total_outstation_sales'] = float(json_data['outstation_sales_reatil']) + json_data['outstation_sales_dealer'] + json_data['outstation_sales_modern_trade'] + json_data['outstation_sales_hawker']

                            json_data['other_sales_retail'] = float(json_data['other_sales_retail']) + sale_data_ss.other_sales_retail
                            json_data['other_sales_dealer'] = float(json_data['other_sales_dealer']) + sale_data_ss.other_sales_dealer
                            json_data['other_sales_modern_trade'] = float(json_data['other_sales_modern_trade']) + sale_data_ss.other_sales_modern_trade
                            json_data['total_other_sales'] = float(json_data['other_sales_retail']) + json_data['other_sales_dealer'] + json_data['other_sales_modern_trade']

                            json_data['other_issues_damage'] = float(json_data['other_issues_damage']) + sale_data_ss.other_issues_damage
                            json_data['other_issues_return'] = float(json_data['other_issues_return']) + sale_data_ss.other_issues_return
                            json_data['other_issues_other'] = float(json_data['other_issues_other']) + sale_data_ss.other_issues_other
                            json_data['total_issue'] = json_data['other_issues_damage'] + json_data['other_issues_return']+json_data['other_issues_other']

                            json_data['grand_total'] = json_data['total_local_sales'] +json_data['total_outstation_sales'] +json_data['total_other_sales']
                            json_data['value'] = float(json_data['unit_price']) * float(json_data['grand_total'])

                            if wdobj:
                                json_data['unit_price'] = wdobj.last_price
                                json_data['cnf_id'] = wdobj.cnf_id
                            else:
                                json_data['unit_price'] = 0
                                json_data['cnf_id'] = None
                            json_data.pop('created_by')
                            json_data.pop('sales_date_time')
                            json_data.pop('created_date')
                            json_data['last_updated'] = "SFA/SFA_LITE_API"
                            json_data['last_updated_date'] = datetime.datetime.now()
                            serializer = SalesDataSerializer(sale_data_ss, data = json_data,partial=True)
                            if serializer.is_valid():
                                serializer.save()
                                repet_flag = repet_flag + 1
                                repeat_sl = {}
                                # print(wd_id,json_data['prodcode'],json_data['town_code'])
                                repeat_sale = Repeat_count.objects.filter(wd_id = json_data['wd_id'], sku_id = json_data['prodcode'],sale_date_time = json_data.get('sale_date'),transaction_source = "SFA/SFA_LITE_API").last()
                                # print(repeat_sale,"=======repeat_sale=====")
                                repeat_sl['wd_id'] = json_data['wd_id']
                                repeat_sl['sku_id'] = json_data['prodcode']
                                repeat_sl['town_id'] = json_data['town_code']
                                repeat_sl['sale_date_time'] = json_data.get('sale_date')
                                repeat_sl['transaction_source'] = "SFA/SFA_LITE_API"
                                if repeat_sale:
                                    repeat_sl['repeat_time'] = int(repeat_sale.repeat_time)+1
                                    repeat_sl['town_id'] = str(repeat_sale.town_id)+"-"+str(json_data['town_code'])
                                    print(repeat_sl,"=======repeat_sl=======")
                                    serializer = Repeat_countSerializers(repeat_sale, data = repeat_sl,partial =True)
                                    # print(repeat_sl['repeat_time'],"===repeat_sl['repeat_time']===")
                                    if serializer.is_valid():
                                        serializer.save()
                                        
                                    else:
                                        return Response(serializer.errors)
                                else:
                                    print("===2===")
                                    repeat_sl['repeat_time'] = int(1)
                                    serializer = Repeat_countSerializers(data = repeat_sl)
                                    if serializer.is_valid():
                                        serializer.save()
                        if sale_data:
                            print("=================if exist update ========================")
                            json_data['status'] = False
                            # json_data['status_3rd_party'] = True
                            serializer = SalesDataSerializer(sale_data,json_data,partial =True)
                            # print(repeat_sl['repeat_time'],"===repeat_sl['repeat_time']===")
                            if serializer.is_valid():
                                serializer.save()
                            else:
                                print(serializer.errors,"======error")
                                errors_list.append(serializer.errors)    
                        else:
                            json_data['status'] = False
                            serializer = SalesDataSerializer(data = json_data)
                            if serializer.is_valid():
                                serializer.save()
                                flag=flag+1
                            else:
                                errors_list.append(serializer.errors)
                        count_success_sale_date_wise(today_date, json_data.get('sale_date'))
                        
                        
                        # sale_date = json_data.get('sale_date')
                        # log_sumry = Integration_log_summary.objects.filter(created_date__date=today_date, sale_date = sale_date)
                        # if not log_sumry:
                        #     aa = Integration_log_summary.objects.create(sale_date = sale_date,total_insart_sale = 1,invalid = 0,tranisition_source = "SFA/SFA_LITE_API")
                        #     print(log_sumry,aa,"======call succ",today_date,sale_date)
                        #     # breakpoint()
                        # else:
                        #     log_sumry = log_sumry.last()
                        #     log_sumry.total_insart_sale = log_sumry.total_insart_sale + 1
                        #     log_sumry.save()
                        #     log_sumry.total_distributer_sale = log_sumry.invalid + log_sumry.total_insart_sale
                        #     log_sumry.save()
        
                        
                        
                    else:
                        count_invalid_sale_date_wise(today_date, json_data.get('sale_date'))
                        # today_date = datetime.datetime.now().date()
                        # sale_date = json_data.get('sale_date')
                        # log_sumry = Integration_log_summary.objects.filter(created_date__date=today_date, sale_date = sale_date)
                        # # print()
                        # if not log_sumry:
                        #     # breakpoint()
                        #     hh = Integration_log_summary.objects.create(sale_date = sale_date,total_insart_sale = 0,invalid = 1,tranisition_source = "SFA/SFA_LITE_API")
                        #     print(log_sumry,hh,"======call succ",today_date,sale_date)
                        # else:
                        #     log_sumry = log_sumry.last()
                        #     log_sumry.invalid = log_sumry.invalid + 1
                        #     log_sumry.save()
                        #     log_sumry.total_distributer_sale = log_sumry.invalid + log_sumry.total_insart_sale
                        #     log_sumry.save()
                        
                        log_list = log_list+1
                        print("=======00=======")
                        json_data['transaction_source'] = "SFA/SFA_LITE_API"
                        log_data_serializer = Invalid_log_dataSerializers(data = json_data)
                        if log_data_serializer.is_valid():
                            log_data_serializer.save()
                            
                            
                        # breakpoint()
                        len_val=len_val+1
                        errors_list.append(json_data)
                        const_time = datetime.datetime(2009, 10, 5, 00, 00, 00)
                        # wdid_town = json_data['distrcode'].split("-")
                        error_detail_dict['sku_id'] = json_data['prodcode']
                        error_detail_dict['wd_id'] = json_data['dist_id']
                        error_detail_dict['town_id'] = json_data['town_code']
                    #     sku.append(json_data['prodcode'])
                    # if not user_id:
                    #     wd.append(wd_id)
                    # if not wdobj:
                    #     town.append(json_data['town_code'])
                        if total_sale == 0:
                            error_detail_dict['reason'] = str(wd_id) +" "+" Received '0' sales for SKU "+ str(json_data['prodcode'])
                        elif len(wd)>0:
                            error_detail_dict['reason'] = str(wd_id) +" "+" wd_id is not present in user_master database."
                            print("===========ss",str(wd_id) +" "+" wd_id is not present in user_master database.")
                            # error_detail_dict['repeat_count'] = len_val
                        elif len(town)>0:
                            error_detail_dict['reason'] = str(wd_id)+ " WD_id is not mapped with " + str(json_data['town_code'])+" town code."
                            print("----------------------rrrr",str(wd_id)+ " WD_id is not mapped with " + str(json_data['town_code'])+" town code.")
                            
                        elif len(unmap_sku) > 0:
                            error_detail_dict['reason'] = str(json_data['prodcode']) +" sku_code is not mapped with "+ str(wd_id) +" wd_id in SS_database:Sales_Hierarchy_Master"
                            print("===========ss",str(json_data['prodcode']) +" sku_code is not mapped with "+ str(wd_id) +" wd_id in SS_database:Sales_Hierarchy_Master")
                            
                        elif len(sku)>0:
                            error_detail_dict['reason'] = str(json_data['prodcode']) +" "+ " sku_code is not present in sku_master database: sku_master"
                            print("===============88",str(json_data['prodcode']) +" "+ " sku_code is not present in sku_master database: sku_master")
                            # error_detail_dict['repeat_count'] = len_val
                        else:
                            # error_detail_dict['repeat_count'] = len_val
                            error_detail_dict['reason'] = str(json_data['town_code']) +"This towm_code is not mapped with any WD"

                        error_detail_dict['tranisition_source'] = "SFA/SFA_LITE_API"
                        # comb_d_t = datetime.datetime.combine(parse(final_data[0].get('sale_date')), const_time.time())
                        comb_d_t = parse(json_data['sale_date'])
                        # print(comb_d_t)
                        error_detail_dict['sales_date_time'] = comb_d_t
                        log_detail = Integration_log_details.objects.filter(tranisition_source = "SFA/SFA_LITE_API",sales_date_time = comb_d_t, sku_id = json_data['prodcode'],town_id = json_data['town_code']).last()
                        # ,town_id = json_data['town_code']
                        # print(log_detail,"======log_detail.repeat_count====")
                        # print(log_detail.repeat_count,"====log_detail.repeat_count====")
                        if log_detail:
                            # error_detail_dict['repeat_count'] = int(log_detail.repeat_count) +1
                            log_serializer = Integration_log_detailsSerializers(log_detail, data = error_detail_dict, partial=True)
                            
                            if log_serializer.is_valid():
                                
                                log_serializer.save()
                        else:
                            # print(error_detail_dict,"======error_detail_dict======")
                            error_detail_dict['repeat_count'] = 1
                            error_detail_dict['created_date'] = today_date
                            log_serializer = Integration_log_detailsSerializers(data = error_detail_dict)
                            
                            if log_serializer.is_valid():
                                
                                log_serializer.save()
                            else:
                                return Response(log_serializer.errors)
                    
                SalesData.objects.filter(wd_id__in = wd_list_4_week).update(status=True)
                    # Summary=============
                # add_sale_summary = Integration_log_summary.objects.filter(sale_date = final_data[0].get('sale_date'),tranisition_source = "SFA/SFA_LITE_API").last()
                # histry_dict["total_distributer_sale"] = int(success_count)
                # histry_dict["created_by"] = 'SFA/SFA_LITE_API'
                # histry_dict["total_insart_sale"]= int(success_count) - int(log_list)
                # histry_dict["sale_date"]=final_data[0].get('sale_date')
                # histry_dict['created_date'] = datetime.datetime.now()
                # histry_dict['tranisition_source'] = 'SFA/SFA_LITE_API'
                # if add_sale_summary:
                #     # print(histry_dict,'====histry_dict ===')
                #     historyserializers = Integration_log_summarySerializers(add_sale_summary, data=histry_dict, partial = True)
                #     if historyserializers.is_valid():
                #         historyserializers.save()
                #     else:
                #         return Response(historyserializers.errors)
                #     if flag!=len(success_count):
                #         print("send mail......",int(success_count) - int(len(errors_list)))
                # else:
                #     histry_dict.pop('created_date')
                #     histry_dict['last_updated_date'] = datetime.datetime.now()
                #     historyserializers = Integration_log_summarySerializers(data=histry_dict)
                #     if historyserializers.is_valid():
                #         historyserializers.save()
                #     if flag!=len(success_count):
                #         print("send mail......",int(success_count) - int(len(errors_list)))
                
                trans_obj = Integration_log_summary.objects.filter(created_date__date=today_date.date(),tranisition_source = "SFA/SFA_LITE_API" ).values("total_distributer_sale","total_insart_sale","sale_date")
                log_detail = Integration_log_details.objects.filter(created_date__date = today_date.date(), tranisition_source = "SFA/SFA_LITE_API").values("sku_id","wd_id","town_id","sales_date_time","reason")
                print(trans_obj,"======trans_obj==========mail=====")
                
                row_no = 0
                # today = utils.now_date
                today = datetime.datetime.now()
                now_day=str(today.date())
                splited_today = now_day.split('-')[2]
                week_details = week_sele_before_date('BRANCH USER')
                if int(splited_today) in date_tuple_weekly and week_details[2]:
                    print(wd_list_4_week,"==============wd_list_4_week==============")
                    create_update_weekly(wd_list_4_week)
                    # weekly_generate_on_weekend(wd_list_4_week)
                if trans_obj:
                    print(trans_obj,"======================================================mail============")
                    # subject = 'UAT Server| SFA/SFA LITE Integration Status Report.'
                    subject = sfa_success_sub
                    html_message = '<table style="width:100% ; border: 1px solid black;">'
                            
                    html_message +=            '<tr style="border: 1px solid black;">'
                    html_message +=                '<th "border: 1px solid black;">SNo.</th>'
                    html_message +=                '<th "border: 1px solid black;">Sale Date</th>'
                    html_message +=                '<th style="border: 1px solid black;">Total distributer sale</th>'
                    html_message +=                '<th style="border: 1px solid black;">Total insert sale</th>'
                    # html_message +=                '<th>town_id</th>'
                    # html_message +=                '<th>reason</th>'
                    html_message +=            '</tr>'
                    for ils in trans_obj:
                        # print(ils,"=================")
                        html_message +=            '<tr style="border: 1px solid red;">'
                        html_message +=                '<td style="border: 1px solid black;">'+str(row_no+1)+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['sale_date'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_distributer_sale'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_insart_sale'])+'</td>'
                        
                        html_message +=            '</tr>'
                    html_message += '</table> </br></br>'



                    # log details==============
                    html_message += '<table style="width:100% ; border: 1px solid black;">'
                            
                    html_message +=            '<tr style="border: 1px solid black;">'
                    html_message +=                '<th style="border: 1px solid black;">SNo.</th>'
                    html_message +=                '<th style="border: 1px solid black;">Sale Date</th>'
                    html_message +=                '<th style="border: 1px solid black;">Sku id</th>'
                    html_message +=                '<th style="border: 1px solid black;">Wd id</th>'
                    html_message +=                '<th style="border: 1px solid black;">Town id</th>'
                    html_message +=                '<th style="border: 1px solid black;">Reason</th>'
                    # html_message +=                '<th style="border: 1px solid black;">Repeat Count</th>'
                    html_message +=            '</tr>'
                    for ld in log_detail:
                        
                        row_no+=1
                        # print(ld['sales_date_time'].date(),"==========date=======")
                        html_message +=            '<tr style="border: 1px solid red;">'
                        html_message +=                '<td style="border: 1px solid black;">'+str(row_no)+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['sales_date_time'].date())+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['sku_id'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['wd_id'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['town_id'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['reason'])+'</td>'
                        # html_message +=                '<td style="border: 1px solid black;">'+str(ld['repeat_count'])+'</td>'
                        html_message +=            '</tr>'
                    html_message += '</table> </br></br>'

                    
                    email_from = settings.EMAIL_HOST_USER
                    cc_email = cc_list
                    # ,'saurabha@triazinesoft.com','mayankg@triazinesoft.com'
                    recipient_list = success_email_list
                    # recipient_list = ["VIPULSHARMA-gpi@modi-ent.com","ankit.rakwal@esobene.com"]

                    reset_email = EmailMessage(
                                subject = subject,
                                body = html_message,
                                from_email = email_from,
                                to = recipient_list,
                                cc = cc_email,
                                reply_to = cc_email,
                                )
                    trans_obj=Integration_log_summary.objects.filter(sale_date=final_data[0].get('sale_date'),email_status = False).update(email_status = True)
                    reset_email.content_subtype = "html"
                    reset_email.send(fail_silently=True)
                return Response({"sku":sku,"wd":wd,"town":town})
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong','error':error}
            # subject = "UAT | SFA/SFA_lite Issue"
            subject = sfa_failed_sub
            recipient_list = failed_list
            # recipient_list = ["VIPULSHARMA-gpi@modi-ent.com","ankit.rakwal@esobene.com"]
            cc_email = cc_list
            html_message = str(context)
            reset_email = EmailMessage(
                                subject = subject,
                                body = html_message,
                                from_email = settings.EMAIL_HOST_USER,
                                to = recipient_list,
                                cc = cc_email,
                                reply_to = cc_email,
                                )
            reset_email.content_subtype = "html"
            reset_email.send(fail_silently=True)
            self.get_queryset()
            return Response(context, status=status.HTTP_200_OK)


class wdupdate(APIView):
    # permission_classes = (IsAuthenticated,)
    serializer_class = BranchMasterSerializers


    def post(self, request):
        file_obj = request.FILES['file']
        wb_obj = openpyxl.load_workbook(file_obj)
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row
        dicts = []
        try:
            for i in range(2, m_row + 1):
                empdata={}
                datas=sheet_obj.cell(row = i, column = 1).value
                a=datas.split("-")[0]
                print("====4===",a)
                # serializer = BranchMasterSerializers(data=empdata)
                # if serializer.is_valid():
                #     serializer.save()
                # else:
                #     return Response(serializer.errors, status=status.HTTP_200_OK)

            message={"ok":"ok"}
            return Response(message, status=status.HTTP_200_OK)
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong'}
            return Response(context, status=status.HTTP_200_OK)



class Surya_API(APIView):
    def get(self, request):
        try:
            print(a=15)
            n= datetime.datetime.now().date()
            from_date = request.GET.get("from_date",None)
            to_date = request.GET.get("to_date",None)
            aaa = WDmaster.objects.filter(wd_ids = "2247",wd_postal_code__icontains = '0212802').values('wd_postal_code')
            # print(week_sele_before_date('BRANCH USER'),"=======================>>>")
            # breakpoint()
            # print(str(to_date),"=====from_date=====")
            # url = "http://gpisfatest.winitsoftware.com/api/api/GPIInterface/GET?StartDate="+str(start_date)+"&EndDate="+str(start_date)+"9&Passcode=GP!W!N"
            # url = "http://gpisfatest.winitsoftware.com/api/api/GPIInterface/GET?StartDate=2022-03-23&EndDate=2022-03-23&Passcode=GP!W!N"
            # url="http://gpisfatest.winitsoftware.com/api/api/GPIInterface/GET?StartDate="+str(from_date)+"&EndDate="+str(to_date)+"&Passcode=GP!W!N"
            print(from_date,"====",from_date)
            if from_date and to_date:
                print("=======dt======")
               # url="https://ss.godfreyphillips.co/api_ss/api/GPIInterface/GET?StartDate="+str(from_date)+"&EndDate="+str(to_date)+"&Passcode=GP!W!N"
                url="https://ss.godfreyphillips.co/GPIInterFaceAPI_new/api/GPIInterface?Passcode=GP!W!N&StartDate="+str(from_date)+"&EndDate="+str(to_date)+"&Passcode=GP!W!N"

            else:
                print("===nn====dt======")
                # url = "https://ss.godfreyphillips.co/api_ss/api/GPIInterface?Passcode=GP!W!N"
                url = "https://ss.godfreyphillips.co/GPIInterFaceAPI_new/api/GPIInterface?Passcode=GP!W!N"
                
            # url = "https://ss.godfreyphillips.co/api_ss/api/GPIInterface?Passcode=GP!W!N"
            surya_data = requests.get(url).json()
            print("surya_data",url,"======surya_data ======")
            errors_list = []
            error_obj_list = []
            valid_in_ss = []
            transaction_history = {}
            error_detail_dict = {}
            flag=0
            wd_list_4_week = []
            date_tuple_weekly = [8,15,22,1]
            
            wd_count = 0
            sku_count = 0
            town_count = 0
            n = datetime.datetime.now().date()-datetime.timedelta(days=1)
            # today = utils.now_date
            today = datetime.datetime.now()
            now_day=str(today.date())
            splited_today = now_day.split('-')[2]
            week_details = week_sele_before_date('BRANCH USER')
        
        
            surya_api_sale_date=parse(surya_data.get('StartDate'))
            
            surya_datas=surya_data.get('Data')
            json_object = json.dumps(surya_datas, indent = 4)
            with open(settings.MEDIA_ROOT + 'surya_api_sales/'+"surya_sales_"+str(datetime.datetime.now().date()) + '.json','w')as outfile:
                outfile.write(json_object)
            if surya_datas:
                response = HttpResponse(content_type='text/csv')
                # response['Content-Disposition'] = 'attachment; filename="sec_sales_data.csv"'
                df = pd.DataFrame.from_dict(list(surya_datas))
                df.to_csv(settings.MEDIA_ROOT + 'surya_api_sales_csv/'+"surya_sales_"+str(datetime.datetime.now().date()) + '.csv', index = False)

            if surya_datas:
                
                for json_data in surya_datas:
                    wd = []
                    sku = []
                    town = []
                    if json_data['sku_code'] and json_data['wd_code'] and json_data['town_id']:
                        sku_obj=SKU_Master_Product.objects.filter(sku_code=json_data['sku_code']).last()
                        json_data['wd_id'] = json_data['wd_code']
                        
                        user = User.objects.filter(user_id = json_data['wd_code'])
                        town_cde = str("0")+str(json_data['town_id'])# 0 concatenate for town issue================
                        town_cde2 = str(json_data['town_id']).strip('-')
                        wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(town_code=json_data['town_id'])|Q(town_code=town_cde)|Q(town_code=town_cde2)|Q(town_code__iexact = str(json_data['town_id'])[1:])),wd_id=json_data.get('wd_id')).values_list('wd_town_id',flat=True)
                        wdobj = WdSkuCatagory.objects.filter(sku_code=json_data['sku_code'],wd_town_id__in=wd_town_list).last()
                        
                        if not sku_obj:
                            sku.append(json_data['sku_code'])
                        if not user:
                            wd.append(json_data['wd_code'])
                        if not wdobj:
                            town.append(json_data['town_id'])
                        total_sale = (
                                float(json_data.get('local_sales_reatil') if json_data.get('local_sales_reatil', 0) else 0) 
                                + float(json_data.get('local_sales_dealer') if json_data.get('local_sales_dealer', 0) else 0)
                                + float(json_data.get('local_sales_modern_trade') if json_data.get('local_sales_modern_trade', 0) else 0)
                                + float(json_data.get('local_sales_hawker') if json_data.get('local_sales_hawker', 0) else 0)
                                
                                + float(json_data.get('outstation_sales_reatil') if json_data.get('outstation_sales_reatil', 0) else 0) 
                                + float(json_data.get('outstation_sales_dealer') if json_data.get('outstation_sales_dealer', 0) else 0)
                                + float(json_data.get('outstation_sales_modern_trade') if json_data.get('outstation_sales_modern_trade', 0) else 0)
                                + float(json_data.get('outstation_sales_hawker') if json_data.get('outstation_sales_hawker', 0) else 0)
                                
                                + float(json_data.get('other_sales_reatil') if json_data.get('other_sales_reatil', 0) else 0) 
                                + float(json_data.get('other_sales_dealer') if json_data.get('other_sales_dealer', 0) else 0)
                                + float(json_data.get('other_sales_hawker') if json_data.get('other_sales_hawker', 0) else 0)
                                + float(json_data.get('other_sales_modern_trade') if json_data.get('other_sales_modern_trade', 0) else 0)
                            )

                        if sku_obj and wdobj and user and (total_sale > 0):
                            valid_in_ss.append(json_data)
                        else:
                            json_data['dist_id']=json_data['wd_code']
                            json_data['prodcode']=json_data['sku_code']
                            json_data['town_code']=json_data['town_id']
                            json_data['sale_date']=json_data['sale_date_time']
                            json_data['region']=json_data['region']
                            json_data['local_retail']=json_data['local_sales_reatil']
                            json_data['local_dealer']=json_data['local_sales_dealer']
                            json_data['local_MT']=json_data['local_sales_modern_trade']
                            json_data['local_HA']=json_data['local_sales_hawker']
                            json_data['out_retail']=json_data['outstation_sales_reatil']
                            json_data['out_dealer']=json_data['outstation_sales_dealer']
                            json_data['out_MT']=json_data['outstation_sales_modern_trade']
                            json_data['out_HA']=json_data['outstation_sales_hawker']
                            json_data['other_retail']=json_data['other_sales_reatil']
                            json_data['other_dealer']=json_data['other_sales_dealer']
                            json_data['other_MT']=json_data['other_sales_modern_trade']
                            json_data['other_issued_damage']=json_data['other_issues_damage']
                            json_data['other_issues_return']=json_data['other_issues_return']
                            json_data['transaction_source'] = "SURYA_API"
                            json_data['plan_type'] = "DAILY"
                            json_data['distrcode'] = str(json_data['wd_code'])+"-"+str(json_data['town_code'])
                            log_data_serializer = Invalid_log_dataSerializers(data = json_data)
                            if log_data_serializer.is_valid():
                                log_data_serializer.save()
                            if total_sale == 0:
                                error_detail_dict['reason'] = str(json_data['wd_code']) +" "+" Received '0' sales for this SKU"+ str(json_data['sku_code'])
                            elif len(wd)>0:
                                wd_count = wd_count+1
                                error_detail_dict['repeat_count'] = wd_count
                                error_detail_dict['reason'] = str(json_data['wd_code']) +" "+" wd_id is not present in user_master database."
                            elif len(sku)>0:
                                sku_count = sku_count+1
                                error_detail_dict['repeat_count'] = sku_count
                                error_detail_dict['reason'] = str(json_data['sku_code']) +" "+ " sku_code is not present in sku_master database: sku_master"
                            elif len(town)>0:
                                town_count = town_count+1
                                error_detail_dict['repeat_count'] = town_count
                                error_detail_dict['reason'] = str(json_data['sku_code']) +" sku_code is not mapped with "+ str(json_data['wd_code']) +" wd_id in SS_database: Sales_Hierarchy_Master"
                            
                            const_time = datetime.datetime(2009, 10, 5, 00, 00, 00)
                            error_detail_dict['sku_id'] = str(json_data['sku_code'])
                            error_detail_dict['wd_id'] = str(json_data['wd_code'])
                            error_detail_dict['town_id'] = str(json_data['town_id'])
                    
                            error_detail_dict['tranisition_source'] = "SURYA_API"
                            comb_d_t = datetime.datetime.combine(parse(json_data['sale_date_time']), const_time.time())
                            error_detail_dict['sales_date_time'] = comb_d_t
                            log_detail = Integration_log_details.objects.filter(tranisition_source = "SURYA_API",sales_date_time = comb_d_t, sku_id = json_data['sku_code'],town_id = json_data['town_id']).last()
                            
                            if log_detail:
                                log_serializer = Integration_log_detailsSerializers(log_detail, data = error_detail_dict , partial = True)
                                
                                if log_serializer.is_valid():
                                    
                                    log_serializer.save()
                            else:
                                log_serializer = Integration_log_detailsSerializers(data = error_detail_dict)
                                
                                if log_serializer.is_valid():
                                    
                                    log_serializer.save()
                                else:
                                    return Response(log_serializer.errors)
                a = 0
                if valid_in_ss:
                    # print(a,"====")
                    for ss_store in valid_in_ss:
                        a=a+1
                        wd_list_4_week.append(ss_store.get('wd_id'))
                        sku_obj=SKU_Master_Product.objects.filter(sku_code=ss_store['sku_code']).last()
                        
                        local_sales_retailcalulated = ss_store.get('local_sales_reatil') if ss_store.get('local_sales_reatil', 0) else 0
                        local_sales_dealercalulated = ss_store.get('local_sales_dealer') if ss_store.get('local_sales_dealer', 0) else 0
                        # calculate suryaAPI sale value======================
                        if sku_obj and sku_obj.category_name == "CIGARETTE":
                            umo_code = sku_obj.weight_uom_code if sku_obj.weight_uom_code else 10
                            calculated_val = sku_obj.calculated_by if sku_obj.calculated_by else 100
                            local_sales_retailcalulated = float(ss_store.get('local_sales_reatil') if ss_store.get('local_sales_reatil', 0) else 0)/float(calculated_val)
                            local_sales_dealercalulated = float(ss_store.get('local_sales_dealer') if ss_store.get('local_sales_dealer', 0) else 0)/float(calculated_val)
                            

                        ss_store['wd_id'] = ss_store['wd_code']
                        wd_town_list =Sales_Hierarchy_Master.objects.filter(Q(wd_id=ss_store.get('wd_id'))|Q(town_code=ss_store['town_id'])|Q(town_code__iexact = str(ss_store['town_id'])[1:])).values_list('wd_town_id',flat=True)
                        wdobj = WdSkuCatagory.objects.filter(sku_code=ss_store['sku_code'],wd_town_id__in=wd_town_list).last()
                        wdid = WDmaster.objects.filter(wd_ids = ss_store.get('wd_id')).last()
                        town_code1 = "0"+str(ss_store['town_id'])# 0 concatenate for town issue================
                        town_cde2 = str(ss_store['town_id']).strip('-')
                        wd_town_name =Sales_Hierarchy_Master.objects.filter(wd_id=ss_store.get('wd_id')).last()
                        wd_town_code =Sales_Hierarchy_Master.objects.filter((Q(town_code=ss_store['town_id'])| Q(town_code__iexact = str(ss_store['town_id'])[1:]) |Q(town_code=town_code1)|Q(town_code=town_cde2)),wd_id=ss_store.get('wd_id')).last()
                        town_name = WDmaster.objects.filter(wd_ids = ss_store.get('wd_id'),wd_postal_code__icontains = wd_town_code.town_code).values('wd_postal_code').last()
                        ss_store['town_code'] = wd_town_code.town_code
                        ss_store['sku_short_name'] = sku_obj.sku_short_name
                        ss_store['sku_code'] = ss_store['sku_code']
                        ss_store['region'] = ss_store['region']
                        ss_store['wd_name'] = wdid.wd_name
                        ss_store['wd_type'] = wdid.wd_type
                        # print(town_name['wd_postal_code'])
                        if town_name:
                            ss_store['town_name'] = town_name['wd_postal_code']
                        else:
                            ss_store['town_name'] = None

                        ss_store['sku_id'] = sku_obj.sku_id
                        ss_store['transaction_source'] = "SURYA_API"
                        ss_store['statename'] = wdid.wd_state if wdid else None
                        ss_store['brand_category'] = sku_obj.category_code
                        ss_store['transaction_type'] = 'DAILY'
                        ss_store['created_by'] = "SURYA_API"
                        ss_store['created_date'] = datetime.datetime.now()
                        ss_store['sales_date_time'] = ss_store.get('sale_date_time')
                        ss_store['local_sales_retail'] = round(float(local_sales_retailcalulated),3)
                        ss_store['local_sales_dealer'] = round(float(local_sales_dealercalulated),3)
                        ss_store['local_sales_modern_trade'] = 0
                        ss_store['local_sales_hawker'] = 0
                        ss_store['outstation_sales_reatil'] = 0
                        ss_store['outstation_sales_dealer'] = 0
                        ss_store['outstation_sales_modern_trade'] = 0
                        ss_store['outstation_sales_hawker'] = 0
                        ss_store['other_sales_reatil'] = 0
                        ss_store['other_sales_dealer'] = 0
                        ss_store['other_sales_modern_trade'] = 0
                        ss_store['other_issues_other'] = 0
                        ss_store['other_issues_damage'] = 0
                        ss_store['other_issues_return'] = 0

                        ss_store['total_local_sales'] = round((float(ss_store['local_sales_retail'],) + float(ss_store['local_sales_dealer'],) + float(ss_store['local_sales_modern_trade'],) + float(ss_store['local_sales_hawker'],)),3)
                        ss_store['total_outstation_sales'] = round((float(ss_store['outstation_sales_reatil']) + float(ss_store['outstation_sales_dealer']) + float(ss_store['outstation_sales_hawker']) + float(ss_store['outstation_sales_modern_trade'])),3)
                        ss_store['total_other_sales'] = round((float(ss_store['other_sales_reatil']) + float(ss_store['other_sales_dealer']) + float(ss_store['other_sales_modern_trade'])),3)
                        ss_store['grand_total'] = round((ss_store['total_local_sales'] + ss_store['total_outstation_sales'] + ss_store['total_other_sales']),3)

                        ss_store['total_issue'] = round((float(ss_store['other_issues_other']) + float(ss_store['other_issues_damage']) + float(ss_store['other_issues_return'])),3)
                        
                        ss_store['company'] = sku_obj.company
                        branch = User.objects.filter(user_id = ss_store['wd_code']).last()
                        region = BranchMaster.objects.filter(branch_code = branch.locationcode).last()
                        ss_store['region'] = region.region
                        sale_data=""
                        if wdobj:
                            ss_store['unit_price'] = wdobj.last_price
                            ss_store['cnf_id'] = wdobj.cnf_id
                            ss_store['town_id'] =  wdobj.wd_town_id
                            ss_store['value'] = float(wdobj.last_price) * float(ss_store['grand_total'])
                            
                            sale_data = SalesData.objects.filter(wd_id = ss_store['wd_id'], sku_id = sku_obj.sku_id, sales_date_time=ss_store['sales_date_time'],town_id = wdobj.wd_town_id,town_code = wd_town_code.town_code).last()
                        
                        if sale_data :
                            ss_store.pop('created_by')
                            ss_store.pop('sales_date_time')
                            ss_store.pop('created_date')
                            ss_store['last_updated'] = "Surya_API"
                            ss_store['last_updated_date'] = datetime.datetime.now()
                            serializer = SalesDataSerializer(sale_data, data = ss_store,partial=True)
                            if serializer.is_valid():
                                serializer.save()
                                flag=flag+1
                            else:
                                errors_list.append(serializer.errors,ss_store['wd_code'],ss_store['sku_code'])
                                
                        else:
                            serializer = SalesDataSerializer(data = ss_store)
                            if serializer.is_valid():
                                serializer.save()
                                flag=flag+1
                            else:
                                errors_list.append(serializer.errors,)

                transaction_history['tranisition_source'] = "SURYA_API"
                transaction_history['sale_date'] = surya_api_sale_date.date()
                transaction_history['total_distributer_sale'] = len(surya_datas)
                transaction_history['total_insart_sale'] = len(valid_in_ss) if len(valid_in_ss) else 0
                transaction_history['created_by'] = "SURYA_API"
                transaction_history['created_date'] = datetime.datetime.now()
                transact_his = Integration_log_summary.objects.filter(sale_date = transaction_history['sale_date'],tranisition_source = "SURYA_API").last()
                print(transact_his,"=====",transaction_history['sale_date'])
                # print(transaction_history,"======transaction_history==== ",transaction_history)
                if transact_his is not None:
                    print("======update")
                    transaction_history.pop('created_date')
                    transaction_history['last_updated_date'] = datetime.datetime.now()
                    serializer_class = Integration_log_summarySerializers(transact_his,data = transaction_history,partial=True)
                    if serializer_class.is_valid():
                        serializer_class.save()
                    else:
                        # return Response(serializer_class.errors)
                        pass
                else:
                    print("======create")
                    
                    serializer_class = Integration_log_summarySerializers(data = transaction_history)
                    if serializer_class.is_valid():
                        # print("hiii__create")
                        serializer_class.save()
                    else:
                        # return Response(serializer_class.errors)
                        pass

                trans_obj=Integration_log_summary.objects.filter(sale_date=transaction_history['sale_date'],email_status = False,tranisition_source = "SURYA_API").values("total_distributer_sale","total_insart_sale","sale_date")
                log_detail = Integration_log_details.objects.filter(sales_date_time__date = surya_datas[0]['sale_date_time'],tranisition_source = "SURYA_API").values("sku_id","wd_id","town_id","sales_date_time","reason","repeat_count")
                row_no = 0
                if trans_obj or log_detail:
                    # subject = '"UAT | SURYA_API Integration Status Report."'
                    subject = surya_success_sub
                    html_message = '<table style="width:100% ; border: 1px solid black;">'
                            
                    html_message +=            '<tr style="border: 1px solid black;">'
                    html_message +=                '<th "border: 1px solid black;">SNo.</th>'
                    html_message +=                '<th "border: 1px solid black;">Sale Date</th>'
                    html_message +=                '<th style="border: 1px solid black;">Total distributer sale</th>'
                    html_message +=                '<th style="border: 1px solid black;">Total insert sale</th>'
                    # html_message +=                '<th>town_id</th>'
                    # html_message +=                '<th>reason</th>'
                    html_message +=            '</tr>'
                    for ils in trans_obj:
                        # print(ils,"=================")
                        html_message +=            '<tr style="border: 1px solid red;">'
                        html_message +=                '<td style="border: 1px solid black;">'+str(row_no+1)+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['sale_date'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_distributer_sale'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_insart_sale'])+'</td>'
                        
                        html_message +=            '</tr>'
                    html_message += '</table> </br></br>'



                    # log details==============
                    html_message += '<table style="width:100% ; border: 1px solid black;">'
                            
                    html_message +=            '<tr style="border: 1px solid black;">'
                    html_message +=                '<th style="border: 1px solid black;">SNo.</th>'
                    html_message +=                '<th style="border: 1px solid black;">Sale Date</th>'
                    html_message +=                '<th style="border: 1px solid black;">Sku id</th>'
                    html_message +=                '<th style="border: 1px solid black;">Wd id</th>'
                    html_message +=                '<th style="border: 1px solid black;">Town id</th>'
                    html_message +=                '<th style="border: 1px solid black;">Reason</th>'
                    # html_message +=                '<th style="border: 1px solid black;">Repeat Count</th>'
                    html_message +=            '</tr>'
                    for ld in log_detail:
                        
                        row_no+=1
                        # print(ld['sales_date_time'].date(),"==========date=======")
                        html_message +=            '<tr style="border: 1px solid red;">'
                        html_message +=                '<td style="border: 1px solid black;">'+str(row_no)+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['sales_date_time'].date())+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['sku_id'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['wd_id'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['town_id'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['reason'])+'</td>'
                        # html_message +=                '<td style="border: 1px solid black;">'+str(ld['repeat_count'])+'</td>'

                        html_message +=            '</tr>'
                    html_message += '</table> </br></br>'

                    
                    email_from = settings.EMAIL_HOST_USER
                    cc_email = cc_list
                    # # ,'saurabha@triazinesoft.com','mayankg@triazinesoft.com'
                    # recipient_list = ["rasmis@triazinesoft.com"]
                    recipient_list = success_email_list

                    reset_email = EmailMessage(
                                subject = subject,
                                body = html_message,
                                from_email = email_from,
                                to = recipient_list,
                                cc = cc_email,
                                reply_to = cc_email,
                                )
                    # trans_obj=Integration_log_summary.objects.filter(sale_date = surya_datas[0]['sale_date_time'],email_status = False,tranisition_source = "SURYA_API").update(email_status = True)
                    reset_email.content_subtype = "html"
                    reset_email.send(fail_silently=True)      
                    
                    
                
                # today = utils.now_date
                today = datetime.datetime.now()
                now_day=str(today.date())
                splited_today = now_day.split('-')[2]
                week_details = week_sele_before_date('BRANCH USER')
                
                if int(splited_today) in date_tuple_weekly and week_details[2]:
                    create_update_weekly(wd_list_4_week)
                    
                return Response({"errors_list":errors_list,"error":error_obj_list})
            else:
                # subject = '"SURYA_API_No Sale"'
                subject = surya_success_sub
                html_message = "SURYA API Having No sales.</p>"
                email_from = settings.EMAIL_HOST_USER
                cc_email = cc_list
                # # ,'saurabha@triazinesoft.com','mayankg@triazinesoft.com'
                recipient_list = failed_list
                # recipient_list = ["VIPULSHARMA-gpi@modi-ent.com","ankit.rakwal@esobene.com"]

                no_sale_mail = EmailMessage(
                            subject = subject,
                            body = html_message,
                            from_email = email_from,
                            to = recipient_list,
                            cc = cc_email,
                            reply_to = cc_email,
                            )
                no_sale_mail.content_subtype = "html"
                no_sale_mail.send(fail_silently=True)
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong','error':error}
            subject = "UAT | SURYA_API Issues"
            subject = surya_failed_sub
            # recipient_list = ["VIPULSHARMA-gpi@modi-ent.com","ankit.rakwal@esobene.com"]
            cc_email = cc_list
            recipient_list = failed_list
            html_message = str(context)
            reset_email = EmailMessage(
                                subject = subject,
                                body = html_message,
                                from_email = settings.EMAIL_HOST_USER,
                                to = recipient_list,
                                cc = cc_email,
                                reply_to = cc_email,
                                )
            reset_email.content_subtype = "html"
            reset_email.send(fail_silently=True)
            self.get(request)
            return Response(context, status=status.HTTP_200_OK)


import json
from dateutil.parser import parse
import pandas as pd
class SFA_API_valid(APIView):
    serializer_class = SalesDataSerializer
    def get(self, request):
        today_date = datetime.datetime.now()
        try:
            # url = 'https://uat.lakshya.rsr.cloware.com/secondarysync/secondary_sales/0'
            # hashers = {'Authorization':'Basic  U0VDT05EQVJZX0lOVEVHUkFUSU9OX1RFQU06U0VDT05EQVJZU3luYzEyMyM=','Auth-Key':'secondaryauth','Content-Type':'application/json'}
            # production=======
            url = "https://drop52.reports.cloware.com/secondarysync/secondary_sales/0"
            hashers = {'Authorization':'Basic  U0VDT05EQVJZX0lOVEVHUkFUSU9OX1RFQU06U0VDT05EQVJZU3luYzEyMyM=','Auth-Key':'secondaryauth','Content-Type':'application/json'}

            data=requests.get(url, headers=hashers)
            success_count=data.json()['data']['success_count']
            print(success_count,"========success_count=========")
            s=0
            all_data=[]
            errors_list=[]
            flag=0
            repet_flag = 0
            histry_dict={}
            error_detail_dict = {}
            len_val = 0
            town_cd =[]
            wd_list_4_week = []
            date_tuple_weekly = [8,15,22,1]
            log_list = 0
            success_list = 0

            for i in range(0,int(success_count),500):
                if s==0:
                    # url = 'https://uat.lakshya.rsr.cloware.com/secondarysync/secondary_sales/0'
                    url = "https://drop52.reports.cloware.com/secondarysync/secondary_sales/0"
                    data=requests.get(url, headers=hashers)
                    total_data=data.json()['data']['distributer_sale']
                    all_data.append(total_data)
                else:
                    # url = 'https://uat.lakshya.rsr.cloware.com/secondarysync/secondary_sales/'+str(i+1)
                    url = "https://drop52.reports.cloware.com/secondarysync/secondary_sales/"+str(i+1)
                    data=requests.get(url, headers=hashers)
                    total_data=data.json()['data']['distributer_sale']
                    all_data.append(total_data)
                s=s+1
            final_data=list(chain(*all_data))
            json_object = json.dumps(all_data, indent = 4)
            # n = datetime.datetime.now().date()-datetime.timedelta(days=1)
            
            with open(settings.MEDIA_ROOT + 'sfa_api_sales/'+"sfa_sales_"+str(datetime.datetime.now().date()) + '.json','w')as outfile:
                outfile.write(json_object)
            if json_object[0]:
                response = HttpResponse(content_type='text/csv')
                df = pd.DataFrame.from_dict(list(final_data))
                df.to_csv(settings.MEDIA_ROOT + 'sfa_api_sales_csv/'+"sfa_sales_"+str(datetime.datetime.now().date()) + '.csv', index = False)
            # =========testcase===========================================
            # from master import sfa
            # success_count = sfa.aa['data']['success_count']
            # final_data = sfa.aa['data']['distributer_sale']
            # print(len(sfa.aa['data']['distributer_sale']),"==========",success_count)
            # breakpoint()
            a = 0
            # ===========================end test case====================
            # from .sfa_test_json import sfa_test
            today_date = datetime.datetime.now()
            if int(success_count)>0:
                # trans_obj=Integration_log_summary.objects.filter(sale_date=final_data[0].get('sale_date')).last()
                Integration_log_summary.objects.all().delete()
                Integration_log_details.objects.all().delete()
                # for json_data in sfa_test: #final_data
                for json_data in final_data: #final_data
                    a = a+1
                    print(a,"=====================================================================================//==")
                    sku_obj=SKU_Master_Product.objects.filter(sku_code__iexact=json_data['prodcode']).last()
                    wd_id = json_data['distrcode'].split("-")[0]
                    
                    town_trim_wd = str(json_data['distrcode']).replace(json_data['town_code'],'')
                    user_id = User.objects.filter(Q(user_id = wd_id)|Q(user_id = town_trim_wd)).last()
                    town_cd.append(json_data['town_code'])
                    # wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(wd_id=wd_id)|Q(wd_id = town_trim_wd)),(Q(town_code=json_data['town_code'])|Q(town_code__iexact=json_data['town_code'])|Q(town_code__iexact=str(json_data['town_code'])[1:])|Q(town_code__iexact='0'+str(json_data['town_code'])))).values_list('wd_town_id',flat=True)
                    wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(wd_id=wd_id)|Q(wd_id = town_trim_wd)),town_code__in=[json_data['town_code'],re.sub('\W+','',json_data['town_code']),str(json_data['town_code']),str(json_data['town_code'])[1:],'0'+str(json_data['town_code'])]).values_list('wd_town_id',flat=True)
                    print(wd_town_list,"====",[json_data['town_code'],re.sub('\W+','',json_data['town_code']),str(json_data['town_code']),str(json_data['town_code'])[1:],'0'+str(json_data['town_code'])])
                    # context = {"data2":wd_town_list,"data":[json_data['town_code'],re.sub('\W+','',json_data['town_code']),str(json_data['town_code']),str(json_data['town_code'])[1:],'0'+str(json_data['town_code'])]}
                    # return Response(context)
                    wdobj = WdSkuCatagory.objects.filter(sku_code__iexact=json_data['prodcode'],wd_town_id__in=wd_town_list).last()
                    sku = []
                    wd = []
                    town = []
                    unmap_sku = []
                    if not user_id:
                        wd.append(wd_id)
                    elif not wd_town_list:
                        town.append(json_data['town_code'])
                    elif not wdobj:
                        unmap_sku.append(json_data['prodcode'])
                    elif not sku_obj:
                        sku.append(json_data['prodcode'])
                    total_sale = (float(json_data.get('local_retail', 0))+float(json_data.get('local_dealer', 0))+float(json_data.get('local_MT', 0))+float(json_data.get('local_HA', 0))
                                +float(json_data.get('out_retail', 0))+float(json_data.get('out_dealer', 0))+float(json_data.get('out_MT', 0))+float(json_data.get('out_HA', 0))
                                +float(json_data.get('other_retail', 0))+float(json_data.get('other_dealer', 0))+float(json_data.get('other_MT', 0)))
                    
                    if sku_obj and wdobj and user_id and (total_sale > 0):
                        success_list = success_list + 1
                        wd_list_4_week.append(user_id.user_id)
                        
                        wds = WDmaster.objects.filter(wd_ids = user_id.user_id).last()
                        # town_nm = Sales_Hierarchy_Master.objects.filter((Q(wd_id = user_id.user_id)|Q(wd_id = town_trim_wd)), (Q(town_code=json_data['town_code'])|Q(town_code__iexact=json_data['town_code'])| Q(town_code__iexact=str(json_data['town_code'])[1:])| Q(town_code__iexact='0'+str(json_data['town_code'])))).last()
                        town_nm = wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(wd_id=wd_id)|Q(wd_id = town_trim_wd)),town_code__in=[json_data['town_code'],re.sub('\W+','',json_data['town_code']),str(json_data['town_code']),str(json_data['town_code'])[1:],'0'+str(json_data['town_code'])]).last()
                        if town_nm:
                            town_name = town_nm.town
                        else:
                            town_name = " "

                        if wds:
                            wd_id_name = wds.wd_name
                        else:
                            wd_id_name = "Name not Avail in wd_master"
                            
                        json_data['distrcode'] = str(user_id.user_id)+"-"+json_data['town_code']
                        json_data['town_code'] = town_nm.town_code if town_nm else json_data['town_code']
                        json_data['sku_short_name'] = sku_obj.sku_short_name
                        json_data['sku_code'] = json_data['prodcode']
                        json_data['town_name'] = town_name

                        json_data['wd_name'] = wd_id_name
                        json_data['wd_type'] = json_data['dist_type']
                        
                        json_data['sku_id'] = sku_obj.sku_id
                        json_data['town_id'] = wdobj.wd_town_id
                        json_data['wd_id'] = user_id.user_id
                        json_data['tranisition_source'] = json_data['dist_type']
                        json_data['transaction_type'] = 'DAILY'
                        json_data['created_by'] = "SFA/SFA_LITE_API"
                        json_data['brand_category'] = json_data.get('catcode')
                        # json_data['created_date'] = datetime.datetime.now()
                        json_data['sales_date_time'] = json_data.get('sale_date')
                        json_data['local_sales_retail'] = json_data.get('local_retail', 0)
                        json_data['local_sales_dealer'] = json_data.get('local_dealer', 0)
                        json_data['local_sales_modern_trade'] = json_data.get('local_MT', 0)
                        json_data['local_sales_hawker'] = json_data.get('local_HA', 0)
                        json_data['outstation_sales_reatil'] = json_data.get('out_retail', 0)
                        json_data['outstation_sales_dealer'] = json_data.get('out_dealer', 0)
                        json_data['outstation_sales_modern_trade'] = json_data.get('out_MT', 0)
                        json_data['outstation_sales_hawker'] = json_data.get('out_HA', 0)
                        json_data['other_sales_retail'] = json_data.get('other_retail', 0)
                        json_data['other_sales_dealer'] = json_data.get('other_dealer', 0)
                        json_data['other_sales_modern_trade'] = json_data.get('other_MT', 0)
                        json_data['other_issues_other'] = json_data.get('other_issues_other', 0)
                        json_data['other_issues_damage'] = json_data.get('other_issued_damage', 0)
                        json_data['other_issues_return'] = json_data.get('other_issued_returns', 0)
                        # json_data['last_updated_date'] = datetime.datetime.now()
                        json_data['total_local_sales'] = float(json_data['local_sales_retail']) + float(json_data['local_sales_dealer']) +float(json_data['local_sales_modern_trade'])+float(json_data['local_sales_hawker'])
                        json_data['total_outstation_sales'] = float(json_data['outstation_sales_reatil']) +float(json_data['outstation_sales_dealer']) +float(json_data['outstation_sales_modern_trade']) +float(json_data['outstation_sales_hawker'])
                        json_data['total_other_sales'] = float(json_data['other_sales_retail']) +float(json_data['other_sales_dealer']) +float(json_data['other_sales_modern_trade'])
                        json_data['total_issue'] = float(json_data['other_issues_other']) +float(json_data['other_issues_damage']) +float(json_data['other_issues_return'])
                        json_data['grand_total'] = json_data['total_local_sales'] +json_data['total_outstation_sales'] +json_data['total_other_sales']
                        if sku_obj:
                            json_data['company'] = sku_obj.company
                        else:
                            json_data['company'] =None
                        if wdobj:
                            json_data['unit_price'] = wdobj.last_price
                            json_data['cnf_id'] = wdobj.cnf_id
                        else:
                            json_data['unit_price'] = 0
                            json_data['cnf_id'] = None
                        branch = User.objects.filter(user_id = json_data.get('wd_id'), user_type = "WD").last()
                        if branch:
                            region = BranchMaster.objects.filter(branch_code = branch.locationcode).last()
                            json_data['region'] = region.region
                        else:
                            json_data['region'] = None

                        json_data['value'] = float(json_data['unit_price']) * float(json_data['grand_total'])
                        sale_data = SalesData.objects.filter(wd_id = json_data['wd_id'],sku_id = json_data['sku_id'],sales_date_time=json_data['sales_date_time'],town_id = json_data['town_id'],town_code = json_data['town_code'],status = True).last()
                        sale_data_ss = SalesData.objects.filter(wd_id = json_data['wd_id'],sku_id = json_data['sku_id'],sales_date_time=json_data['sales_date_time'],town_id = json_data['town_id'],town_code = json_data['town_code'],status = False).last()
                        # print(sale_data,"======")
                        if sale_data_ss :

                            # json_data['town_code'] = str(sale_data.town_code)
                            # json_data['sku_short_name'] = sku_obj.sku_short_name
                            # json_data['sku_code'] = json_data['prodcode']

                            # json_data['wd_name'] = wd_id_name
                            # json_data['wd_type'] = json_data['dist_type']

                            json_data['local_sales_retail'] = float(json_data['local_sales_retail']) + sale_data_ss.local_sales_retail
                            json_data['local_sales_dealer'] = float(json_data['local_sales_dealer']) + sale_data_ss.local_sales_dealer 
                            json_data['local_sales_modern_trade'] = float(json_data['local_sales_modern_trade']) + sale_data_ss.local_sales_modern_trade
                            json_data['local_sales_hawker'] = float(json_data['local_sales_hawker']) + sale_data_ss.local_sales_hawker
                            json_data['total_local_sales'] = float(json_data['local_sales_retail']) + json_data['local_sales_dealer'] +json_data['local_sales_modern_trade']+json_data['local_sales_hawker']

                            json_data['outstation_sales_reatil'] = float(json_data['outstation_sales_reatil']) + sale_data_ss.outstation_sales_reatil
                            json_data['outstation_sales_dealer'] = float(json_data['outstation_sales_dealer']) + sale_data_ss.outstation_sales_dealer
                            json_data['outstation_sales_modern_trade'] = float(json_data['outstation_sales_modern_trade']) + sale_data_ss.outstation_sales_modern_trade
                            json_data['outstation_sales_hawker'] = float(json_data['outstation_sales_hawker']) + sale_data_ss.outstation_sales_hawker
                            json_data['total_outstation_sales'] = float(json_data['outstation_sales_reatil']) + json_data['outstation_sales_dealer'] + json_data['outstation_sales_modern_trade'] + json_data['outstation_sales_hawker']

                            json_data['other_sales_retail'] = float(json_data['other_sales_retail']) + sale_data_ss.other_sales_retail
                            json_data['other_sales_dealer'] = float(json_data['other_sales_dealer']) + sale_data_ss.other_sales_dealer
                            json_data['other_sales_modern_trade'] = float(json_data['other_sales_modern_trade']) + sale_data_ss.other_sales_modern_trade
                            json_data['total_other_sales'] = float(json_data['other_sales_retail']) + json_data['other_sales_dealer'] + json_data['other_sales_modern_trade']

                            json_data['other_issues_damage'] = float(json_data['other_issues_damage']) + sale_data_ss.other_issues_damage
                            json_data['other_issues_return'] = float(json_data['other_issues_return']) + sale_data_ss.other_issues_return
                            json_data['other_issues_other'] = float(json_data['other_issues_other']) + sale_data_ss.other_issues_other
                            json_data['total_issue'] = json_data['other_issues_damage'] + json_data['other_issues_return']+json_data['other_issues_other']

                            json_data['grand_total'] = json_data['total_local_sales'] +json_data['total_outstation_sales'] +json_data['total_other_sales']
                            json_data['value'] = float(json_data['unit_price']) * float(json_data['grand_total'])

                            if wdobj:
                                json_data['unit_price'] = wdobj.last_price
                                json_data['cnf_id'] = wdobj.cnf_id
                            else:
                                json_data['unit_price'] = 0
                                json_data['cnf_id'] = None
                            json_data.pop('created_by')
                            json_data.pop('sales_date_time')
                            # json_data.pop('created_date')
                            json_data['last_updated'] = "SFA/SFA_LITE_API"
                            json_data['last_updated_date'] = datetime.datetime.now()
                            serializer = SalesDataSerializer(sale_data_ss, data = json_data,partial=True)
                            if serializer.is_valid():
                                serializer.save()
                                repet_flag = repet_flag + 1
                                repeat_sl = {}
                                # print(wd_id,json_data['prodcode'],json_data['town_code'])
                                repeat_sale = Repeat_count.objects.filter(wd_id = json_data['wd_id'], sku_id = json_data['prodcode'],sale_date_time = json_data.get('sale_date'),transaction_source = "SFA/SFA_LITE_API").last()
                                # print(repeat_sale,"=======repeat_sale=====")
                                repeat_sl['wd_id'] = json_data['wd_id']
                                repeat_sl['sku_id'] = json_data['prodcode']
                                repeat_sl['town_id'] = json_data['town_code']
                                repeat_sl['sale_date_time'] = json_data.get('sale_date')
                                repeat_sl['transaction_source'] = "SFA/SFA_LITE_API"
                                if repeat_sale:
                                    repeat_sl['repeat_time'] = int(repeat_sale.repeat_time)+1
                                    repeat_sl['town_id'] = str(repeat_sale.town_id)+"-"+str(json_data['town_code'])
                                    print(repeat_sl,"=======repeat_sl=======")
                                    serializer = Repeat_countSerializers(repeat_sale, data = repeat_sl,partial =True)
                                    # print(repeat_sl['repeat_time'],"===repeat_sl['repeat_time']===")
                                    if serializer.is_valid():
                                        serializer.save()
                                        
                                    else:
                                        return Response(serializer.errors)
                                else:
                                    print("===2===")
                                    repeat_sl['repeat_time'] = int(1)
                                    json_data['created_date'] = datetime.datetime.now()
                                    serializer = Repeat_countSerializers(data = repeat_sl)
                                    if serializer.is_valid():
                                        serializer.save()
                        if sale_data:
                            print("=================if exist update ========================")
                            json_data['status'] = False
                            # json_data['status_3rd_party'] = True
                            json_data['last_updated_date'] = datetime.datetime.now()
                            serializer = SalesDataSerializer(sale_data,json_data,partial =True)
                            # print(repeat_sl['repeat_time'],"===repeat_sl['repeat_time']===")
                            if serializer.is_valid():
                                serializer.save()
                                
                            else:
                                print(serializer.errors,"======error")
                                errors_list.append(serializer.errors)    
                        else:
                            json_data['status'] = False
                            json_data['created_date'] = datetime.datetime.now()
                            serializer = SalesDataSerializer(data = json_data)
                            if serializer.is_valid():
                                serializer.save()
                                flag=flag+1
                            else:
                                errors_list.append(serializer.errors)
                                
                        count_success_sale_date_wise(today_date, json_data.get('sale_date'))

                        
                    else:
                        count_invalid_sale_date_wise(today_date, json_data.get('sale_date'))
                        log_list = log_list+1
                        print("=======00=======")
                        json_data['transaction_source'] = "SFA/SFA_LITE_API"
                        log_data_serializer = Invalid_log_dataSerializers(data = json_data)
                        if log_data_serializer.is_valid():
                            log_data_serializer.save()
                            
                            
                        # breakpoint()
                        len_val=len_val+1
                        errors_list.append(json_data)
                        const_time = datetime.datetime(2009, 10, 5, 00, 00, 00)
                        # wdid_town = json_data['distrcode'].split("-")
                        error_detail_dict['sku_id'] = json_data['prodcode']
                        error_detail_dict['wd_id'] = json_data['dist_id']
                        error_detail_dict['town_id'] = json_data['town_code']
                    #     sku.append(json_data['prodcode'])
                    # if not user_id:
                    #     wd.append(wd_id)
                    # if not wdobj:
                    #     town.append(json_data['town_code'])
                        if total_sale == 0:
                            error_detail_dict['reason'] = str(wd_id) +" "+" Received '0' sales for SKU "+ str(json_data['prodcode'])
                        elif len(wd)>0:
                            error_detail_dict['reason'] = str(wd_id) +" "+" wd_id is not present in user_master database."
                            print("===========ss",str(wd_id) +" "+" wd_id is not present in user_master database.")
                            # error_detail_dict['repeat_count'] = len_val
                        elif len(town)>0:
                            error_detail_dict['reason'] = str(wd_id)+ " WD_id is not mapped with " + str(json_data['town_code'])+" town code."
                            print("----------------------rrrr",str(wd_id)+ " WD_id is not mapped with " + str(json_data['town_code'])+" town code.")
                            
                        elif len(unmap_sku) > 0:
                            error_detail_dict['reason'] = str(json_data['prodcode']) +" sku_code is not mapped with "+ str(wd_id) +" wd_id in SS_database:Sales_Hierarchy_Master"
                            print("===========ss",str(json_data['prodcode']) +" sku_code is not mapped with "+ str(wd_id) +" wd_id in SS_database:Sales_Hierarchy_Master")
                            
                        elif len(sku)>0:
                            error_detail_dict['reason'] = str(json_data['prodcode']) +" "+ " sku_code is not present in sku_master database: sku_master"
                            print("===============88",str(json_data['prodcode']) +" "+ " sku_code is not present in sku_master database: sku_master")
                            # error_detail_dict['repeat_count'] = len_val
                        else:
                            # error_detail_dict['repeat_count'] = len_val
                            error_detail_dict['reason'] = str(json_data['town_code']) +"This towm_code is not mapped with any WD"

                        error_detail_dict['tranisition_source'] = "SFA/SFA_LITE_API"
                        # comb_d_t = datetime.datetime.combine(parse(final_data[0].get('sale_date')), const_time.time())
                        comb_d_t = parse(json_data['sale_date'])
                        # print(comb_d_t)
                        error_detail_dict['sales_date_time'] = comb_d_t
                        log_detail = Integration_log_details.objects.filter(tranisition_source = "SFA/SFA_LITE_API",sales_date_time = comb_d_t, sku_id = json_data['prodcode'],town_id = json_data['town_code']).last()
                        # ,town_id = json_data['town_code']
                        # print(log_detail,"======log_detail.repeat_count====")
                        # print(log_detail.repeat_count,"====log_detail.repeat_count====")
                        if log_detail:
                            error_detail_dict['created_date'] = today_date
                            # error_detail_dict['repeat_count'] = int(log_detail.repeat_count) +1
                            log_serializer = Integration_log_detailsSerializers(log_detail, data = error_detail_dict, partial=True)
                            
                            if log_serializer.is_valid():
                                
                                log_serializer.save()
                        else:
                            # print(error_detail_dict,"======error_detail_dict======")
                            error_detail_dict['repeat_count'] = 1
                            error_detail_dict['created_date'] = today_date
                            log_serializer = Integration_log_detailsSerializers(data = error_detail_dict)
                            
                            if log_serializer.is_valid():
                                
                                log_serializer.save()
                            else:
                                return Response(log_serializer.errors)
                    
                SalesData.objects.filter(wd_id__in = wd_list_4_week).update(status=True)
                    # Summary=============
                # add_sale_summary = Integration_log_summary.objects.filter(sale_date = final_data[0].get('sale_date'),tranisition_source = "SFA/SFA_LITE_API").last()
                # histry_dict["total_distributer_sale"] = int(success_count)
                # histry_dict["created_by"] = 'SFA/SFA_LITE_API'
                # histry_dict["total_insart_sale"]= int(success_count) - int(log_list)
                # histry_dict["sale_date"]=final_data[0].get('sale_date')
                # histry_dict['created_date'] = datetime.datetime.now()
                # histry_dict['tranisition_source'] = 'SFA/SFA_LITE_API'
                # if add_sale_summary:
                #     # print(histry_dict,'====histry_dict ===')
                #     historyserializers = Integration_log_summarySerializers(add_sale_summary, data=histry_dict, partial = True)
                #     if historyserializers.is_valid():
                #         historyserializers.save()
                #     else:
                #         return Response(historyserializers.errors)
                #     if flag!=len(success_count):
                #         print("send mail......",int(success_count) - int(len(errors_list)))
                # else:
                #     histry_dict.pop('created_date')
                #     histry_dict['last_updated_date'] = datetime.datetime.now()
                #     historyserializers = Integration_log_summarySerializers(data=histry_dict)
                #     if historyserializers.is_valid():
                #         historyserializers.save()
                #     if flag!=len(success_count):
                #         print("send mail......",int(success_count) - int(len(errors_list)))
                
                trans_obj = Integration_log_summary.objects.filter(created_date__date=today_date.date(),tranisition_source = "SFA/SFA_LITE_API" ).values("total_distributer_sale","total_insart_sale","sale_date")
                log_detail = Integration_log_details.objects.filter(created_date__date = today_date.date(), tranisition_source = "SFA/SFA_LITE_API").values("sku_id","wd_id","town_id","sales_date_time","reason")
                print(trans_obj,"======trans_obj==========mail=====")
                
                row_no = 0
                # today = utils.now_date
                today = datetime.datetime.now()
                now_day=str(today.date())
                splited_today = now_day.split('-')[2]
                week_details = week_sele_before_date('BRANCH USER')
                if int(splited_today) in date_tuple_weekly and week_details[2]:
                    print(wd_list_4_week,"==============wd_list_4_week==============")
                    create_update_weekly(wd_list_4_week)
                    # weekly_generate_on_weekend(wd_list_4_week)
                success_count = success_count if success_count else 0
                count_dist_sale = 0
                insert_sale = 0
                if trans_obj:
                    print(trans_obj,"======================================================mail============")
                    # subject = 'UAT Server| SFA/SFA LITE Integration Status Report.'
                    subject = sfa_success_sub
                    html_message = '<table style="width:100% ; border: 1px solid black;">'
                            
                    html_message +=            '<tr style="border: 1px solid black;">'
                    html_message +=                '<th "border: 1px solid black;">SNo.</th>'
                    html_message +=                '<th "border: 1px solid black;">Sale Date</th>'
                    html_message +=                '<th style="border: 1px solid black;">Total distributer sale</th>'
                    html_message +=                '<th style="border: 1px solid black;">Total insert sale</th>'
                    # html_message +=                '<th>town_id</th>'
                    # html_message +=                '<th>reason</th>'
                    html_message +=            '</tr>'
                    for ils in trans_obj:
                        row_no = row_no+1
                        # print(ils,"=================")
                        html_message +=            '<tr style="border: 1px solid red;">'
                        html_message +=                '<td style="border: 1px solid black;">'+str(row_no)+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['sale_date'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_distributer_sale'])+'</td>'
                        count_dist_sale = count_dist_sale + ils['total_distributer_sale']
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_insart_sale'])+'</td>'
                        insert_sale = insert_sale + ils['total_insart_sale']
                        
                        html_message +=            '</tr>'
                        row_no+=1
                    html_message +=            '<tr>'
                    html_message +=                '<td style="border: 1px solid black;">'+str(count_dist_sale)+'</td>'
                    html_message +=                '<td style="border: 1px solid black;">'+str("Total")+'</td>'
                    html_message +=                '<td style="border: 1px solid black;">'+str(count_dist_sale)+'</td>'
                    html_message +=                '<td style="border: 1px solid black;">'+str(insert_sale)+'</td>'
                    html_message +=            '</tr>'
                    html_message += '</table> </br></br>'



                    # log details==============
                    html_message += '<table style="width:100% ; border: 1px solid black;">'
                            
                    html_message +=            '<tr style="border: 1px solid black;">'
                    html_message +=                '<th style="border: 1px solid black;">SNo.</th>'
                    html_message +=                '<th style="border: 1px solid black;">Sale Date</th>'
                    html_message +=                '<th style="border: 1px solid black;">Sku id</th>'
                    html_message +=                '<th style="border: 1px solid black;">Wd id</th>'
                    html_message +=                '<th style="border: 1px solid black;">Town id</th>'
                    html_message +=                '<th style="border: 1px solid black;">Reason</th>'
                    # html_message +=                '<th style="border: 1px solid black;">Repeat Count</th>'
                    html_message +=            '</tr>'
                    for ld in log_detail:
                        
                        row_no+=1
                        # print(ld['sales_date_time'].date(),"==========date=======")
                        html_message +=            '<tr style="border: 1px solid red;">'
                        html_message +=                '<td style="border: 1px solid black;">'+str(row_no)+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['sales_date_time'].date())+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['sku_id'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['wd_id'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['town_id'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['reason'])+'</td>'
                        # html_message +=                '<td style="border: 1px solid black;">'+str(ld['repeat_count'])+'</td>'
                        html_message +=            '</tr>'
                    html_message += '</table> </br></br>'

                    
                    email_from = settings.EMAIL_HOST_USER
                    cc_email = cc_list
                    # ,'saurabha@triazinesoft.com','mayankg@triazinesoft.com'
                    recipient_list = success_email_list

                    reset_email = EmailMessage(
                                subject = subject,
                                body = html_message,
                                from_email = email_from,
                                to = recipient_list,
                                cc = cc_email,
                                reply_to = cc_email,
                                )
                    trans_obj=Integration_log_summary.objects.filter(sale_date=final_data[0].get('sale_date'),email_status = False).update(email_status = True)
                    reset_email.content_subtype = "html"
                    reset_email.send(fail_silently=True)
                # return Response({"sku":sku,"wd":wd,"town":town})
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong','error':error}
            # subject = "UAT | SFA/SFA_lite Issue"
            subject = sfa_failed_sub
            recipient_list = failed_list
            cc_email = cc_list
            # recipient_list = ["VIPULSHARMA-gpi@modi-ent.com","ankit.rakwal@esobene.com"]
            html_message = str(context)
            reset_email = EmailMessage(
                                subject = subject,
                                body = html_message,
                                from_email = settings.EMAIL_HOST_USER,
                                to = recipient_list,
                                cc = cc_email,
                                reply_to = cc_email,
                                )
            reset_email.content_subtype = "html"
            reset_email.send(fail_silently=True)
            return Response(context, status=status.HTTP_200_OK)


from django.db.models import Count          
class unique_log(APIView):
    def get(self, request):
        from_date = request.GET.get("from_dt",None)
        to_date = request.GET.get("to_td",None)
        # print()
        # value = Integration_log_details.objects.filter(sales_date_time__date =  date).values('sku_id','wd_id','town_id')
        try:
            value_list = Integration_log_details.objects.filter(sales_date_time__date__gte = from_date,sales_date_time__date__lte = to_date).values_list('reason',flat=True).distinct()
            # value = Integration_log_details.objects.filter(sales_date_time__date__gte = from_date,sales_date_time__date__lte = to_date,reason__in = value_list).values('reason').distinct()

            
            # 'sku_id','wd_id','town_id',
            # value = Integration_log_details.objects.all().distinct('reason')   'sales_date_time','tranisition_source'
            # print(value[0])
            # print(len(value))
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="log_dtail.csv"'
                
            writer = csv.DictWriter(response,fieldnames=[
                                'S No.','WD ID','SKU ID','Town','REGION', 'Sale Date Time','tranisition_source'                                       
                                ])
            writer = writer.writeheader()
            snum = 0 
            for val in value_list:
                print(val,"======value====")
                value = Integration_log_details.objects.filter(sales_date_time__date__gte = from_date,sales_date_time__date__lte = to_date,reason= val).last()
                print(value)
                snum+=1
                writer = csv.writer(response)
                writer.writerow([
                    snum,
                    value.wd_id,
                    value.sku_id,
                    value.town_id,
                    value.reason,
                    value.sales_date_time,
                    value.tranisition_source,

                ])
            return response
            # return Response({"logs":value})
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong','error':error}
            return Response(context, status=status.HTTP_200_OK)

from django.views.decorators.csrf import csrf_exempt
class SFA_API_read_json_file(APIView):
    @csrf_exempt
    def post(self, request):
        try:
            # with open(settings.MEDIA_ROOT + 'sfa_api_sales/'+"sfa_sales_"+str(datetime.datetime.now().date()) + '.json','w')as outfile:
            #     outfile.write(json_object)
            s=0
            all_data=[]
            errors_list=[]
            flag=0
            repet_flag = 0
            histry_dict={}
            error_detail_dict = {}
            len_val = 0
            town_cd =[]
            success_list = 0
            # json_file = open(settings.MEDIA_ROOT + 'sfa_api_sales/'+"sfa_sales_"+str('2022-05-31') + '.json')
            # final_data_wc = json.load(json_file)
            # final_data_wc = request.data
            final_data=request.data
            print(len(final_data),"======len======")
            # breakpoint()
            wd_list_4_week = []
            log_list = 0
                
            # breakpoint()
            success_count = len(final_data)
            today_date = datetime.datetime.now()
            date_tuple_weekly = [8,15,22,1]
            a = 0
            if int(success_count)>0:
                trans_obj=Integration_log_summary.objects.filter(sale_date=final_data[0].get('sale_date')).last()
                temp_sale_data_objs = Temp_Total_sku_sales.objects.all()
                print(temp_sale_data_objs)
                temp_sale_data_objs.delete()
                Integration_log_summary.objects.all().delete()
                Integration_log_details.objects.all().delete()
                # breakpoint()
                for json_data in final_data:
                    a = a+1
                    print(a,"===================================>>>>>55<<<<===============================//==")
                    sku_obj=SKU_Master_Product.objects.filter(sku_code__iexact=json_data['prodcode']).last()
                    wd_id = json_data['distrcode'].split("-")[0]
                    
                    town_trim_wd = str(json_data['distrcode']).replace(json_data['town_code'],'')
                    user_id = User.objects.filter(Q(user_id = wd_id)|Q(user_id = town_trim_wd)).last()
                    town_cd.append(json_data['town_code'])
                    wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(wd_id=wd_id)|Q(wd_id = town_trim_wd)),town_code__in=[json_data['town_code'],re.sub('\W+','',json_data['town_code']),str(json_data['town_code']),str(json_data['town_code'])[1:],'0'+str(json_data['town_code'])]).values_list('wd_town_id',flat=True)
                    sku_obj=SKU_Master_Product.objects.filter(sku_code__iexact=json_data['prodcode']).last()
                    wd_id = json_data['distrcode'].split("-")[0]
                    
                    town_trim_wd = str(json_data['distrcode']).replace(json_data['town_code'],'')
                    user_id = User.objects.filter(Q(user_id = wd_id)|Q(user_id = town_trim_wd)).last()
                    town_cd.append(json_data['town_code'])
                    # wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(wd_id=wd_id)|Q(wd_id = town_trim_wd)),(Q(town_code=json_data['town_code'])|Q(town_code__iexact=json_data['town_code'])|Q(town_code__iexact=str(json_data['town_code'])[1:])|Q(town_code__iexact='0'+str(json_data['town_code'])))).values_list('wd_town_id',flat=True)
                    wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(wd_id=wd_id)|Q(wd_id = town_trim_wd)),town_code__in=[json_data['town_code'],re.sub('\W+','',json_data['town_code']),str(json_data['town_code']),str(json_data['town_code'])[1:],'0'+str(json_data['town_code'])]).values_list('wd_town_id',flat=True)
                    wdobj = WdSkuCatagory.objects.filter(sku_code__iexact=json_data['prodcode'],wd_town_id__in=wd_town_list).last()
                    sku = []
                    wd = []
                    town = []
                    unmap_sku = []
                    if not user_id:
                        wd.append(wd_id)
                    elif not wd_town_list:
                        town.append(json_data['town_code'])
                    elif not wdobj:
                        unmap_sku.append(json_data['prodcode'])
                    elif not sku_obj:
                        sku.append(json_data['prodcode'])
                    total_sale = (float(json_data.get('local_retail', 0))+float(json_data.get('local_dealer', 0))+float(json_data.get('local_MT', 0))+float(json_data.get('local_HA', 0))
                                +float(json_data.get('out_retail', 0))+float(json_data.get('out_dealer', 0))+float(json_data.get('out_MT', 0))+float(json_data.get('out_HA', 0))
                                +float(json_data.get('other_retail', 0))+float(json_data.get('other_dealer', 0))+float(json_data.get('other_MT', 0)))
                    if sku_obj and wdobj and user_id and (total_sale > 0):
                        success_list = success_list + 1
                        wd_list_4_week.append(user_id.user_id)
                        
                        wds = WDmaster.objects.filter(wd_ids = user_id.user_id).last()
                        # town_nm = Sales_Hierarchy_Master.objects.filter((Q(wd_id = user_id.user_id)|Q(wd_id = town_trim_wd)),(Q(town_code__iexact=json_data['town_code'])|Q(town_code__iexact = str(json_data['town_code'])[1:])|Q(town_code__iexact='0'+str(json_data['town_code'])))).last()
                        town_nm = wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(wd_id=wd_id)|Q(wd_id = town_trim_wd)),town_code__in=[json_data['town_code'],re.sub('\W+','',json_data['town_code']),str(json_data['town_code']),str(json_data['town_code'])[1:],'0'+str(json_data['town_code'])]).last()
                        if town_nm:
                            town_name = town_nm.town
                        else:
                            town_name = " "

                        if wds:
                            wd_type_master = wds.wd_type
                            wd_id_name = wds.wd_name
                        else:
                            wd_id_name = "Name not Avail in wd_master"
                            
                        json_data['distrcode'] = str(user_id.user_id)+"-"+json_data['town_code']
                        json_data['town_code'] = json_data['town_code']
                        json_data['sku_short_name'] = sku_obj.sku_short_name
                        json_data['sku_code'] = json_data['prodcode']
                        json_data['town_name'] = town_name

                        json_data['wd_name'] = wd_id_name
                        json_data['wd_type'] = wd_type_master
                        
                        json_data['sku_id'] = sku_obj.sku_id
                        json_data['town_id'] = wdobj.wd_town_id
                        json_data['wd_id'] = user_id.user_id
                        json_data['tranisition_source'] = json_data['dist_type']
                        json_data['transaction_type'] = 'DAILY'
                        json_data['created_by'] = "SFA/SFA_LITE_API"
                        json_data['brand_category'] = json_data.get('catcode')
                        json_data['created_date'] = datetime.datetime.now()
                        json_data['sales_date_time'] = json_data.get('sale_date')
                        json_data['local_sales_retail'] = json_data.get('local_retail', 0)
                        json_data['local_sales_dealer'] = json_data.get('local_dealer', 0)
                        json_data['local_sales_modern_trade'] = json_data.get('local_MT', 0)
                        json_data['local_sales_hawker'] = json_data.get('local_HA', 0)
                        json_data['outstation_sales_reatil'] = json_data.get('out_retail', 0)
                        json_data['outstation_sales_dealer'] = json_data.get('out_dealer', 0)
                        json_data['outstation_sales_modern_trade'] = json_data.get('out_MT', 0)
                        json_data['outstation_sales_hawker'] = json_data.get('out_HA', 0)
                        json_data['other_sales_retail'] = json_data.get('other_retail', 0)
                        json_data['other_sales_dealer'] = json_data.get('other_dealer', 0)
                        json_data['other_sales_modern_trade'] = json_data.get('other_MT', 0)
                        json_data['other_issues_other'] = json_data.get('other_issues_other', 0)
                        json_data['other_issues_damage'] = json_data.get('other_issued_damage', 0)
                        json_data['other_issues_return'] = json_data.get('other_issued_returns', 0)
                        # json_data['last_updated_date'] = datetime.datetime.now()
                        json_data['total_local_sales'] = float(json_data['local_sales_retail']) + float(json_data['local_sales_dealer']) +float(json_data['local_sales_modern_trade'])+float(json_data['local_sales_hawker'])
                        json_data['total_outstation_sales'] = float(json_data['outstation_sales_reatil']) +float(json_data['outstation_sales_dealer']) +float(json_data['outstation_sales_modern_trade']) +float(json_data['outstation_sales_hawker'])
                        json_data['total_other_sales'] = float(json_data['other_sales_retail']) +float(json_data['other_sales_dealer']) +float(json_data['other_sales_modern_trade'])
                        json_data['total_issue'] = float(json_data['other_issues_other']) +float(json_data['other_issues_damage']) +float(json_data['other_issues_return'])
                        json_data['grand_total'] = json_data['total_local_sales'] +json_data['total_outstation_sales'] +json_data['total_other_sales']
                        if sku_obj:
                            json_data['company'] = sku_obj.company
                        else:
                            json_data['company'] =None
                        if wdobj:
                            json_data['unit_price'] = wdobj.last_price
                            json_data['cnf_id'] = wdobj.cnf_id
                        else:
                            json_data['unit_price'] = 0
                            json_data['cnf_id'] = None
                        branch = User.objects.filter(user_id = json_data.get('wd_id'), user_type = "WD").last()
                        if branch:
                            region = BranchMaster.objects.filter(branch_code = branch.locationcode).last()
                            json_data['region'] = region.region
                        else:
                            json_data['region'] = None

                        json_data['value'] = float(json_data['unit_price']) * float(json_data['grand_total'])
                        json_data['value'] = float(json_data['unit_price']) * float(json_data['grand_total'])
                        temp_sale_data = Temp_Total_sku_sales.objects.filter(wd_id = json_data['wd_id'],sku_id = json_data['sku_id'],transaction_type = 'DAILY',sales_date_time=json_data['sales_date_time'],town_id = json_data['town_id'],town_code = json_data['town_code']).last()
                        if not temp_sale_data:
                            serializer_temp_sale_data = Temp_Total_sku_salesSerializer(data = json_data, many = False)
                            if serializer_temp_sale_data.is_valid():
                                serializer_temp_sale_data.save()
                            else:
                                print(serializer_temp_sale_data.errors)
                                pass
                            print(total_sale,"===================>>createing===================sfa======",temp_sale_data)
                            # breakpoint()
                        else:
                            json_data['local_sales_retail'] = float(json_data['local_sales_retail']) + temp_sale_data.local_sales_retail
                            json_data['local_sales_dealer'] = float(json_data['local_sales_dealer']) + temp_sale_data.local_sales_dealer 
                            json_data['local_sales_modern_trade'] = float(json_data['local_sales_modern_trade']) + temp_sale_data.local_sales_modern_trade
                            json_data['local_sales_hawker'] = float(json_data['local_sales_hawker']) + temp_sale_data.local_sales_hawker
                            json_data['total_local_sales'] = float(json_data['local_sales_retail']) + json_data['local_sales_dealer'] +json_data['local_sales_modern_trade']+json_data['local_sales_hawker']

                            json_data['outstation_sales_reatil'] = float(json_data['outstation_sales_reatil']) + temp_sale_data.outstation_sales_reatil
                            json_data['outstation_sales_dealer'] = float(json_data['outstation_sales_dealer']) + temp_sale_data.outstation_sales_dealer
                            json_data['outstation_sales_modern_trade'] = float(json_data['outstation_sales_modern_trade']) + temp_sale_data.outstation_sales_modern_trade
                            json_data['outstation_sales_hawker'] = float(json_data['outstation_sales_hawker']) + temp_sale_data.outstation_sales_hawker
                            json_data['total_outstation_sales'] = float(json_data['outstation_sales_reatil']) + json_data['outstation_sales_dealer'] + json_data['outstation_sales_modern_trade'] + json_data['outstation_sales_hawker']

                            json_data['other_sales_retail'] = float(json_data['other_sales_retail']) + temp_sale_data.other_sales_retail
                            json_data['other_sales_dealer'] = float(json_data['other_sales_dealer']) + temp_sale_data.other_sales_dealer
                            json_data['other_sales_modern_trade'] = float(json_data['other_sales_modern_trade']) + temp_sale_data.other_sales_modern_trade
                            json_data['total_other_sales'] = float(json_data['other_sales_retail']) + json_data['other_sales_dealer'] + json_data['other_sales_modern_trade']

                            json_data['other_issues_damage'] = float(json_data['other_issues_damage']) + temp_sale_data.other_issues_damage
                            json_data['other_issues_return'] = float(json_data['other_issues_return']) + temp_sale_data.other_issues_return
                            json_data['other_issues_other'] = float(json_data['other_issues_other']) + temp_sale_data.other_issues_other
                            json_data['total_issue'] = json_data['other_issues_damage'] + json_data['other_issues_return']+json_data['other_issues_other']

                            json_data['grand_total'] = json_data['total_local_sales'] +json_data['total_outstation_sales'] +json_data['total_other_sales']
                            json_data['value'] = float(json_data['unit_price']) * float(json_data['grand_total'])

                            if wdobj:
                                json_data['unit_price'] = wdobj.last_price
                                json_data['cnf_id'] = wdobj.cnf_id
                            else:
                                json_data['unit_price'] = 0
                                json_data['cnf_id'] = None
                            json_data.pop('created_by')
                            json_data.pop('sales_date_time')
                            json_data.pop('created_date')
                            json_data['last_updated'] = "SFA/SFA_LITE_API"
                            json_data['last_updated_date'] = datetime.datetime.now()
                            serializer_temp_sale_data = Temp_Total_sku_salesSerializer(temp_sale_data,data = json_data, partial = True)
                            if serializer_temp_sale_data.is_valid():
                                serializer_temp_sale_data.save()
                        
                        count_success_sale_date_wise(today_date, json_data.get('sale_date'))
                        
                    else:
                        count_invalid_sale_date_wise(today_date, json_data.get('sale_date'))
                        
                        log_list = log_list+1
                        print("=======00=======")
                        json_data['transaction_source'] = "SFA/SFA_LITE_API"
                        log_data_serializer = Invalid_log_dataSerializers(data = json_data)
                        if log_data_serializer.is_valid():
                            log_data_serializer.save()
                            
                            
                        # breakpoint()
                        len_val=len_val+1
                        errors_list.append(json_data)
                        const_time = datetime.datetime(2009, 10, 5, 00, 00, 00)
                        # wdid_town = json_data['distrcode'].split("-")
                        error_detail_dict['sku_id'] = json_data['prodcode']
                        error_detail_dict['wd_id'] = json_data['dist_id']
                        error_detail_dict['town_id'] = json_data['town_code']

                        if total_sale == 0:
                            error_detail_dict['reason'] = str(wd_id) +" "+" Received '0' sales for SKU "+ str(json_data['prodcode'])
                        elif len(wd)>0:
                            error_detail_dict['reason'] = str(wd_id) +" "+" wd_id is not present in user_master database."
                            print("===========ss",str(wd_id) +" "+" wd_id is not present in user_master database.")
                            # error_detail_dict['repeat_count'] = len_val
                        elif len(town)>0:
                            error_detail_dict['reason'] = str(wd_id)+ " WD_id is not mapped with " + str(json_data['town_code'])+" town code."
                            print("----------------------rrrr",str(wd_id)+ " WD_id is not mapped with " + str(json_data['town_code'])+" town code.")
                            
                        elif len(unmap_sku) > 0:
                            error_detail_dict['reason'] = str(json_data['prodcode']) +" sku_code is not mapped with "+ str(wd_id) +" wd_id in SS_database:Sales_Hierarchy_Master"
                            print("===========ss",str(json_data['prodcode']) +" sku_code is not mapped with "+ str(wd_id) +" wd_id in SS_database:Sales_Hierarchy_Master")
                            
                        elif len(sku)>0:
                            error_detail_dict['reason'] = str(json_data['prodcode']) +" "+ " sku_code is not present in sku_master database: sku_master"
                            print("===============88",str(json_data['prodcode']) +" "+ " sku_code is not present in sku_master database: sku_master")
                            # error_detail_dict['repeat_count'] = len_val
                        else:
                            # error_detail_dict['repeat_count'] = len_val
                            error_detail_dict['reason'] = str(json_data['town_code']) +"This towm_code is not mapped with any WD"

                        error_detail_dict['tranisition_source'] = "SFA/SFA_LITE_API"
                        # comb_d_t = datetime.datetime.combine(parse(final_data[0].get('sale_date')), const_time.time())
                        comb_d_t = parse(json_data['sale_date'])
                        # print(comb_d_t)
                        error_detail_dict['sales_date_time'] = comb_d_t
                        log_detail = Integration_log_details.objects.filter(tranisition_source = "SFA/SFA_LITE_API",sales_date_time = comb_d_t, sku_id = json_data['prodcode'],town_id = json_data['town_code']).last()
                        # ,town_id = json_data['town_code']
                        # print(log_detail,"======log_detail.repeat_count====")
                        # print(log_detail.repeat_count,"====log_detail.repeat_count====")
                        if log_detail:
                            # error_detail_dict['repeat_count'] = int(log_detail.repeat_count) +1
                            log_serializer = Integration_log_detailsSerializers(log_detail, data = error_detail_dict, partial=True)
                            
                            if log_serializer.is_valid():
                                
                                log_serializer.save()
                        else:
                            # print(error_detail_dict,"======error_detail_dict======")
                            error_detail_dict['repeat_count'] = 1
                            error_detail_dict['created_date'] = today_date
                            log_serializer = Integration_log_detailsSerializers(data = error_detail_dict)
                            
                            if log_serializer.is_valid():
                                
                                log_serializer.save()
                            else:
                                return Response(log_serializer.errors)
                    
                from_temp_sale = Temp_Total_sku_sales.objects.all().values('brand_category','sku_id','wd_id','town_id','sales_date_time',
                                                                'local_sales_retail','local_sales_dealer','local_sales_modern_trade','local_sales_hawker','total_local_sales',
                                                                'outstation_sales_reatil','outstation_sales_dealer','outstation_sales_modern_trade','outstation_sales_hawker','total_outstation_sales',
                                                                'other_sales_retail','other_sales_dealer','other_sales_modern_trade','total_other_sales',
                                                                'other_issues_damage','other_issues_return','other_issues_other','total_issue',
                                                                'grand_total','created_by','last_updated','transaction_source','created_date','last_updated_date',
                                                                'status','freeze_status','transaction_type','company','unit_price','region','cnf_id','value',
                                                                'wd_name','wd_type','sku_code','sku_short_name','town_name','town_code','distrcode',
                                                            )

                if from_temp_sale:
                    for fts in from_temp_sale:
                        retive_cre_update = SalesData.objects.filter(wd_id = fts['wd_id'],sku_id = fts['sku_id'],sales_date_time=fts['sales_date_time'],transaction_type = 'DAILY',town_id = fts['town_id'],town_code = fts['town_code']).last()
                        if not retive_cre_update:
                            sales_data_serializer = SalesDataSerializer(data = fts, many = False)
                            if sales_data_serializer.is_valid():
                                sales_data_serializer.save()
                                # print(sales_data_serializer.data,"=========ccrr")
                            else:
                                print(sales_data_serializer.errors)
                        else:
                            sales_data_serializer = SalesDataSerializer(retive_cre_update,data = fts, partial = True)
                            if sales_data_serializer.is_valid():
                                sales_data_serializer.save()
                                # print(sales_data_serializer.data,"=========upup")
                                
                            else:
                                print(sales_data_serializer.errors)


                    
                SalesData.objects.filter(wd_id__in = wd_list_4_week).update(status=True)
                trans_obj = Integration_log_summary.objects.filter(created_date__date=today_date.date(),tranisition_source = "SFA/SFA_LITE_API" ).values("total_distributer_sale","total_insart_sale","sale_date")
                log_detail = Integration_log_details.objects.filter(created_date__date = today_date.date(), tranisition_source = "SFA/SFA_LITE_API").values("sku_id","wd_id","town_id","sales_date_time","reason")
                print(trans_obj,"======trans_obj==========mail=====")
                
                row_no = 0
                detail_row_no = 0
                # today = utils.now_date
                today = datetime.datetime.now()
                now_day=str(today.date())
                splited_today = now_day.split('-')[2]
                week_details = week_sele_before_date('BRANCH USER')
                if int(splited_today) in date_tuple_weekly and week_details[2]:
                    print(wd_list_4_week,"==============wd_list_4_week==============")
                    create_update_weekly(wd_list_4_week)
                count_dist_sale = 0
                insert_sale = 0
                if trans_obj:
                    print(trans_obj,"======================================================mail============")
                    subject = sfa_success_sub
                    html_message = '<table style="width:100% ; border: 1px solid black;">'
                            
                    html_message +=            '<tr style="border: 1px solid black;">'
                    html_message +=                '<th "border: 1px solid black;">SNo.</th>'
                    html_message +=                '<th "border: 1px solid black;">Sale Date</th>'
                    html_message +=                '<th style="border: 1px solid black;">Total distributer sale</th>'
                    html_message +=                '<th style="border: 1px solid black;">Total insert sale</th>'
                    # html_message +=                '<th>town_id</th>'
                    # html_message +=                '<th>reason</th>'
                    html_message +=            '</tr>'
                    for ils in trans_obj:
                        # print(ils,"=================")
                        row_no+=1
                        html_message +=            '<tr style="border: 1px solid red;">'
                        html_message +=                '<td style="border: 1px solid black;">'+str(row_no)+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['sale_date'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_distributer_sale'])+'</td>'
                        count_dist_sale = count_dist_sale + ils['total_distributer_sale']
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_insart_sale'])+'</td>'
                        insert_sale = insert_sale + ils['total_insart_sale']
                        
                        html_message +=            '</tr>'
                        row_no+=1
                    html_message +=            '<tr>'
                    html_message +=                '<td style="border: 1px solid black;">'+str(str(success_count))+'</td>'
                    html_message +=                '<td style="border: 1px solid black;">'+str("Total")+'</td>'
                    html_message +=                '<td style="border: 1px solid black;">'+str(count_dist_sale)+'</td>'
                    html_message +=                '<td style="border: 1px solid black;">'+str(insert_sale)+'</td>'
                    html_message +=            '</tr>'
                    html_message += '</table> </br></br>'



                    # log details==============
                    html_message += '<table style="width:100% ; border: 1px solid black;">'
                            
                    html_message +=            '<tr style="border: 1px solid black;">'
                    html_message +=                '<th style="border: 1px solid black;">SNo.</th>'
                    html_message +=                '<th style="border: 1px solid black;">Sale Date</th>'
                    html_message +=                '<th style="border: 1px solid black;">Sku id</th>'
                    html_message +=                '<th style="border: 1px solid black;">Wd id</th>'
                    html_message +=                '<th style="border: 1px solid black;">Town id</th>'
                    html_message +=                '<th style="border: 1px solid black;">Reason</th>'
                    # html_message +=                '<th style="border: 1px solid black;">Repeat Count</th>'
                    html_message +=            '</tr>'
                    for ld in log_detail:
                        
                        detail_row_no+=1
                        # print(ld['sales_date_time'].date(),"==========date=======")
                        html_message +=            '<tr style="border: 1px solid red;">'
                        html_message +=                '<td style="border: 1px solid black;">'+str(detail_row_no)+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['sales_date_time'].date())+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['sku_id'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['wd_id'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['town_id'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ld['reason'])+'</td>'
                        # html_message +=                '<td style="border: 1px solid black;">'+str(ld['repeat_count'])+'</td>'
                        html_message +=            '</tr>'
                    html_message += '</table> </br></br>'

                    
                    email_from = settings.EMAIL_HOST_USER
                    cc_email = cc_list
                    # ,'saurabha@triazinesoft.com','mayankg@triazinesoft.com'
                    recipient_list = success_email_list
                    # recipient_list = ["VIPULSHARMA-gpi@modi-ent.com","ankit.rakwal@esobene.com"]

                    reset_email = EmailMessage(
                                subject = subject,
                                body = html_message,
                                from_email = email_from,
                                to = recipient_list,
                                cc = cc_email,
                                reply_to = cc_email,
                                )
                    trans_obj=Integration_log_summary.objects.filter(sale_date=final_data[0].get('sale_date'),email_status = False).update(email_status = True)
                    reset_email.content_subtype = "html"
                    reset_email.send(fail_silently=True)

            return None
        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong'}
            return Response(context, status=status.HTTP_200_OK)

from master import tests
class SFA_API_valid_V2(generics.ListCreateAPIView):
    # http_method_names = ['get', ]
    def get_queryset(self):
        try:
            
            print('success_count',"========success_count=========")
            # print(aa = 66)
            # breakpoint()
            # url = 'https://uat.lakshya.rsr.cloware.com/secondarysync/secondary_sales/0'
            # hashers = {'Authorization':'Basic  U0VDT05EQVJZX0lOVEVHUkFUSU9OX1RFQU06U0VDT05EQVJZU3luYzEyMyM=','Auth-Key':'secondaryauth','Content-Type':'application/json'}
            # production=======
            url = "https://drop52.reports.cloware.com/secondarysync/secondary_sales/0"
            hashers = {'Authorization':'Basic  U0VDT05EQVJZX0lOVEVHUkFUSU9OX1RFQU06U0VDT05EQVJZU3luYzEyMyM=','Auth-Key':'secondaryauth','Content-Type':'application/json'}

            data=requests.get(url, headers=hashers)
            success_count=data.json()['data']['success_count']
            s=0
            all_data=[]
            errors_list=[]
            flag=0
            repet_flag = 0
            histry_dict={}
            error_detail_dict = {}
            len_val = 0
            town_cd =[]
            wd_list_4_week = []
            date_tuple_weekly = [8,15,22,1]
            log_list = 0
            success_list = 0

            for i in range(0,int(success_count),500):
                if s==0:
                    # url = 'https://uat.lakshya.rsr.cloware.com/secondarysync/secondary_sales/0'
                    url = "https://drop52.reports.cloware.com/secondarysync/secondary_sales/0"
                    data=requests.get(url, headers=hashers)
                    total_data=data.json()['data']['distributer_sale']
                    all_data.append(total_data)
                else:
                    # url = 'https://uat.lakshya.rsr.cloware.com/secondarysync/secondary_sales/'+str(i+1)
                    url = "https://drop52.reports.cloware.com/secondarysync/secondary_sales/"+str(i+1)
                    data=requests.get(url, headers=hashers)
                    total_data=data.json()['data']['distributer_sale']
                    all_data.append(total_data)
                s=s+1
            # all_data = tests.sfa
            final_data = list(chain(*all_data))
            json_object = json.dumps(all_data, indent = 4)
            # n = datetime.datetime.now().date()-datetime.timedelta(days=1)
            with open(settings.MEDIA_ROOT + 'sfa_api_sales/'+"sfa_sales_"+str(datetime.datetime.now().date()) + '.json','w')as outfile:
                outfile.write(json_object)
            if json_object[0]:
                response = HttpResponse(content_type='text/csv')
                df = pd.DataFrame.from_dict(list(final_data))
                df.to_csv(settings.MEDIA_ROOT + 'sfa_api_sales_csv/'+"sfa_sales_"+str(datetime.datetime.now().date()) + '.csv', index = False)
            # =========testcase===========================================
            # from master import sfa
            # success_count = sfa.aa['data']['success_count']
            # final_data = sfa.aa['data']['distributer_sale']
            # print(len(sfa.aa['data']['distributer_sale']),"==========",success_count)
            # breakpoint()
            a = 0
            # ===========================end test case====================
            today_date = datetime.datetime.now()
            if json_object[0]:
                response = HttpResponse(content_type='text/csv')
                df = pd.DataFrame.from_dict(list(final_data))
                df.to_csv(settings.MEDIA_ROOT + 'sfa_api_sales_csv/'+"sfa_sales_"+str(datetime.datetime.now().date()) + '.csv', index = False)
            
            status_data = Apistatus.objects.filter(running_date=today_date.date(),api="SFA_API").last()
            running_state =  status_data.status if status_data and status_data.status == True else False
            print(running_state,"========running_state",success_count,"=======bb",today_date.date())
            if int(success_count) > 0  and (running_state == False) :
                running_status_generate = Apistatus.objects.create(running_date=today_date.date(),status=True,api="SFA_API")
                trans_obj=Integration_log_summary.objects.filter(sale_date=final_data[0].get('sale_date')).last()
                temp_sale_data_objs = Temp_Total_sku_sales.objects.all()
                
                print(temp_sale_data_objs)
                temp_sale_data_objs.delete()
                Integration_log_summary.objects.all().delete()
                Integration_log_details.objects.all().delete()
                # breakpoint()
                # final_data = final_data[:100]
                less_30_day = today_date.date()-datetime.timedelta(days=8)
                for json_data in final_data:
                    # breakpoint()
                    a = a+1
                    print(a,"=====================================================================================//==")
                    sku_obj=SKU_Master_Product.objects.filter(sku_code__iexact=json_data['prodcode']).last()
                    wd_id = json_data['distrcode'].split("-")[0]
                    
                    town_trim_wd = str(json_data['distrcode']).replace(json_data['town_code'],'')
                    user_id = User.objects.filter(Q(user_id = wd_id)|Q(user_id = town_trim_wd)).last()
                    town_cd.append(json_data['town_code'])
                    wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(wd_id=wd_id)|Q(wd_id = town_trim_wd)),town_code__in=[json_data['town_code'],re.sub('\W+','',json_data['town_code']),str(json_data['town_code']),str(json_data['town_code'])[1:],'0'+str(json_data['town_code'])]).values_list('wd_town_id',flat=True)
                    sku_obj=SKU_Master_Product.objects.filter(sku_code__iexact=json_data['prodcode']).last()
                    wd_id = json_data['distrcode'].split("-")[0]

                    # wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(wd_id=wd_id)|Q(wd_id = town_trim_wd)),(Q(town_code=json_data['town_code'])|Q(town_code__iexact=json_data['town_code'])|Q(town_code__iexact=str(json_data['town_code'])[1:])|Q(town_code__iexact='0'+str(json_data['town_code'])))).values_list('wd_town_id',flat=True)
                    wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(wd_id=wd_id)|Q(wd_id = town_trim_wd)),town_code__in=[json_data['town_code'],re.sub('\W+','',json_data['town_code']),str(json_data['town_code']),str(json_data['town_code'])[1:],'0'+str(json_data['town_code'])]).values_list('wd_town_id',flat=True)
                    wdobj = WdSkuCatagory.objects.filter(sku_code__iexact=json_data['prodcode'],wd_town_id__in=wd_town_list, status = True).last()
                    sku = []
                    wd = []
                    town = []
                    unmap_sku = []
                    if not user_id:
                        wd.append(wd_id)
                    elif not wd_town_list:
                        town.append(json_data['town_code'])
                    elif not wdobj:
                        unmap_sku.append(json_data['prodcode'])
                    elif not sku_obj:
                        sku.append(json_data['prodcode'])
                    total_sale = (float(json_data.get('local_retail', 0))+float(json_data.get('local_dealer', 0))+float(json_data.get('local_MT', 0))+float(json_data.get('local_HA', 0))
                                +float(json_data.get('out_retail', 0))+float(json_data.get('out_dealer', 0))+float(json_data.get('out_MT', 0))+float(json_data.get('out_HA', 0))
                                +float(json_data.get('other_retail', 0))+float(json_data.get('other_dealer', 0))+float(json_data.get('other_MT', 0)))
                    if sku_obj and wdobj and user_id and (total_sale > 0):
                        success_list = success_list + 1
                        wd_list_4_week.append(user_id.user_id)
                        
                        wds = WDmaster.objects.filter(wd_ids = user_id.user_id).last()
                        # town_nm = Sales_Hierarchy_Master.objects.filter((Q(wd_id = user_id.user_id)|Q(wd_id = town_trim_wd)),(Q(town_code__iexact=json_data['town_code'])|Q(town_code__iexact = str(json_data['town_code'])[1:])|Q(town_code__iexact='0'+str(json_data['town_code'])))).last()
                        town_nm = wd_town_list =Sales_Hierarchy_Master.objects.filter((Q(wd_id=wd_id)|Q(wd_id = town_trim_wd)),town_code__in=[json_data['town_code'],re.sub('\W+','',json_data['town_code']),str(json_data['town_code']),str(json_data['town_code'])[1:],'0'+str(json_data['town_code'])]).last()
                        if town_nm:
                            town_name = town_nm.town
                        else:
                            town_name = " "

                        if wds:
                            wd_type_master = wds.wd_type
                            wd_id_name = wds.wd_name
                        else:
                            wd_id_name = "Name not Avail in wd_master"
                            
                        json_data['distrcode'] = str(user_id.user_id)+"-"+json_data['town_code']
                        json_data['town_code'] = town_nm.town_code
                        json_data['sku_short_name'] = sku_obj.sku_short_name
                        json_data['sku_code'] = json_data['prodcode']
                        json_data['town_name'] = town_name
                        json_data['statename'] = wds.wd_state

                        json_data['wd_name'] = wd_id_name
                        json_data['wd_type'] = wd_type_master
                        
                        json_data['sku_id'] = sku_obj.sku_id
                        json_data['town_id'] = wdobj.wd_town_id
                        json_data['wd_id'] = user_id.user_id
                        json_data['tranisition_source'] = json_data['dist_type']
                        json_data['transaction_type'] = 'DAILY'
                        json_data['created_by'] = "SFA/SFA_LITE_API"
                        json_data['brand_category'] = json_data.get('catcode')
                        json_data['created_date'] = datetime.datetime.now()
                        json_data['sales_date_time'] = json_data.get('sale_date')
                        json_data['local_sales_retail'] = json_data.get('local_retail', 0)
                        json_data['local_sales_dealer'] = json_data.get('local_dealer', 0)
                        json_data['local_sales_modern_trade'] = json_data.get('local_MT', 0)
                        json_data['local_sales_hawker'] = json_data.get('local_HA', 0)
                        json_data['outstation_sales_reatil'] = json_data.get('out_retail', 0)
                        json_data['outstation_sales_dealer'] = json_data.get('out_dealer', 0)
                        json_data['outstation_sales_modern_trade'] = json_data.get('out_MT', 0)
                        json_data['outstation_sales_hawker'] = json_data.get('out_HA', 0)
                        json_data['other_sales_retail'] = json_data.get('other_retail', 0)
                        json_data['other_sales_dealer'] = json_data.get('other_dealer', 0)
                        json_data['other_sales_modern_trade'] = json_data.get('other_MT', 0)
                        json_data['other_issues_other'] = json_data.get('other_issues_other', 0)
                        json_data['other_issues_damage'] = json_data.get('other_issued_damage', 0)
                        json_data['other_issues_return'] = json_data.get('other_issued_returns', 0)
                        # json_data['last_updated_date'] = datetime.datetime.now()
                        json_data['total_local_sales'] = float(json_data['local_sales_retail']) + float(json_data['local_sales_dealer']) +float(json_data['local_sales_modern_trade'])+float(json_data['local_sales_hawker'])
                        json_data['total_outstation_sales'] = float(json_data['outstation_sales_reatil']) +float(json_data['outstation_sales_dealer']) +float(json_data['outstation_sales_modern_trade']) +float(json_data['outstation_sales_hawker'])
                        json_data['total_other_sales'] = float(json_data['other_sales_retail']) +float(json_data['other_sales_dealer']) +float(json_data['other_sales_modern_trade'])
                        json_data['total_issue'] = float(json_data['other_issues_other']) +float(json_data['other_issues_damage']) +float(json_data['other_issues_return'])
                        json_data['grand_total'] = round(json_data['total_local_sales'] +json_data['total_outstation_sales'] +json_data['total_other_sales'],3)
                        if sku_obj:
                            json_data['company'] = sku_obj.company
                        else:
                            json_data['company'] =None
                        if wdobj:
                            json_data['unit_price'] = wdobj.last_price
                            json_data['cnf_id'] = wdobj.cnf_id
                        else:
                            json_data['unit_price'] = 0
                            json_data['cnf_id'] = None
                        branch = User.objects.filter(user_id = json_data.get('wd_id'), user_type = "WD").last()
                        if branch:
                            region = BranchMaster.objects.filter(branch_code = branch.locationcode).last()
                            json_data['region'] = region.region
                        else:
                            json_data['region'] = None

                        json_data['value'] = float(json_data['unit_price']) * float(json_data['grand_total'])
                        # sales_obj = sales_data.filter(wd_id = json_data['wd_id'],sku_id = json_data['sku_id'],transaction_type = 'DAILY',sales_date_time=json_data['sales_date_time'],town_id = json_data['town_id'],town_code = json_data['town_code']).last()
                        # sales_obj = SalesData.objects.filter(wd_id = json_data['wd_id'],sku_id = json_data['sku_id'],transaction_type = 'DAILY',sales_date_time=json_data['sales_date_time'],town_id = json_data['town_id'],town_code = json_data['town_code']).last()
                        # if not sales_obj:
                        #     SaleData_create(json_data).sale_create()
                        # else:
                        #     SaleData_update(sales_obj,json_data).sale_update()
                        
                        # this section connented due to duplicate sale will not comming on SFA API===================== 
                        temp_sale_data = Temp_Total_sku_sales.objects.filter(wd_id = json_data['wd_id'],sku_id = json_data['sku_id'],transaction_type = 'DAILY',sales_date_time=json_data['sales_date_time'],town_id = json_data['town_id'],town_code = json_data['town_code']).last()
                        if not temp_sale_data:
                            serializer_temp_sale_data = Temp_Total_sku_salesSerializer(data = json_data, many = False)
                            if serializer_temp_sale_data.is_valid():
                                serializer_temp_sale_data.save()
                            else:
                                print(serializer_temp_sale_data.errors)
                                pass
                            # print(total_sale,"===================>>createing===================sfa======",temp_sale_data)
                            # breakpoint()
                        else:
                            json_data['local_sales_retail'] = float(json_data['local_sales_retail']) + temp_sale_data.local_sales_retail
                            json_data['local_sales_dealer'] = float(json_data['local_sales_dealer']) + temp_sale_data.local_sales_dealer 
                            json_data['local_sales_modern_trade'] = float(json_data['local_sales_modern_trade']) + temp_sale_data.local_sales_modern_trade
                            json_data['local_sales_hawker'] = float(json_data['local_sales_hawker']) + temp_sale_data.local_sales_hawker
                            json_data['total_local_sales'] = float(json_data['local_sales_retail']) + json_data['local_sales_dealer'] +json_data['local_sales_modern_trade']+json_data['local_sales_hawker']

                            json_data['outstation_sales_reatil'] = float(json_data['outstation_sales_reatil']) + temp_sale_data.outstation_sales_reatil
                            json_data['outstation_sales_dealer'] = float(json_data['outstation_sales_dealer']) + temp_sale_data.outstation_sales_dealer
                            json_data['outstation_sales_modern_trade'] = float(json_data['outstation_sales_modern_trade']) + temp_sale_data.outstation_sales_modern_trade
                            json_data['outstation_sales_hawker'] = float(json_data['outstation_sales_hawker']) + temp_sale_data.outstation_sales_hawker
                            json_data['total_outstation_sales'] = float(json_data['outstation_sales_reatil']) + json_data['outstation_sales_dealer'] + json_data['outstation_sales_modern_trade'] + json_data['outstation_sales_hawker']

                            json_data['other_sales_retail'] = float(json_data['other_sales_retail']) + temp_sale_data.other_sales_retail
                            json_data['other_sales_dealer'] = float(json_data['other_sales_dealer']) + temp_sale_data.other_sales_dealer
                            json_data['other_sales_modern_trade'] = float(json_data['other_sales_modern_trade']) + temp_sale_data.other_sales_modern_trade
                            json_data['total_other_sales'] = float(json_data['other_sales_retail']) + json_data['other_sales_dealer'] + json_data['other_sales_modern_trade']

                            json_data['other_issues_damage'] = float(json_data['other_issues_damage']) + temp_sale_data.other_issues_damage
                            json_data['other_issues_return'] = float(json_data['other_issues_return']) + temp_sale_data.other_issues_return
                            json_data['other_issues_other'] = float(json_data['other_issues_other']) + temp_sale_data.other_issues_other
                            json_data['total_issue'] = json_data['other_issues_damage'] + json_data['other_issues_return']+json_data['other_issues_other']

                            json_data['grand_total'] = round(json_data['total_local_sales'] +json_data['total_outstation_sales'] +json_data['total_other_sales'],3)
                            json_data['value'] = float(json_data['unit_price']) * float(json_data['grand_total'])

                            if wdobj:
                                json_data['unit_price'] = wdobj.last_price
                                json_data['cnf_id'] = wdobj.cnf_id
                            else:
                                json_data['unit_price'] = 0
                                json_data['cnf_id'] = None
                            json_data.pop('created_by')
                            json_data.pop('sales_date_time')
                            json_data.pop('created_date')
                            json_data['last_updated'] = "SFA/SFA_LITE_API"
                            json_data['last_updated_date'] = datetime.datetime.now()
                            serializer_temp_sale_data = Temp_Total_sku_salesSerializer(temp_sale_data,data = json_data, partial = True)
                            if serializer_temp_sale_data.is_valid():
                                serializer_temp_sale_data.save()
                        
                        count_success_sale_date_wise(today_date, json_data.get('sale_date'))
                        
                    else:
                        count_invalid_sale_date_wise(today_date, json_data.get('sale_date'))
                        
                        log_list = log_list+1
                        # print("=======00=======")
                        # json_data['transaction_source'] = "SFA/SFA_LITE_API"
                        # log_data_serializer = Invalid_log_dataSerializers(data = json_data)
                        # if log_data_serializer.is_valid():
                        #     log_data_serializer.save()
                            
                            
                        # breakpoint()
                        len_val=len_val+1
                        errors_list.append(json_data)
                        # const_time = datetime.datetime(2009, 10, 5, 00, 00, 00)
                        # wdid_town = json_data['distrcode'].split("-")
                        error_detail_dict['sku_id'] = json_data['prodcode']
                        error_detail_dict['wd_id'] = json_data['dist_id']
                        error_detail_dict['town_id'] = json_data['town_code']

                        if total_sale == 0:
                            error_detail_dict['reason'] = str(wd_id) +" "+" Received '0' sales for SKU "+ str(json_data['prodcode'])
                        elif len(wd)>0:
                            error_detail_dict['reason'] = str(wd_id) +" "+" wd_id is not present in user_master database."
                            print("===========ss",str(wd_id) +" "+" wd_id is not present in user_master database.")
                            # error_detail_dict['repeat_count'] = len_val
                        elif len(town)>0:
                            error_detail_dict['reason'] = str(wd_id)+ " WD_id is not mapped with " + str(json_data['town_code'])+" town code."
                            print("----------------------rrrr",str(wd_id)+ " WD_id is not mapped with " + str(json_data['town_code'])+" town code.")
                            
                        elif len(unmap_sku) > 0:
                            error_detail_dict['reason'] = str(json_data['prodcode']) +" sku_code is not mapped with "+ str(wd_id) +" wd_id in SS_database:Sales_Hierarchy_Master"
                            print("===========ss",str(json_data['prodcode']) +" sku_code is not mapped with "+ str(wd_id) +" wd_id in SS_database:Sales_Hierarchy_Master")
                            
                        elif len(sku)>0:
                            error_detail_dict['reason'] = str(json_data['prodcode']) +" "+ " sku_code is not present in sku_master database: sku_master"
                            print("===============88",str(json_data['prodcode']) +" "+ " sku_code is not present in sku_master database: sku_master")
                            # error_detail_dict['repeat_count'] = len_val
                        else:
                            # error_detail_dict['repeat_count'] = len_val
                            error_detail_dict['reason'] = str(json_data['town_code']) +"This towm_code is not mapped with any WD"

                        error_detail_dict['tranisition_source'] = "SFA/SFA_LITE_API"
                        # comb_d_t = datetime.datetime.combine(parse(final_data[0].get('sale_date')), const_time.time())
                        comb_d_t = parse(json_data['sale_date'])
                        # print(comb_d_t)
                        error_detail_dict['sales_date_time'] = comb_d_t
                        log_detail = Integration_log_details.objects.filter(tranisition_source = "SFA/SFA_LITE_API",sales_date_time = comb_d_t, sku_id = json_data['prodcode'],town_id = json_data['town_code']).last()
                        if log_detail:
                            # error_detail_dict['repeat_count'] = int(log_detail.repeat_count) +1
                            log_serializer = Integration_log_detailsSerializers(log_detail, data = error_detail_dict, partial=True)
                            
                            if log_serializer.is_valid():
                                
                                log_serializer.save()
                        else:
                            # print(error_detail_dict,"======error_detail_dict======")
                            error_detail_dict['repeat_count'] = 1
                            error_detail_dict['created_date'] = today_date
                            log_serializer = Integration_log_detailsSerializers(data = error_detail_dict)
                            
                            if log_serializer.is_valid():
                                
                                log_serializer.save()
                            else:
                                pass
                                # return Response(log_serializer.errors)
                    
                # from_temp_sale = Temp_Total_sku_sales.objects.all().values('brand_category','sku_id','wd_id','town_id','sales_date_time',
                #                                                 'local_sales_retail','local_sales_dealer','local_sales_modern_trade','local_sales_hawker','total_local_sales',
                #                                                 'outstation_sales_reatil','outstation_sales_dealer','outstation_sales_modern_trade','outstation_sales_hawker','total_outstation_sales',
                #                                                 'other_sales_retail','other_sales_dealer','other_sales_modern_trade','total_other_sales',
                #                                                 'other_issues_damage','other_issues_return','other_issues_other','total_issue',
                #                                                 'grand_total','created_by','last_updated','transaction_source','created_date','last_updated_date',
                #                                                 'status','freeze_status','transaction_type','company','unit_price','region','cnf_id','value',
                #                                                 'wd_name','wd_type','sku_code','sku_short_name','town_name','town_code','distrcode',
                #                                             )

                # if from_temp_sale:
                # #     # n = len(from_temp_sale)//4
                #     n = 4
                # # #     # print([from_temp_sale[i:i + n] for i in range(0, len(from_temp_sale), n)])
                #     aaa = np.array_split(from_temp_sale, 4)
                #     print(len(aaa),"-------------->")
                #     print(len(aaa[0]),len(aaa[1]),len(aaa[2]),len(aaa[3]))
                #     from_temp_sale = [from_temp_sale[i:i + n] for i in range(0, len(from_temp_sale), n)]
                #     thread_access1 = SaleData_create_and_update_1(aaa[0])
                #     # thread_access1.create_and_update()
                #     thread_access2 = SaleData_create_and_update_2(aaa[1])
                #     # thread_access2.create_and_update()
                #     thread_access3 = SaleData_create_and_update_3(aaa[2])
                #     # thread_access3.create_and_update()
                #     thread_access4 = SaleData_create_and_update_4(aaa[3])
                #     # thread_access4.create_and_update()
                    
                    
                    # thread_access1.start()
                    # thread_access2.start()
                    # thread_access3.start()
                    # thread_access4.start()
                    # thread_access1.join()
                    # thread_access2.join()
                    # thread_access3.join()
                    # thread_access4.join()
                    # breakpoint()
                    print('======================================================================its over now============')
                    
                    # for fts in from_temp_sale:                       
                    #     retive_cre_update = SalesData.objects.filter(wd_id = fts['wd_id'],sku_id = fts['sku_id'],sales_date_time=fts['sales_date_time'],transaction_type = 'DAILY',town_id = fts['town_id'],town_code = fts['town_code']).last()
                    #     # retive_cre_update = sales_data.filter(wd_id = fts['wd_id'],sku_id = fts['sku_id'],sales_date_time=fts['sales_date_time'],transaction_type = 'DAILY',town_id = fts['town_id'],town_code = fts['town_code']).last()
                    #     if not retive_cre_update:
                    #         # SaleData_create(fts).sale_create()
                            
                            
                            
                    #         # commented for create sale thread=============
                            
                    #         sales_data_serializer = SalesDataSerializer(data = fts, many = False)
                    #         if sales_data_serializer.is_valid():
                    #             sales_data_serializer.save()
                    #             print(sales_data_serializer.data,"=========ccrr")
                    #         else:
                    #             print(sales_data_serializer.errors)
                    #     else:
                    #         # SaleData_update(retive_cre_update,fts).sale_update()
                            
                    #         # commented for update sale thread=============
                    #         sales_data_serializer = SalesDataSerializer(retive_cre_update,data = fts, partial = True)
                    #         if sales_data_serializer.is_valid():
                    #             sales_data_serializer.save()
                    #             print(sales_data_serializer.data,"=========upup")
                                
                    #         else:
                    #             print(sales_data_serializer.errors)

                from sqlalchemy import create_engine

                # ################################################
                print("-----------------Before db connection-----------------")
                # my_eng = create_engine('mysql+mysqlconnector://secsales:s_ecGpi1_23@10.250.1.191:3306/gpi_ss')
                # my_eng = create_engine('mysql+mysqlconnector://python1_db:sec&Gpi#123@49.50.69.231:3306/secondary_sales_gpi') # DEV ----------------------
                # print(my_eng)
                print("-----------------AFTER DB connection-----------------")

                past_seven_day_date = datetime.datetime.today() - datetime.timedelta(days=20)

                start_date = datetime.datetime.now()
                
                # flag mentain for sale would not be save through front-end.
                running_status_flag = running_status_generate
                running_status_flag.sfa_bulk_flag = True
                running_status_flag.save()
                
                data = SalesData.objects.filter(sales_date_time__gte=past_seven_day_date).values()

                st = datetime.datetime.now()
                # main table data
                df = pd.DataFrame.from_records(data)
                # main_df = df.loc[:, df.columns != 'id']

                # temp table data
                temp_data = Temp_Total_sku_sales.objects.all().values()

                temp_df = pd.DataFrame.from_records(temp_data)
                main_temp_df = temp_df.loc[:, temp_df.columns != 'id']

                main_dataframe = pd.concat((df, temp_df))

                ids = main_dataframe.loc[main_dataframe.duplicated(
                    subset=['wd_id', 'sku_id', 'sales_date_time', 'transaction_type', 'town_code'],
                    keep='last'), :]
                ii = ids['id']
                fi_li = ii.tolist()
                fi_li = list(map(int, ii.tolist()))
                cc = SalesData.objects.filter(id__in=fi_li).delete()
                d = main_temp_df.to_sql('transaction_salesdata', my_eng, if_exists='append', index=False)
                print(datetime.datetime.now()-start_date,"====================================duration")
                ##################################
                running_status_flag = running_status_generate
                running_status_flag.sfa_bulk_flag = False
                running_status_flag.save()

                    
                SalesData.objects.filter(wd_id__in = wd_list_4_week).update(status=True)
                trans_obj = Integration_log_summary.objects.filter(created_date__date=today_date.date(),tranisition_source = "SFA/SFA_LITE_API" ).values("total_distributer_sale","total_insart_sale","sale_date")
                log_detail = Integration_log_details.objects.filter(created_date__date = today_date.date(), tranisition_source = "SFA/SFA_LITE_API").values("sku_id","wd_id","town_id","sales_date_time","reason")
                # print(trans_obj,"======trans_obj==========mail=====")
                success_count = success_count if success_count else 0
                row_no = 0
                detail_row_no = 0
                # today = utils.now_date
                today = datetime.datetime.now()
                now_day=str(today.date())
                splited_today = now_day.split('-')[2]
                # week_details = week_sele_before_date('BRANCH USER')
                # if int(splited_today) in date_tuple_weekly and week_details[2]:
                #     print(wd_list_4_week,"==============wd_list_4_week==============")
                #     create_update_weekly(wd_list_4_week)
                count_dist_sale = 0
                insert_sale = 0
                if trans_obj:
                    print(trans_obj,"======================================================mail============")
                    subject = sfa_success_sub
                    html_message = '<table style="width:100% ; border: 1px solid black;">'

                    html_message +=            '<tr style="border: 1px solid black;">'
                    html_message +=                '<th "border: 1px solid black;">SNo.</th>'
                    html_message +=                '<th "border: 1px solid black;">Sale Date</th>'
                    html_message +=                '<th style="border: 1px solid black;">Total distributer sale</th>'
                    html_message +=                '<th style="border: 1px solid black;">Total insert sale</th>'
                    # html_message +=                '<th>town_id</th>'
                    # html_message +=                '<th>reason</th>'
                    html_message +=            '</tr>'
                    for ils in trans_obj:
                        print(ils,"============+=====")
                        row_no+=1
                        html_message +=            '<tr style="border: 1px solid red;">'
                        html_message +=                '<td style="border: 1px solid black;">'+str(row_no)+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['sale_date'])+'</td>'
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_distributer_sale'])+'</td>'
                        count_dist_sale = count_dist_sale + ils['total_distributer_sale']
                        html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_insart_sale'])+'</td>'
                        insert_sale = insert_sale + ils['total_insart_sale']

                        html_message +=            '</tr>'
                    html_message +=            '<tr>'
                    html_message +=                '<td style="border: 1px solid black;">'+str(str(success_count))+'</td>'
                    html_message +=                '<td style="border: 1px solid black;">'+str("Total")+'</td>'
                    html_message +=                '<td style="border: 1px solid black;">'+str(count_dist_sale)+'</td>'
                    html_message +=                '<td style="border: 1px solid black;">'+str(insert_sale)+'</td>'
                    html_message +=            '</tr>'


                    html_message += '</table> </br></br>'
            #
            #
            #
            #         # log details==============
            #         # html_message += '<table style="width:100% ; border: 1px solid black;">'
            #
            #         # html_message +=            '<tr style="border: 1px solid black;">'
            #         # html_message +=                '<th style="border: 1px solid black;">SNo.</th>'
            #         # html_message +=                '<th style="border: 1px solid black;">Sale Date</th>'
            #         # html_message +=                '<th style="border: 1px solid black;">Sku id</th>'
            #         # html_message +=                '<th style="border: 1px solid black;">Wd id</th>'
            #         # html_message +=                '<th style="border: 1px solid black;">Town id</th>'
            #         # html_message +=                '<th style="border: 1px solid black;">Reason</th>'
            #         # # html_message +=                '<th style="border: 1px solid black;">Repeat Count</th>'
            #         # html_message +=            '</tr>'
            #         # for ld in log_detail:
            #
            #         #     detail_row_no+=1
            #         #     # print(ld['sales_date_time'].date(),"==========date=======")
            #         #     html_message +=            '<tr style="border: 1px solid red;">'
            #         #     html_message +=                '<td style="border: 1px solid black;">'+str(detail_row_no)+'</td>'
            #         #     html_message +=                '<td style="border: 1px solid black;">'+str(ld['sales_date_time'].date())+'</td>'
            #         #     html_message +=                '<td style="border: 1px solid black;">'+str(ld['sku_id'])+'</td>'
            #         #     html_message +=                '<td style="border: 1px solid black;">'+str(ld['wd_id'])+'</td>'
            #         #     html_message +=                '<td style="border: 1px solid black;">'+str(ld['town_id'])+'</td>'
            #         #     html_message +=                '<td style="border: 1px solid black;">'+str(ld['reason'])+'</td>'
            #         #     # html_message +=                '<td style="border: 1px solid black;">'+str(ld['repeat_count'])+'</td>'
            #         #     html_message +=            '</tr>'
            #         # html_message += '</table> </br></br>'
            #
                    email_from = settings.EMAIL_HOST_USER
                    cc_email = cc_list
                    # ,'saurabha@triazinesoft.com','mayankg@triazinesoft.com'
                    recipient_list = success_email_list
                    # recipient_list = ["VIPULSHARMA-gpi@modi-ent.com","ankit.rakwal@esobene.com"]

                    reset_email = EmailMessage(
                                subject = subject,
                                body = html_message,
                                from_email = email_from,
                                to = recipient_list,
                                cc = cc_email,
                                reply_to = cc_email,
                                )
                    trans_obj=Integration_log_summary.objects.filter(sale_date=final_data[0].get('sale_date'),email_status = False).update(email_status = True)
                    Apistatus.objects.filter(running_date=today_date.date(),api="SFA_API").update(status=False)
                    reset_email.content_subtype = "html"
                    reset_email.send(fail_silently=True)
                    print("============mail success============")
                week_details = week_sele_before_date('BRANCH USER')
                if int(splited_today) in date_tuple_weekly and week_details[2]:
                    print(wd_list_4_week,"==============wd_list_4_week==============")
                    create_update_weekly(wd_list_4_week)
            else:
                if int(success_count)<=0:

                    emailsend_zoro_or_running("There is no sale in SFA/SFA lite API.")
                elif running_state==True:

                    emailsend_zoro_or_running("SFA API is still running.")

                return None

        except Exception as e:
            error = getattr(e, 'message', repr(e))
            logger.error(error)
            context = {'status': False,'message':'Something Went Wrong','error':error}
            # subject = "UAT | SFA/SFA_lite Issue"
            subject = sfa_failed_sub
            recipient_list = failed_list
            cc_email = cc_list
            # recipient_list = ["VIPULSHARMA-gpi@modi-ent.com","ankit.rakwal@esobene.com"]
            html_message = str(context)
            reset_email = EmailMessage(
                                subject = subject,
                                body = html_message,
                                from_email = settings.EMAIL_HOST_USER,
                                to = recipient_list,
                                cc = cc_email,
                                reply_to = cc_email,
                                )
            reset_email.content_subtype = "html"
            today_date = datetime.datetime.now()
            Apistatus.objects.filter(running_date=today_date.date(),api="SFA_API").update(status=False)
            reset_email.send(fail_silently=True)
            # self.get_queryset()
            return Response(context, status=status.HTTP_200_OK)



# class Surya_API_json(APIView):
#     def get(self,request):
        
#         dat = request.GET.get("date")
#         errors_list = []
#         error_obj_list = []
#         valid_in_ss = []
#         unavail_error = []
#         invalid_in_ss = []
#         transaction_history = {}
#         error_detail_dict = {}
#         flag=0
        
#         wd_count = 0
#         sku_count = 0
#         town_count = 0
#         json_file = open(settings.MEDIA_ROOT + 'surya_api_sales/'+"surya_sales_"+str(dat) + '.json')
#         surya_datas = json.load(json_file)
#         try:
#             if surya_datas:
#                 for json_data in surya_datas:
#                     wd = []
#                     sku = []
#                     town = []
#                     if json_data['sku_code'] and json_data['wd_code'] and json_data['town_id']:
#                         sku_obj=SKU_Master_Product.objects.filter(sku_code=json_data['sku_code']).last()
#                         json_data['wd_id'] = json_data['wd_code']
#                         # .split("-")[0]
#                         # user = User.objects.filter(user_id = json_data['wd_code'].split("-")[0])
#                         user = User.objects.filter(user_id = json_data['wd_code'])

#                         wd_town_list =Sales_Hierarchy_Master.objects.filter(town_code=json_data['town_id'],wd_id=json_data.get('wd_id')).values_list('wd_town_id',flat=True)
#                         print(type(json_data['town_id']),"===json_data['town_id']===",wd_town_list,"-----",json_data['town_id'])
#                         # breakpoint()
#                         # town_code=json_data['town_id'],
#                         # if wd_town_list:
#                         wdobj = WdSkuCatagory.objects.filter(sku_code=json_data['sku_code'],wd_town_id__in=wd_town_list).last()
#                         if not sku_obj:
#                             sku.append(json_data['sku_code'])
#                             # error_detail_dict['reason'] = str(json_data['sku_code']) +" sku_code is not present in sku_master database "
#                         if not user:
#                             wd.append(json_data['wd_code'])
#                             # error_detail_dict['reason'] = str(json_data['wd_code']) +" "+" wd_id is not present in user_master database."
#                         if not wdobj:
#                             town.append(json_data['town_id'])
#                             # error_detail_dict['reason'] = str(json_data['town_id']) +"This towm_code is not mapped with any WD"

#                         if sku_obj and wdobj and user:
#                             # print("=====avail=====")
#                             valid_in_ss.append(json_data)
#                         else:
#                             json_data['dist_id']=json_data['wd_code']
#                             json_data['prodcode']=json_data['sku_code']
#                             json_data['town_code']=str(json_data['town_id'])
#                             json_data['sale_date']=json_data['sale_date_time']
#                             # json_data[region]=json_data['branch_code']
#                             json_data['region']=json_data['region']
#                             # json_data['prodcode']=json_data['category_code']
#                             json_data['local_retail']=json_data['local_sales_reatil']
#                             json_data['local_dealer']=json_data['local_sales_dealer']
#                             json_data['local_MT']=json_data['local_sales_modern_trade']
#                             json_data['local_HA']=json_data['local_sales_hawker']
#                             json_data['out_retail']=json_data['outstation_sales_reatil']
#                             json_data['out_dealer']=json_data['outstation_sales_dealer']
#                             json_data['out_MT']=json_data['outstation_sales_modern_trade']
#                             json_data['out_HA']=json_data['outstation_sales_hawker']
#                             json_data['other_retail']=json_data['other_sales_reatil']
#                             json_data['other_dealer']=json_data['other_sales_dealer']
#                             json_data['other_MT']=json_data['other_sales_modern_trade']
#                             json_data['other_issued_damage']=json_data['other_issues_damage']
#                             json_data['other_issues_return']=json_data['other_issues_return']
#                             # json_data[]=json_data['other_issues_return']
#                             json_data['transaction_source'] = "SURYA"
#                             json_data['plan_type'] = "DAILY"
#                             json_data['distrcode'] = str(json_data['wd_code'])+"-"+str(json_data['town_code'])
#                             log_data_serializer = Invalid_log_dataSerializers(data = json_data)
#                             if log_data_serializer.is_valid():
#                                 log_data_serializer.save()
#                             #     return Response(log_data_serializer.data)
#                             # return Response(log_data_serializer.errors)


#                             if len(wd)>0:
#                                 wd_count = wd_count+1
#                                 error_detail_dict['repeat_count'] = wd_count
#                                 error_detail_dict['reason'] = str(json_data['wd_code']) +" "+" wd_id is not present in user_master database."
#                             elif len(sku)>0:
#                                 sku_count = sku_count+1
#                                 error_detail_dict['repeat_count'] = sku_count
#                                 error_detail_dict['reason'] = str(json_data['sku_code']) +" "+ " sku_code is not present in sku_master database: sku_master"
#                             elif len(town)>0:
#                                 town_count = town_count+1
#                                 error_detail_dict['repeat_count'] = town_count
#                                 error_detail_dict['reason'] = str(json_data['sku_code']) +" sku_code is not mapped with "+ str(json_data['wd_code']) +" wd_id in SS_database: Sales_Hierarchy_Master"
#                             else:
#                                 error_detail_dict['repeat_count'] = 0
#                                 error_detail_dict['reason'] = str(json_data['town_id']) +"This towm_code is not mapped with any WD"

#                             const_time = datetime.datetime(2009, 10, 5, 00, 00, 00)
#                             # wdid_town = json_data['distrcode'].split("-")
#                             error_detail_dict['sku_id'] = json_data['sku_code']
#                             error_detail_dict['wd_id'] = json_data['wd_code']
#                             error_detail_dict['town_id'] = json_data['town_id']
                    
#                             error_detail_dict['tranisition_source'] = "SURYA_API"
#                             comb_d_t = datetime.datetime.combine(parse(json_data['sale_date_time']), const_time.time())
#                             # print(comb_d_t)
#                             error_detail_dict['sales_date_time'] = comb_d_t
#                             log_detail = Integration_log_details.objects.filter(tranisition_source = "SURYA_API",sales_date_time = comb_d_t, sku_id = json_data['sku_code'],town_id = json_data['town_id']).last()
#                             # ,town_id = json_data['town_code']
#                             # print(error_detail_dict)
#                             if log_detail:
#                                 print("=======update===")
#                                 log_serializer = Integration_log_detailsSerializers(log_detail, data = error_detail_dict , partial = True)
                                
#                                 if log_serializer.is_valid():
                                    
#                                     log_serializer.save()
#                             else:
#                                 print("=====create====")
#                                 log_serializer = Integration_log_detailsSerializers(data = error_detail_dict)
                                
#                                 if log_serializer.is_valid():
                                    
#                                     log_serializer.save()
#                                 else:
#                                     return Response(log_serializer.errors)
#                 a = 0
#                 if valid_in_ss:
#                     # print(a,"====")
#                     for ss_store in valid_in_ss:
#                         a=a+1
#                         # print(a,"====")
#                         sku_obj=SKU_Master_Product.objects.filter(sku_code=ss_store['sku_code']).last()
#                         # print(sku_obj.category_code,"======sku_obj===",sku_obj)
#                         ss_store['wd_id'] = ss_store['wd_code']
#                         # .split("-")[0]
#                         wd_town_list =Sales_Hierarchy_Master.objects.filter(Q(wd_id=ss_store.get('wd_id'))|Q(town_code=ss_store['town_id'])).values_list('wd_town_id',flat=True)
#                         # town_code=ss_store['town_id'],
#                         wdobj = WdSkuCatagory.objects.filter(sku_code=ss_store['sku_code'],wd_town_id__in=wd_town_list).last()
#                         # print(ss_store['sku_code'],"======ss_store======",wdobj.wd_town_id)
#                         wdid = WDmaster.objects.filter(wd_ids = ss_store.get('wd_id')).last()
#                         wd_town_name =Sales_Hierarchy_Master.objects.filter(wd_id=ss_store.get('wd_id')).last()
#                         ss_store['town_code'] = ss_store['town_id']
#                         ss_store['sku_short_name'] = sku_obj.sku_short_name
#                         ss_store['sku_code'] = ss_store['sku_code']
#                         print(wdid.wd_name,"======================wdid=======================")
#                         ss_store['region'] = ss_store['region']
#                         ss_store['wd_name'] = wdid.wd_name
#                         ss_store['wd_type'] = wdid.wd_type
#                         if wd_town_name:
#                             ss_store['town_name'] = wd_town_name.town
#                         else:
#                             ss_store['town_name'] = None

#                         ss_store['sku_id'] = sku_obj.sku_id
#                         ss_store['transaction_source'] = "SURYA"
                        
#                         ss_store['brand_category'] = sku_obj.category_code
#                         ss_store['transaction_type'] = 'DAILY'
#                         ss_store['created_by'] = "SURYA"
#                         # ss_store['brand_category'] = ss_store.get('catcode')
#                         ss_store['created_date'] = datetime.datetime.now()
#                         ss_store['sales_date_time'] = ss_store.get('sale_date_time')
#                         ss_store['local_sales_retail'] = ss_store.get('local_sales_reatil') if ss_store.get('local_sales_reatil', 0) else 0
#                         ss_store['local_sales_dealer'] = 0
#                         ss_store['local_sales_modern_trade'] = 0
#                         ss_store['local_sales_hawker'] = 0
#                         ss_store['outstation_sales_reatil'] = 0
#                         ss_store['outstation_sales_dealer'] = 0
#                         ss_store['outstation_sales_modern_trade'] = 0
#                         ss_store['outstation_sales_hawker'] = 0
#                         ss_store['other_sales_reatil'] = 0
#                         ss_store['other_sales_dealer'] = 0
#                         ss_store['other_sales_modern_trade'] = 0
#                         ss_store['other_issues_other'] = 0
#                         ss_store['other_issues_damage'] = 0
#                         ss_store['other_issues_return'] = 0

#                         # print(type(ss_store['local_sales_dealer'])  ,"=====lllll=====")

#                         ss_store['total_local_sales'] = float(ss_store['local_sales_retail']) + float(ss_store['local_sales_dealer']) + float(ss_store['local_sales_modern_trade']) + float(ss_store['local_sales_hawker'])
#                         ss_store['total_outstation_sales'] = float(ss_store['outstation_sales_reatil']) + float(ss_store['outstation_sales_dealer']) + float(ss_store['outstation_sales_hawker']) + float(ss_store['outstation_sales_modern_trade'])
#                         ss_store['total_other_sales'] = float(ss_store['other_sales_reatil']) + float(ss_store['other_sales_dealer']) + float(ss_store['other_sales_modern_trade'])
#                         ss_store['grand_total'] = ss_store['total_local_sales'] + ss_store['total_outstation_sales'] + ss_store['total_other_sales']

#                         ss_store['total_issue'] = float(ss_store['other_issues_other']) + float(ss_store['other_issues_damage']) + float(ss_store['other_issues_return'])
                        
#                         # cat_val = unit_val_calc(ss_store['sku_code'], ss_store['wd_code'].split("-")[1], ss_store['grand_total'])
#                         # print(cat_val,"=====cat_val====")
#                         ss_store['company'] = sku_obj.company
#                         branch = User.objects.filter(user_id = ss_store['wd_code']).last()
#                         region = BranchMaster.objects.filter(branch_code = branch.locationcode).last()
#                         ss_store['region'] = region.region
#                         sale_data=""
#                         if wdobj:
#                             ss_store['unit_price'] = wdobj.last_price
#                             ss_store['cnf_id'] = wdobj.cnf_id
#                             ss_store['town_id'] =  wdobj.wd_town_id
#                             ss_store['value'] = float(wdobj.last_price) * float(ss_store['grand_total'])
                            
#                             sale_data = SalesData.objects.filter(wd_id = ss_store['wd_id'], sku_id = sku_obj.sku_id, sales_date_time=ss_store['sales_date_time'],town_id = wdobj.wd_town_id,town_code = ss_store['town_id']).last()
#                         # print(ss_store,"=======sale_data=====")
#                         # with translation.atomic():
#                         if sale_data :
#                             ss_store.pop('created_by')
#                             ss_store.pop('sales_date_time')
#                             ss_store.pop('created_date')
#                             ss_store['last_updated'] = "Surya_API"
#                             ss_store['last_updated_date'] = datetime.datetime.now()
#                             serializer = SalesDataSerializer(sale_data, data = ss_store,partial=True)
#                             if serializer.is_valid():
#                                 serializer.save()
#                                 flag=flag+1
#                             else:
#                                 errors_list.append(serializer.errors,ss_store['wd_code'],ss_store['sku_code'])
                                
#                         else:
#                             serializer = SalesDataSerializer(data = ss_store)
#                             if serializer.is_valid():
#                                 serializer.save()
#                                 flag=flag+1
#                             else:
#                                 errors_list.append(serializer.errors,)
#                                 # errors_list.append(ss_store['wd_code'])
#                                 # errors_list.append(ss_store['sku_code'])

#                     transaction_history['tranisition_source'] = "SURYA_API"
#                     transaction_history['sale_date'] = valid_in_ss[0]['sale_date_time']
#                     transaction_history['total_distributer_sale'] = len(surya_datas)
#                     transaction_history['total_insart_sale'] = len(valid_in_ss)
#                     transaction_history['created_by'] = "SURYA_API"
#                     transaction_history['created_date'] = datetime.datetime.now()
#                     print("hiiii=====")
#                     transact_his = Integration_log_summary.objects.filter(sale_date = transaction_history['sale_date'],tranisition_source = "SURYA_API").last()
#                     print("hiiii=2222====")

#                     # print(transaction_history,"======transaction_history==== ",transaction_history)
#                     if transact_his is not None:
#                         transaction_history.pop('created_date')
#                         transaction_history['last_updated_date'] = datetime.datetime.now()
#                         serializer_class = Integration_log_summarySerializers(transact_his,data = transaction_history,partial=True)
#                         if serializer_class.is_valid():
#                             # print("hiii__update")
#                             serializer_class.save()
#                             # print("======ajay=====")
#                         else:
#                             return Response(serializer_class.errors)
#                     else:
#                         serializer_class = Integration_log_summarySerializers(data = transaction_history)
#                         if serializer_class.is_valid():
#                             # print("hiii__create")
#                             serializer_class.save()
#                         else:
#                             return Response(serializer_class.errors)

#                 trans_obj=Integration_log_summary.objects.filter(sale_date=surya_datas[0]['sale_date_time'],email_status = False,tranisition_source = "SURYA_API").values("total_distributer_sale","total_insart_sale","sale_date")
#                 log_detail = Integration_log_details.objects.filter(sales_date_time__date = surya_datas[0]['sale_date_time'],tranisition_source = "SURYA_API").values("sku_id","wd_id","town_id","sales_date_time","reason","repeat_count")
#                 # print(trans_obj,"====",log_detail)
#                 row_no = 0
#                 if trans_obj:
#                     subject = '"SURYA_API"'
                    
#                     html_message = '<table style="width:100% ; border: 1px solid black;">'
                            
#                     html_message +=            '<tr style="border: 1px solid black;">'
#                     html_message +=                '<th "border: 1px solid black;">SNo.</th>'
#                     html_message +=                '<th "border: 1px solid black;">Sale Date</th>'
#                     html_message +=                '<th style="border: 1px solid black;">Total distributer sale</th>'
#                     html_message +=                '<th style="border: 1px solid black;">Total insert sale</th>'
#                     # html_message +=                '<th>town_id</th>'
#                     # html_message +=                '<th>reason</th>'
#                     html_message +=            '</tr>'
#                     for ils in trans_obj:
#                         # print(ils,"=================")
#                         html_message +=            '<tr style="border: 1px solid red;">'
#                         html_message +=                '<td style="border: 1px solid black;">'+str(row_no+1)+'</td>'
#                         html_message +=                '<td style="border: 1px solid black;">'+str(ils['sale_date'])+'</td>'
#                         html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_distributer_sale'])+'</td>'
#                         html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_insart_sale'])+'</td>'
                        
#                         html_message +=            '</tr>'
#                     html_message += '</table> </br></br>'



#                     # log details==============
#                     html_message += '<table style="width:100% ; border: 1px solid black;">'
                            
#                     html_message +=            '<tr style="border: 1px solid black;">'
#                     html_message +=                '<th style="border: 1px solid black;">SNo.</th>'
#                     html_message +=                '<th style="border: 1px solid black;">Sale Date</th>'
#                     html_message +=                '<th style="border: 1px solid black;">Sku id</th>'
#                     html_message +=                '<th style="border: 1px solid black;">Wd id</th>'
#                     html_message +=                '<th style="border: 1px solid black;">Town id</th>'
#                     html_message +=                '<th style="border: 1px solid black;">Reason</th>'
#                     # html_message +=                '<th style="border: 1px solid black;">Repeat Count</th>'
#                     html_message +=            '</tr>'
#                     for ld in log_detail:
                        
#                         row_no+=1
#                         # print(ld['sales_date_time'].date(),"==========date=======")
#                         html_message +=            '<tr style="border: 1px solid red;">'
#                         html_message +=                '<td style="border: 1px solid black;">'+str(row_no)+'</td>'
#                         html_message +=                '<td style="border: 1px solid black;">'+str(ld['sales_date_time'].date())+'</td>'
#                         html_message +=                '<td style="border: 1px solid black;">'+str(ld['sku_id'])+'</td>'
#                         html_message +=                '<td style="border: 1px solid black;">'+str(ld['wd_id'])+'</td>'
#                         html_message +=                '<td style="border: 1px solid black;">'+str(ld['town_id'])+'</td>'
#                         html_message +=                '<td style="border: 1px solid black;">'+str(ld['reason'])+'</td>'
#                         # html_message +=                '<td style="border: 1px solid black;">'+str(ld['repeat_count'])+'</td>'

#                         html_message +=            '</tr>'
#                     html_message += '</table> </br></br>'

                    
#                     email_from = settings.EMAIL_HOST_USER
#                     cc_email = ['rasmis@triazinesoft.com','saurabha@triazinesoft.com']
#                     # ,'saurabha@triazinesoft.com','mayankg@triazinesoft.com'
#                     # recipient_list = ["rasmis@triazinesoft.com"]
#                     recipient_list = ["VIPULSHARMA-gpi@modi-ent.com","ankit.rakwal@esobene.com"]

#                     reset_email = EmailMessage(
#                                 subject = subject,
#                                 body = html_message,
#                                 from_email = email_from,
#                                 to = recipient_list,
#                                 cc = cc_email,
#                                 reply_to = cc_email,
#                                 )
#                     trans_obj=Integration_log_summary.objects.filter(sale_date = surya_datas[0]['sale_date_time'],email_status = False,tranisition_source = "SURYA_API").update(email_status = True)
#                     reset_email.content_subtype = "html"
#                     reset_email.send(fail_silently=True)                
#                 return Response({"sku":sku,"wd":wd,"town":town})
#         except Exception as e:
#             error = getattr(e, 'message', repr(e))
#             logger.error(error)
#             context = {'status': False,'message':'Something Went Wrong'}
#             return Response(context, status=status.HTTP_200_OK)



        
        
        
class LogoutallUser(viewsets.ModelViewSet):
    
    def get_queryset(self):
        users=User.objects.all().update(is_logedin=False,token=None)
        
        
        
        
   

class Test_Sch(generics.ListCreateAPIView):
    serializers_calss = Integration_log_detailsSerializers
    def get_queryset(self):
        subject = "Test Scheduller"
        recipient_list = ['rasmis@triazinesoft.com']
        cc_email = ['rasmis@triazinesoft.com']
        html_message = test_sch
        reset_email = EmailMessage(
                            subject = subject,
                            body = html_message,
                            from_email = settings.EMAIL_HOST_USER,
                            to = recipient_list,
                            cc = cc_email,
                            reply_to = cc_email,
                            )
        reset_email.content_subtype = "html"
        reset_email.send(fail_silently=True)
        
        
        
class LogoutallUser(viewsets.ModelViewSet):
    
    def get_queryset(self):
        users=User.objects.all().update(is_logedin=False,token=None)
        
        
        
        
   
    
class Sfa_email(APIView):
    def get(self, request):
        trans_obj = Integration_log_summary.objects.filter(tranisition_source = "SFA/SFA_LITE_API" ).values("total_distributer_sale","total_insart_sale","sale_date")
        log_detail = Integration_log_details.objects.filter(tranisition_source = "SFA/SFA_LITE_API").values("sku_id","wd_id","town_id","sales_date_time","reason")
        # print(trans_obj,"======trans_obj==========mail=====")
        # success_count = success_count if success_count else 0
        row_no = 0
        detail_row_no = 0
        # today = utils.now_date
        today = datetime.datetime.now()
        now_day=str(today.date())
        splited_today = now_day.split('-')[2]
        # week_details = week_sele_before_date('BRANCH USER')
        # if int(splited_today) in date_tuple_weekly and week_details[2]:
        #     print(wd_list_4_week,"==============wd_list_4_week==============")
        #     create_update_weekly(wd_list_4_week)
        count_dist_sale = 0
        insert_sale = 0
        if trans_obj:
            print(trans_obj,"======================================================mail============")
            subject = sfa_success_sub
            html_message = '<table style="width:100% ; border: 1px solid black;">'
                    
            html_message +=            '<tr style="border: 1px solid black;">'
            html_message +=                '<th "border: 1px solid black;">SNo.</th>'
            html_message +=                '<th "border: 1px solid black;">Sale Date</th>'
            html_message +=                '<th style="border: 1px solid black;">Total distributer sale</th>'
            html_message +=                '<th style="border: 1px solid black;">Total insert sale</th>'
            # html_message +=                '<th>town_id</th>'
            # html_message +=                '<th>reason</th>'
            html_message +=            '</tr>'
            for ils in trans_obj:
                # print(ils,"=================")
                row_no+=1
                html_message +=            '<tr style="border: 1px solid red;">'
                html_message +=                '<td style="border: 1px solid black;">'+str(row_no)+'</td>'
                html_message +=                '<td style="border: 1px solid black;">'+str(ils['sale_date'])+'</td>'
                html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_distributer_sale'])+'</td>'
                count_dist_sale = count_dist_sale + ils['total_distributer_sale']
                html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_insart_sale'])+'</td>'
                insert_sale = insert_sale + ils['total_insart_sale']
                
                html_message +=            '</tr>'
            html_message +=            '<tr>'
            html_message +=                '<td style="border: 1px solid black;">'+str(str(" "))+'</td>'
            html_message +=                '<td style="border: 1px solid black;">'+str("Total")+'</td>'
            html_message +=                '<td style="border: 1px solid black;">'+str(count_dist_sale)+'</td>'
            html_message +=                '<td style="border: 1px solid black;">'+str(insert_sale)+'</td>'
            html_message +=            '</tr>'

            
            html_message += '</table> </br></br>'



            # log details==============
            # html_message += '<table style="width:100% ; border: 1px solid black;">'
                    
            # html_message +=            '<tr style="border: 1px solid black;">'
            # html_message +=                '<th style="border: 1px solid black;">SNo.</th>'
            # html_message +=                '<th style="border: 1px solid black;">Sale Date</th>'
            # html_message +=                '<th style="border: 1px solid black;">Sku id</th>'
            # html_message +=                '<th style="border: 1px solid black;">Wd id</th>'
            # html_message +=                '<th style="border: 1px solid black;">Town id</th>'
            # html_message +=                '<th style="border: 1px solid black;">Reason</th>'
            # # html_message +=                '<th style="border: 1px solid black;">Repeat Count</th>'
            # html_message +=            '</tr>'
            # for ld in log_detail:
                
            #     detail_row_no+=1
            #     # print(ld['sales_date_time'].date(),"==========date=======")
            #     html_message +=            '<tr style="border: 1px solid red;">'
            #     html_message +=                '<td style="border: 1px solid black;">'+str(detail_row_no)+'</td>'
            #     html_message +=                '<td style="border: 1px solid black;">'+str(ld['sales_date_time'].date())+'</td>'
            #     html_message +=                '<td style="border: 1px solid black;">'+str(ld['sku_id'])+'</td>'
            #     html_message +=                '<td style="border: 1px solid black;">'+str(ld['wd_id'])+'</td>'
            #     html_message +=                '<td style="border: 1px solid black;">'+str(ld['town_id'])+'</td>'
            #     html_message +=                '<td style="border: 1px solid black;">'+str(ld['reason'])+'</td>'
            #     # html_message +=                '<td style="border: 1px solid black;">'+str(ld['repeat_count'])+'</td>'
            #     html_message +=            '</tr>'
            # html_message += '</table> </br></br>'

            # subjects = Subject.objects.filter(use_for__icontains = "sfa_api").last()
            # emails = Email_users.objects.filter(is_active = True,subject = subjects.id)
#             subject = subjects.name
            # emails = Email_users.objects.filter(is_active = True)
            # pass_cc_email = emails.filter(email_cc = True).values_list('email',flat=True) if emails else []
            # pass_to_email = emails.filter(email_to = True).values_list('email',flat=True) if emails else []
            subject = sfa_success_sub
            email_from = settings.EMAIL_HOST_USER
            cc_email = cc_list
            # ,'saurabha@triazinesoft.com','mayankg@triazinesoft.com'
            recipient_list = success_email_list
            # recipient_list = ["VIPULSHARMA-gpi@modi-ent.com","ankit.rakwal@esobene.com"]

            reset_email = EmailMessage(
                        subject = subject,
                        body = html_message,
                        from_email = email_from,
                        to = recipient_list,
                        cc = cc_email,
                        reply_to = cc_email,
                        )
            today_date = datetime.datetime.now()
            trans_obj=Integration_log_summary.objects.filter(email_status = False).update(email_status = True)
            Apistatus.objects.filter(running_date=today_date.date(),api="SFA_API").update(status=False)
            reset_email.content_subtype = "html"
            reset_email.send(fail_silently=True)
        
        else:
            if not trans_obj:
                emailsend_zoro_or_running("There is no sale in SFA/SFA lite API.")

            return None



class Surya_API_email(APIView):
    def get(self, request):
        trans_obj=Integration_log_summary.objects.filter(email_status = False,tranisition_source = "SURYA_API").values("total_distributer_sale","total_insart_sale","sale_date")
        log_detail = Integration_log_details.objects.filter(tranisition_source = "SURYA_API").values("sku_id","wd_id","town_id","sales_date_time","reason","repeat_count")
        row_no = 0
        if trans_obj or log_detail:
            # subject = '"UAT | SURYA_API Integration Status Report."'
            subject = surya_success_sub
            
            html_message = '<table style="width:100% ; border: 1px solid black;">'
                    
            html_message +=            '<tr style="border: 1px solid black;">'
            html_message +=                '<th "border: 1px solid black;">SNo.</th>'
            html_message +=                '<th "border: 1px solid black;">Sale Date</th>'
            html_message +=                '<th style="border: 1px solid black;">Total distributer sale</th>'
            html_message +=                '<th style="border: 1px solid black;">Total insert sale</th>'
            # html_message +=                '<th>town_id</th>'
            # html_message +=                '<th>reason</th>'
            html_message +=            '</tr>'
            for ils in trans_obj:
                # print(ils,"=================")
                html_message +=            '<tr style="border: 1px solid red;">'
                html_message +=                '<td style="border: 1px solid black;">'+str(row_no+1)+'</td>'
                html_message +=                '<td style="border: 1px solid black;">'+str(ils['sale_date'])+'</td>'
                html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_distributer_sale'])+'</td>'
                html_message +=                '<td style="border: 1px solid black;">'+str(ils['total_insart_sale'])+'</td>'
                
                html_message +=            '</tr>'
            html_message += '</table> </br></br>'



            # log details==============
            html_message += '<table style="width:100% ; border: 1px solid black;">'
                    
            html_message +=            '<tr style="border: 1px solid black;">'
            html_message +=                '<th style="border: 1px solid black;">SNo.</th>'
            html_message +=                '<th style="border: 1px solid black;">Sale Date</th>'
            html_message +=                '<th style="border: 1px solid black;">Sku id</th>'
            html_message +=                '<th style="border: 1px solid black;">Wd id</th>'
            html_message +=                '<th style="border: 1px solid black;">Town id</th>'
            html_message +=                '<th style="border: 1px solid black;">Reason</th>'
            # html_message +=                '<th style="border: 1px solid black;">Repeat Count</th>'
            html_message +=            '</tr>'
            for ld in log_detail:
                
                row_no+=1
                # print(ld['sales_date_time'].date(),"==========date=======")
                html_message +=            '<tr style="border: 1px solid red;">'
                html_message +=                '<td style="border: 1px solid black;">'+str(row_no)+'</td>'
                html_message +=                '<td style="border: 1px solid black;">'+str(ld['sales_date_time'].date())+'</td>'
                html_message +=                '<td style="border: 1px solid black;">'+str(ld['sku_id'])+'</td>'
                html_message +=                '<td style="border: 1px solid black;">'+str(ld['wd_id'])+'</td>'
                html_message +=                '<td style="border: 1px solid black;">'+str(ld['town_id'])+'</td>'
                html_message +=                '<td style="border: 1px solid black;">'+str(ld['reason'])+'</td>'
                # html_message +=                '<td style="border: 1px solid black;">'+str(ld['repeat_count'])+'</td>'

                html_message +=            '</tr>'
            html_message += '</table> </br></br>'

            
            email_from = settings.EMAIL_HOST_USER
            cc_email = cc_list
            # # ,'saurabha@triazinesoft.com','mayankg@triazinesoft.com'
            recipient_list = success_email_list
            # recipient_list = ["VIPULSHARMA-gpi@modi-ent.com","ankit.rakwal@esobene.com"]

            reset_email = EmailMessage(
                        subject = subject,
                        body = html_message,
                        from_email = email_from,
                        to = recipient_list,
                        cc = cc_email,
                        reply_to = cc_email,
                        )
            # trans_obj=Integration_log_summary.objects.filter(sale_date = surya_datas[0]['sale_date_time'],email_status = False,tranisition_source = "SURYA_API").update(email_status = True)
            reset_email.content_subtype = "html"
            reset_email.send(fail_silently=True)      
            
            
        
        # today = utils.now_date
        today = datetime.datetime.now()
        now_day=str(today.date())
        splited_today = now_day.split('-')[2]
        week_details = week_sele_before_date('BRANCH USER')
        Apistatus.objects.filter(running_date=today.date(),api="SFA_API").update(status=False)
            
        return None
        
        
        
        

from rest_framework.parsers import FormParser, MultiPartParser
class SalesUpload(APIView):
    
    def post(self, request):
        file = request.data['file']
        csv_record = pd.read_csv(file)
        list_of_column_names = list(csv_record.columns)
        wrong_columns = 0
        column=['BRANCH','TOWN NAME','WD NAME','WD ID','CATEGORY','BRAND CODE','BRANDSHORT NAME','SECSALE DATE','LOCAL RETAILER','LOCAL DEALER',
                'LOCAL MODERN','LOCAL HAWKER','OUTSTATION RETAILER','OUTSTATION DEALER','OUTSTATION MODERN','OUTSTATION HAWKER','OTHER RETAILER','OTHER DEALER',
                'OTHER MODERN','TOTAL RETAIL','TOTAL DEALER','TOTAL MODERN TRADE SALES','TOTAL HAWKER','TOTAL LOCAL','TOTAL OUTSTATION SALES',
                'GRAND TOTAL','TRANSACTION SOURCE']
        for columns in list_of_column_names: 

            print(columns,"=======")
            if columns not in column:
                wrong_columns+=1
        if len(column) != len(list_of_column_names) or wrong_columns > 0:
            raise serializers.ValidationError({'error' : ('CSV File format is wrong.Please check and correct.')})
        df_2_json = csv_record.to_json(orient='records')
        df_2_json = json.loads(df_2_json)
        
        if not df_2_json:
            raise serializers.ValidationError({'error' : ('There is no data in this CSV.')})
        with_remark = []
        for row in df_2_json:
            wd_id = row['WD ID']
            categoey_code = row['CATEGORY']
            sku_code = row['BRAND CODE']
            branch = row['BRANCH']
            town_name = row['TOWN NAME']
            print(row['SECSALE DATE'],"=======", type(row['SECSALE DATE']))
            sale_date = parse(row['SECSALE DATE']).date()
            
            user_table = User.objects.filter(user_id = wd_id)
            wd_master = WDmaster.objects.filter(wd_ids = wd_id, wd_postal_code__icontains = town_name).last()
            hierarchy_master = Sales_Hierarchy_Master.objects.filter(wd_id = wd_id, town__icontains = town_name)
            town_list = hierarchy_master.values_list('wd_town_id', flat= True)
            mapping_table = WdSkuCatagory.objects.filter(wd_town_id__in = town_list,sku_code = sku_code).last()
            master_product = SKU_Master_Product.objects.filter(sku_code = sku_code).last()
            
            
            
            if not user_table and not wd_master:
                row['Remark'] = "This user not exist in User_master or WD_master."
            elif not town_list:
                row['Remark'] = "This user not exist in Hierarchy master."
            elif not mapping_table:
                row['Remark'] = "This sku_code not mapped with wd_id and town"
            elif not master_product:
                row['Remark'] = "This SKU not present in master_product."
            elif user_table and wd_master and town_list and mapping_table and master_product and row['SECSALE DATE']:
                sales_obj = {}
                sales_obj['local_sales_retail'] = round(row.get('LOCAL RETAILER',0),3)
                sales_obj['local_sales_dealer'] = round(row.get('LOCAL DEALER',0),3)
                sales_obj['local_sales_modern_trade'] = round(row.get('LOCAL MODERN',0),3)
                sales_obj['local_sales_hawker'] = round(row.get('LOCAL HAWKER',0),3)
                sales_obj['total_local_sales'] = row.get('LOCAL RETAILER',0) + row.get('LOCAL DEALER',0) + row.get('LOCAL MODERN',0) + row.get('LOCAL HAWKER',0)
                
                sales_obj['outstation_sales_reatil'] = round(row.get('OUTSTATION RETAILER',0),3)
                sales_obj['outstation_sales_dealer'] = round(row.get('OUTSTATION DEALER',0),3)
                sales_obj['outstation_sales_modern_trade'] = round(row.get('OUTSTATION MODERN',0),3)
                sales_obj['outstation_sales_hawker'] = round(row.get('OUTSTATION HAWKER',0),3)
                sales_obj['total_outstation_sales'] = row.get('OUTSTATION RETAILER',0) + row.get('OUTSTATION DEALER',0) + row.get('OUTSTATION MODERN',0) + row.get('OUTSTATION HAWKER',0)
                # row.get(,0)
                
                
                sales_obj['other_sales_retail'] = round(row.get('OTHER RETAILER',0),3)
                sales_obj['other_sales_dealer'] = round(row.get('OTHER DEALER',0),3)
                sales_obj['other_sales_modern_trade'] = round(row.get('OTHER MODERN',0),3)
                sales_obj['total_other_sales'] = row.get('OTHER RETAILER',0) + row.get('OTHER DEALER',0) + row.get('OTHER MODERN',0)
                
                sales_obj['other_issues_damage'] = 0
                sales_obj['other_issues_return'] = 0
                sales_obj['other_issues_other'] = 0
                sales_obj['total_issue'] = 0
                
                sales_obj['grand_total'] = sales_obj['total_local_sales'] + sales_obj['total_outstation_sales'] + sales_obj['total_other_sales']
                
                sales_obj['wd_id'] = wd_id
                sales_obj['wd_name'] = wd_master.wd_name
                sales_obj['sku_id'] = master_product.sku_id
                sales_obj['town_code'] = hierarchy_master.last().town_code
                sales_obj['brand_category'] = categoey_code
                sales_obj['wd_type'] = row.get('TRANSACTION SOURCE', None)
                
                sales_obj['transaction_type'] = "DAILY"
                sales_obj['sku_short_name'] = master_product.sku_short_name
                sales_obj['sales_date_time'] = sale_date
                sales_obj['town_id'] = hierarchy_master.last().wd_town_id
                sales_obj['distrcode'] = f"{wd_id}-{hierarchy_master.last().town_code}"
                sales_obj['sku_code'] = sku_code
                sales_obj['town_name'] = hierarchy_master.last().town
                sales_obj['statename'] = wd_master.wd_state if wd_master else None
                sales_obj['gpi_state'] = wd_master.wd_state if wd_master else None
                sales_obj['cnf_id'] = mapping_table.cnf_id if mapping_table else None
                sales_obj['region'] = row['BRANCH']
                sales_obj['unit_price'] = mapping_table.last_price if mapping_table else 0
                sales_obj['value'] = sales_obj['grand_total'] * sales_obj['unit_price']
                sales_obj['company'] = master_product.company if master_product else None
                
                
                
                transaction_sales_data = SalesData.objects.filter(wd_id = wd_id,town_id__in = town_list, town_code = hierarchy_master.last().town_code, sales_date_time = sale_date, 
                                                                  brand_category = categoey_code, sku_id = master_product.sku_id,sku_code = sku_code, transaction_type = 'DAILY').last()
                
                if not transaction_sales_data:
                    sales_obj['created_by'] = "7th march CSV"
                    # create===========
                    print("==============create")
                    sales_obj['created_date'] = datetime.datetime.now()
                    serializer_class = SalesDataSerializer(data = sales_obj, many = False)
                    if serializer_class.is_valid():
                        serializer_class.save()
                    
                    row['Remark'] = "Create Success."
                    
                else:
                    sales_obj['last_updated'] = "7th march CSV"
                    
                    sales_obj['last_updated_date'] = datetime.datetime.now()
                    serializer_class = SalesDataSerializer(transaction_sales_data,sales_obj, many = False)
                    print("==============update")
                    
                    if serializer_class.is_valid():
                        serializer_class.save()
                    # update============
                    row['Remark'] = "Update Success."
            with_remark.append(row)        
        return Response({'massage': 'Sales data has been saved successfully' ,'status': True,'with_remark':with_remark})
